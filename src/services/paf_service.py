"""Regras de negócio do PAF (Plano Anual de Férias) — um dos serviços
centrais do sistema, usado para planejar e controlar as férias de todos os
militares. As rotas em src/routes/ferias_pafs.py ficam só com a orquestração
HTTP; toda consulta, validação e gravação relacionada a férias mora aqui.
"""
from base64 import b64encode
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

import matplotlib.pyplot as plt
from flask import current_app
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, func
from sqlalchemy.orm import aliased, joinedload

from src import database
from src.authz import is_super, is_super_or_perm
from src.decorators.control import militar_esta_no_escopo, obms_permitidas_para_usuario
from src.models import (
    Meses,
    Militar,
    MilitarObmFuncao,
    Obm,
    Paf,
    PostoGrad,
    Quadro,
    User,
    now_manaus_naive,
)


# ---------------------------------------------------------------------------
# Vigência e datas
# ---------------------------------------------------------------------------

def paf_ano_vigente() -> int:
    return current_app.config.get('PAF_ANO_VIGENTE', datetime.now().year)


def parse_date(date_string):
    try:
        return datetime.strptime(date_string, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def first_day_next_month(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def validate_vacation_period(start_date, days):
    """As férias de um período não podem ultrapassar 31/12 do ano seguinte
    ao atual (regra de negócio do PAF)."""
    next_year = date.today().year + 1
    end_date = start_date + timedelta(days=days - 1)
    if end_date > date(next_year, 12, 31):
        raise ValueError("As férias não podem ultrapassar 31 de dezembro.")


NOMES_MESES = {
    "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4, "Maio": 5, "Junho": 6,
    "Julho": 7, "Agosto": 8, "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12,
}


def calcular_janela_edicao_paf(ano: int) -> dict:
    """Calcula, para o ano de referência da tabela da OBM, a partir de qual
    data um usuário comum pode marcar férias (não é permitido marcar no
    passado). SUPER USER não tem essa trava e pode editar qualquer data."""
    if is_super():
        return {
            "min_iso": f"{ano}-01-01",
            "min_year": 0,
            "min_month": 1,
            "bloqueio_mes_atual": False,
            "is_super": True,
        }

    min_global = first_day_next_month(datetime.now().date())
    min_year = min_global.year
    min_month = min_global.month

    if ano < min_year:
        min_iso = f"{ano}-12-31"
    elif ano == min_year:
        min_iso = min_global.isoformat()
    else:
        min_iso = f"{ano}-01-01"

    return {
        "min_iso": min_iso,
        "min_year": min_year,
        "min_month": min_month,
        "bloqueio_mes_atual": True,
        "is_super": False,
    }


# ---------------------------------------------------------------------------
# Acesso por OBM
# ---------------------------------------------------------------------------

def usuario_tem_acesso_obm(obm_id: int) -> bool:
    """SUPER USER acessa qualquer OBM; os demais, só as OBMs sob sua gestão."""
    if is_super():
        return True
    return obm_id in obms_permitidas_para_usuario(current_user)


# ---------------------------------------------------------------------------
# Listagem / DataTables (tela "Férias" do super usuário)
# ---------------------------------------------------------------------------

def montar_query_pafs_datatable(ano: int, search_value: Optional[str]):
    """Monta a query base (Militar + Paf do ano + Usuário que alterou) usada
    pelo DataTables server-side da tela de férias."""
    Usuario = aliased(User)
    query = (
        database.session.query(Militar, Paf, Usuario)
        .outerjoin(
            Paf,
            and_(Militar.id == Paf.militar_id, Paf.ano_referencia == ano)
        )
        .outerjoin(Usuario, Usuario.id == Paf.usuario_id)
    )

    if search_value:
        query = query.filter(
            (Militar.nome_completo.ilike(f'%{search_value}%')) |
            (Militar.matricula.ilike(f'%{search_value}%')) |
            (Militar.quadro.has(quadro=search_value))
        )

    return query


def serializar_linha_paf_datatable(militar: Militar, paf: Optional[Paf], usuario: Optional[User]) -> dict:
    return {
        "posto_grad": militar.posto_grad.sigla if militar.posto_grad else "",
        "nome_completo": militar.nome_completo,
        "matricula": militar.matricula,
        "quadro": militar.quadro.quadro if militar.quadro else "",

        "mes_usufruto": paf.mes_usufruto if paf else "",
        "qtd_dias_1": paf.qtd_dias_primeiro_periodo if paf else "",
        "inicio_1": str(paf.primeiro_periodo_ferias) if paf and paf.primeiro_periodo_ferias else "",
        "fim_1": str(paf.fim_primeiro_periodo) if paf and paf.fim_primeiro_periodo else "",
        "qtd_dias_2": paf.qtd_dias_segundo_periodo if paf else "",
        "inicio_2": str(paf.segundo_periodo_ferias) if paf and paf.segundo_periodo_ferias else "",
        "fim_2": str(paf.fim_segundo_periodo) if paf and paf.fim_segundo_periodo else "",
        "qtd_dias_3": paf.qtd_dias_terceiro_periodo if paf else "",
        "inicio_3": str(paf.terceiro_periodo_ferias) if paf and paf.terceiro_periodo_ferias else "",
        "fim_3": str(paf.fim_terceiro_periodo) if paf and paf.fim_terceiro_periodo else "",

        "alterado_por": (
            usuario.nome if (paf and usuario and getattr(usuario, "nome", None))
            else (usuario.email if (paf and usuario and getattr(usuario, "email", None)) else "")
        ),
        "alterado_em": (
            paf.data_alteracao.strftime("%d/%m/%Y %H:%M")
            if (paf and paf.data_alteracao) else ""
        ),

        "id": militar.id,
    }


# ---------------------------------------------------------------------------
# Militares sem PAF cadastrado
# ---------------------------------------------------------------------------

def listar_militares_sem_paf():
    """Militares com vínculo de OBM ativo que ainda não têm PAF registrado,
    ordenados por prioridade de OBM (Gabinete do Subcomandante-Geral primeiro)
    e depois pela hierarquia de posto/graduação."""
    subquery_pafs = database.session.query(Paf.militar_id).subquery()

    prioridade_obm = case(
        (Obm.sigla == 'GAB SUBCMT-GERAL', 1),
        else_=2
    )

    ordem_posto = case(
        (PostoGrad.sigla == 'CEL', 1),
        (PostoGrad.sigla == 'TC', 2),
        (PostoGrad.sigla == 'MAJ', 3),
        (PostoGrad.sigla == 'CAP', 4),
        (PostoGrad.sigla == '1 TEN', 5),
        (PostoGrad.sigla == '2 TEN', 6),
        (PostoGrad.sigla == 'AL OF', 7),
        (PostoGrad.sigla == 'ALUNO OFICIAL', 8),
        (PostoGrad.sigla == 'SUBTENENTE', 9),
        (PostoGrad.sigla == '1 SGT', 10),
        (PostoGrad.sigla == '2 SGT', 11),
        (PostoGrad.sigla == '3 SGT', 12),
        (PostoGrad.sigla == 'AL SGT', 13),
        (PostoGrad.sigla == 'CB', 14),
        (PostoGrad.sigla == 'SD', 15),
        (PostoGrad.sigla == 'AL SD', 16),
        else_=99
    )

    sub_militares = (
        database.session.query(
            Militar.id.label("militar_id"),
            Militar.nome_completo,
            PostoGrad.sigla.label("posto_grad"),
            Quadro.quadro.label("quadro"),
            Obm.sigla.label("obm"),
            ordem_posto.label("ordem"),
            func.row_number().over(
                partition_by=Militar.id,
                order_by=[prioridade_obm.asc(), MilitarObmFuncao.id.desc()]
            ).label("linha")
        )
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .join(Obm, Obm.id == MilitarObmFuncao.obm_id)
        .join(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .join(Quadro, Quadro.id == Militar.quadro_id)
        .filter(MilitarObmFuncao.data_fim.is_(None))  # OBMs ativas
        .filter(~Militar.id.in_(subquery_pafs))  # sem PAF
        .subquery()
    )

    return (
        database.session.query(
            sub_militares.c.nome_completo,
            sub_militares.c.posto_grad,
            sub_militares.c.quadro,
            sub_militares.c.obm
        )
        .filter(sub_militares.c.linha == 1)
        .order_by(sub_militares.c.ordem, sub_militares.c.obm)
        .all()
    )


# ---------------------------------------------------------------------------
# Tabela de PAFs por OBM (tela da chefia)
# ---------------------------------------------------------------------------

def listar_militares_pafs_para_tabela(obm_id: int, ano: int):
    """Militares ativos com vínculo ativo na OBM + PAF do ano (se existir),
    para a tabela de lançamento de férias da chefia."""
    return (
        database.session.query(Militar, Paf)
        .outerjoin(Paf, and_(
            Paf.militar_id == Militar.id,
            Paf.ano_referencia == ano
        ))
        .options(joinedload(Militar.obm_funcoes))
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
            Militar.inativo.is_(False)
        )
        .all()
    )


# ---------------------------------------------------------------------------
# Exportação Excel dos PAFs de uma OBM
# ---------------------------------------------------------------------------

COLUNAS_EXPORTACAO_PAF = [
    "OBM", "Ano", "Posto/Grad", "Nome", "Matrícula", "Quadro", "Mês Usufruto",
    "Qtd. Dias 1º", "Início 1º", "Fim 1º",
    "Qtd. Dias 2º", "Início 2º", "Fim 2º",
    "Qtd. Dias 3º", "Início 3º", "Fim 3º",
]


def resolver_ano_referencia_pafs(ano_querystring: Optional[int]) -> Optional[int]:
    """Se o ano não vier explícito na URL, usa o último ano com PAF cadastrado."""
    if ano_querystring:
        return ano_querystring
    return database.session.query(func.max(Paf.ano_referencia)).scalar()


def listar_militares_pafs_para_exportacao(obm_id: int, ano: int):
    """Militares com vínculo ativo na OBM + PAF do ano, ordenados por nome —
    para a planilha exportada (aqui não filtramos por `inativo`, diferente da
    tabela da chefia, para manter o histórico completo na exportação)."""
    return (
        database.session.query(Militar, Paf)
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None)
        )
        .outerjoin(Paf, and_(
            Paf.militar_id == Militar.id,
            Paf.ano_referencia == ano
        ))
        .order_by(Militar.nome_completo.asc())
        .all()
    )


def gerar_planilha_pafs_obm(obm: Obm, ano: int, militares_pafs) -> BytesIO:
    """Gera o .xlsx do PAF de uma OBM/ano a partir da lista (Militar, Paf)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{obm.sigla} {ano}"

    ws.append(COLUNAS_EXPORTACAO_PAF)
    for col_num, col_name in enumerate(COLUNAS_EXPORTACAO_PAF, 1):
        celula = ws.cell(row=1, column=col_num)
        celula.value = col_name
        celula.font = Font(bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS_EXPORTACAO_PAF))}1"

    def fmt(dt):
        return dt.strftime("%d/%m/%Y") if dt else ""

    for militar, paf in militares_pafs:
        ws.append([
            obm.sigla,
            ano,
            militar.posto_grad.sigla if militar.posto_grad else "",
            militar.nome_completo,
            militar.matricula,
            militar.quadro.quadro if militar.quadro else "",
            paf.mes_usufruto if paf else "",

            paf.qtd_dias_primeiro_periodo if paf else "",
            fmt(paf.primeiro_periodo_ferias) if paf else "",
            fmt(paf.fim_primeiro_periodo) if paf else "",

            paf.qtd_dias_segundo_periodo if paf else "",
            fmt(paf.segundo_periodo_ferias) if paf else "",
            fmt(paf.fim_segundo_periodo) if paf else "",

            paf.qtd_dias_terceiro_periodo if paf else "",
            fmt(paf.terceiro_periodo_ferias) if paf else "",
            fmt(paf.fim_terceiro_periodo) if paf else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Gráfico de férias por mês
# ---------------------------------------------------------------------------

OBMS_ADICIONAIS_GRAFICO_FERIAS = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25]


def contar_ferias_por_mes(obm_id: int) -> dict:
    """Conta, entre os 3 períodos de férias do PAF (sem filtro de ano — soma
    todos os PAFs do militar), quantos militares têm férias marcadas em cada
    mês. A OBM 16 agrega também um conjunto fixo de OBMs subordinadas."""
    militares = (
        database.session.query(Militar, Paf)
        .outerjoin(Paf, Paf.militar_id == Militar.id)
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            (MilitarObmFuncao.obm_id == obm_id) |
            (MilitarObmFuncao.obm_id.in_(OBMS_ADICIONAIS_GRAFICO_FERIAS) if obm_id == 16 else False)
        )
        .all()
    )

    ferias_por_mes = {mes.id: 0 for mes in Meses.query.all()}
    for militar, paf in militares:
        if not paf:
            continue
        for data_periodo in (
            paf.primeiro_periodo_ferias,
            paf.segundo_periodo_ferias,
            paf.terceiro_periodo_ferias,
        ):
            if data_periodo:
                ferias_por_mes[data_periodo.month] += 1

    return ferias_por_mes


def gerar_grafico_ferias_base64(ferias_por_mes: dict) -> str:
    """Renderiza o gráfico de barras de férias/mês e devolve como PNG em base64."""
    meses = Meses.query.all()
    labels = [mes.mes for mes in meses]
    values = [ferias_por_mes[mes.id] for mes in meses]

    fig = plt.figure(figsize=(10, 6))
    plt.bar(labels, values, color='skyblue')
    plt.xlabel('Mês')
    plt.ylabel('Número de Militares de Férias')
    plt.title('Militares de Férias por Mês')
    plt.xticks(rotation=25)

    buf = BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    image_base64 = b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    plt.close(fig)  # evita acumular figuras na memória do processo a cada chamada

    return image_base64


# ---------------------------------------------------------------------------
# Exceção de virada de ano
# ---------------------------------------------------------------------------

def alternar_excecao_virada_ano(militar_id: int, ano: int, excecao: bool, usuario_id: int) -> Paf:
    """Cria o PAF (se ainda não existir) e liga/desliga a flag de exceção de
    virada de ano. Não comita a sessão — quem chama decide o commit."""
    paf = Paf.query.filter_by(militar_id=militar_id, ano_referencia=ano).first()

    if not paf:
        paf = Paf(militar_id=militar_id, ano_referencia=ano, usuario_id=usuario_id)
        database.session.add(paf)

    paf.excecao_virada_ano = excecao
    paf.usuario_id = usuario_id

    try:
        paf.data_alteracao = now_manaus_naive()
    except Exception:
        paf.data_alteracao = datetime.now()

    return paf


# ---------------------------------------------------------------------------
# Atualização de PAF (lançamento dos períodos de férias)
# ---------------------------------------------------------------------------

@dataclass
class PeriodoFerias:
    numero: int
    qtd_dias: int
    inicio: Optional[date]
    fim: Optional[date]


def extrair_periodos_ferias(form) -> list:
    """Lê os 3 períodos de férias enviados pelo formulário de lançamento do PAF."""
    return [
        PeriodoFerias(
            numero=1,
            qtd_dias=int(form.get('qtd_dias_1') or 0),
            inicio=parse_date(form.get('inicio_1')),
            fim=parse_date(form.get('fim_1')),
        ),
        PeriodoFerias(
            numero=2,
            qtd_dias=int(form.get('qtd_dias_2') or 0),
            inicio=parse_date(form.get('inicio_2')),
            fim=parse_date(form.get('fim_2')),
        ),
        PeriodoFerias(
            numero=3,
            qtd_dias=int(form.get('qtd_dias_3') or 0),
            inicio=parse_date(form.get('inicio_3')),
            fim=parse_date(form.get('fim_3')),
        ),
    ]


def validar_periodos_ferias(periodos: list) -> None:
    """Levanta ValueError no primeiro período inválido (mesmo comportamento
    de curto-circuito da validação original: para na primeira falha)."""
    for periodo in periodos:
        if periodo.inicio:
            validate_vacation_period(periodo.inicio, periodo.qtd_dias)


def salvar_paf(militar_id: int, ano: int, mes_usufruto, periodos: list, usuario_id: int) -> Paf:
    """Cria (se preciso) e grava os 3 períodos de férias no PAF do militar/ano.
    Não comita a sessão — quem chama decide o commit."""
    paf = Paf.query.filter_by(militar_id=militar_id, ano_referencia=ano).first()
    if not paf:
        paf = Paf(militar_id=militar_id, ano_referencia=ano)
        database.session.add(paf)

    p1, p2, p3 = periodos

    paf.mes_usufruto = mes_usufruto
    paf.qtd_dias_primeiro_periodo = p1.qtd_dias
    paf.primeiro_periodo_ferias = p1.inicio
    paf.fim_primeiro_periodo = p1.fim

    paf.qtd_dias_segundo_periodo = p2.qtd_dias
    paf.segundo_periodo_ferias = p2.inicio
    paf.fim_segundo_periodo = p2.fim

    paf.qtd_dias_terceiro_periodo = p3.qtd_dias
    paf.terceiro_periodo_ferias = p3.inicio
    paf.fim_terceiro_periodo = p3.fim

    paf.usuario_id = usuario_id
    paf.data_alteracao = now_manaus_naive()

    return paf


def dentro_da_janela_de_edicao_mensal() -> bool:
    """PAFs só podem ser alterados entre os dias 10 e 20 de cada mês
    (fora essa janela, só quem tem bypass explícito)."""
    hoje = datetime.now().day
    return 10 <= hoje <= 20


def usuario_pode_atualizar_paf() -> bool:
    return is_super() or is_super_or_perm("FERIAS_UPDATE") or is_super_or_perm("FERIAS_SUPER")


def usuario_tem_escopo_sobre_militar(militar_id: int) -> bool:
    if is_super() or is_super_or_perm("FERIAS_SUPER"):
        return True
    permitidas = obms_permitidas_para_usuario(current_user)
    return militar_esta_no_escopo(militar_id, permitidas)
