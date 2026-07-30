from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def formatar_planilha(planilha):
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    preenchimento_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    fonte_cabecalho = Font(
        color="FFFFFF",
        bold=True,
    )

    for celula in planilha[1]:
        celula.fill = preenchimento_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for coluna in planilha.columns:
        maior_tamanho = 0
        letra_coluna = get_column_letter(coluna[0].column)

        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            maior_tamanho = max(maior_tamanho, len(valor))

            celula.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        planilha.column_dimensions[letra_coluna].width = min(
            maior_tamanho + 2,
            50,
        )

    # Aplica cor visual conforme o status.
    cabecalhos = {
        celula.value: celula.column
        for celula in planilha[1]
    }

    coluna_status = cabecalhos.get("status")

    if not coluna_status:
        return

    preenchimento_encontrado = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    preenchimento_nao_encontrado = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    preenchimento_atencao = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    for numero_linha in range(2, planilha.max_row + 1):
        celula_status = planilha.cell(
            row=numero_linha,
            column=coluna_status,
        )

        status = str(celula_status.value or "").upper()

        if status == "ENCONTRADO":
            celula_status.fill = preenchimento_encontrado

        elif status in {"NÃO ENCONTRADO", "NOME VAZIO"}:
            celula_status.fill = preenchimento_nao_encontrado

        else:
            celula_status.fill = preenchimento_atencao