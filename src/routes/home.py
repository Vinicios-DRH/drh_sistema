
from flask import current_app
from flask_login import login_required
from flask import abort, jsonify, current_app
from src.formatar_cpf import get_militar_por_user
from flask import render_template, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from src import app, database
from src.querys import obter_estatisticas_militares
from src.decorators.control import checar_ocupacao
from src.decorators.business_logic import processar_militares_a_disposicao, processar_militares_agregados, \
    processar_militares_le, processar_militares_lts
from src.services.militar_situacao_service import processar_inicio_situacoes_extras
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.routes.helpers import (
    _agora_manaus,
    _preview_licencas_especiais,
    _preview_lts,
    _filtros_painel_cbmc,
    _estatisticas_operacional_capital,
    _preview_le_operacional_capital,
    _preview_lts_operacional_capital,
    _militares_disponiveis_operacional_capital,
    _opcoes_filtros_cbmc,
)


@app.context_processor
def inject_militar_atual():
    mil = None
    try:
        if current_user.is_authenticated:
            mil = get_militar_por_user(current_user)  # usa teu helper
    except Exception:
        mil = None
    return {
        "militar_atual": mil,
        "militar_id_atual": (mil.id if mil else None),
    }


@app.route('/api/estatisticas', methods=['GET'])
def estatisticas():
    """Retorna as estatísticas dos militares em formato JSON."""
    estatisticas = obter_estatisticas_militares()
    return jsonify(estatisticas)


@app.route("/")
@login_required
def home():
    if current_user.funcao_user_id == 12:
        return redirect(url_for('home_atualizacao'))

    # DIRETOR e CHEFE: home sem dados sensíveis
    if current_user.funcao_user_id in (1, 2):
        return render_template(
            "home_gestao2.html",
            agora=_agora_manaus(),
            nome_usuario=getattr(current_user, "nome", "Usuário"),
            perfil_nome="Diretor" if current_user.funcao_user_id == 1 else "Chefe",
        )

    try:
        estatisticas = obter_estatisticas_militares()

        licencas_especiais_preview = _preview_licencas_especiais(limit=5)
        lts_preview = _preview_lts(limit=5)

        return render_template(
            "home.html",
            **estatisticas,
            licencas_especiais_preview=licencas_especiais_preview,
            lts_preview=lts_preview
        )
    except Exception as e:
        current_app.logger.exception("Erro ao carregar home")
        return jsonify({
            "error": "Erro ao carregar a página",
            "details": str(e)
        }), 500


OBMS_OPERACIONAIS_CAPITAL = [2, 5, 7, 15, 26, 35, 59, 60, 61, 62, 63, 65, 86]
LOCALIDADE_CAPITAL_ID = 1
USER_ID_CHEFE_CBMC = [13]


@app.route("/painel-cbmc-operacional")
@login_required
def painel_cbmc_operacional():
    if current_user.id != 13:
        abort(403)

    try:
        filtros = _filtros_painel_cbmc()
        estatisticas = _estatisticas_operacional_capital(filtros=filtros)
        licencas_especiais_preview = _preview_le_operacional_capital(
            limit=5, filtros=filtros)
        lts_preview = _preview_lts_operacional_capital(
            limit=5, filtros=filtros)
        militares_disponiveis = _militares_disponiveis_operacional_capital(
            filtros=filtros, limit=300)
        postos_grad, quadros = _opcoes_filtros_cbmc()

        return render_template(
            "painel_cbmc_operacional.html",
            **estatisticas,
            licencas_especiais_preview=licencas_especiais_preview,
            lts_preview=lts_preview,
            militares_disponiveis=militares_disponiveis,
            postos_grad=postos_grad,
            quadros=quadros,
            filtros=filtros,
            obms_operacionais=OBMS_OPERACIONAIS_CAPITAL,
        )
    except Exception as e:
        print(f"Erro no painel operacional do CBMC: {e}")
        return jsonify({"error": "Erro ao carregar o painel operacional do CBMC"}), 500


@app.route("/painel-cbmc-operacional/exportar-excel")
@login_required
def exportar_excel_painel_cbmc_operacional():
    if current_user.id != 13:
        abort(403)

    try:
        filtros = _filtros_painel_cbmc()
        militares_disponiveis = _militares_disponiveis_operacional_capital(
            filtros=filtros, limit=100000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Efetivo Operacional CBMC"

        headers = ["NOME", "MATRÍCULA", "POSTO/GRAD", "QUADRO", "OBM"]
        ws.append(headers)

        header_fill = PatternFill(
            fill_type="solid", start_color="B5121B", end_color="B5121B")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for item in militares_disponiveis:
            ws.append([
                item.nome_completo or "",
                item.matricula or "",
                item.posto_grad or "",
                item.quadro or "",
                item.obm_sigla or "",
            ])

        widths = {
            "A": 45,
            "B": 18,
            "C": 16,
            "D": 20,
            "E": 14,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="painel_cbmc_operacional.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Erro ao exportar Excel do painel CBMC: {e}")
        return jsonify({"error": "Erro ao exportar Excel"}), 500


@app.route("/admin/reprocessar-vigencias", methods=["POST"])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'DRH', 'SUPER USER', 'DIRETOR DRH')
def reprocessar_vigencias():
    try:
        processar_militares_agregados()
        processar_militares_a_disposicao()
        processar_militares_le()
        processar_militares_lts()
        processar_inicio_situacoes_extras()
        database.session.commit()
        return jsonify({"ok": True, "message": "Vigências reprocessadas com sucesso."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
