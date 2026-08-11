import unicodedata

from flask import current_app
import io
import math
from zoneinfo import ZoneInfo
from flask_wtf.csrf import validate_csrf
from flask_login import login_required
from flask import abort, json, request, jsonify, make_response, current_app
from random import shuffle
import os
import zipfile
import qrcode
import pytz
import pandas as pd
import base64
import matplotlib.pyplot as plt
import requests
import urllib
from src.decorators.utils_acumulo import b2_bucket_name, b2_client, b2_delete_all_versions, b2_upload_fileobj
from src.identificacao import buscar_pessoa_por_cpf, normaliza_matricula
from src.formatar_cpf import cadete_restantes, formatar_cpf, get_militar_por_user, is_cadete
from flask import render_template, redirect, url_for, request, flash, jsonify, session, send_file, make_response, \
    Response, stream_with_context
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import validate_csrf, generate_csrf
from werkzeug.utils import secure_filename
from src import app, database, bcrypt
from src.forms import (AtualizacaoCadastralForm, ControleConvocacaoForm, CriarSenhaForm, FichaAlunosForm, FormEsqueciSenha, FormFiltroMilitar, FormMilitarInativo, FormResetarSenhaPublica, FormViatura,
                       IdentificacaoForm, ImpactoForm, FormLogin, FormMilitar, FormCriarUsuario, FormMotoristas, FormFiltroMotorista, LtsAlunoForm, RecompensaAlunoForm,
                       RestricaoAlunoForm, SancaoAlunoForm, TabelaVencimentoForm, InativarAlunoForm, MatriculaConfirmForm)
from src.models import (ControleConvocacao, Convocacao, DocumentoMilitar, EfetivoDiarioOBM, HistoricoEfetivoDiario, ImportacaoMilitarHistorico, LogAcesso, LtsAlunos, Militar, MilitaresInativos, NomeConvocado, PostoGrad, Quadro, Obm, Localidade, Funcao, RecompensaAluno, RestricaoAluno, SancaoAluno, SegundoVinculo, SituacaoConvocacao, User, FuncaoUser, PublicacaoBg,
                        EstadoCivil, Especialidade, Destino, Motivo, Modalidade, Punicao, Comportamento, MilitarObmFuncao,
                        FuncaoGratificada,
                        MilitaresAgregados, MilitaresADisposicao, LicencaEspecial, LicencaParaTratamentoDeSaude, Paf,
                        Meses, Motoristas, Categoria, TabelaVencimento, ValorDetalhadoPostoGrad, FichaAlunos, AlunoInativo, Viaturas, ViaturaMilitar, MilitarGraduacao, MilitarContatoEmergencia, MilitarConjuge, LogExportacaoExcel, Curso, MilitarCurso, AuditoriaAtualizacaoCadastral, now_manaus_naive)
from src.querys import dados_para_mapa, obter_estatisticas_militares, login_usuario
from src.decorators.control import checar_ocupacao, militar_esta_no_escopo, obms_permitidas_para_usuario, sync_user_admin_obms_from_militar
from src.decorators.business_logic import processar_militares_a_disposicao, processar_militares_agregados, \
    processar_militares_le, processar_militares_lts
from datetime import datetime, date, timedelta
from io import BytesIO
from sqlalchemy.orm import joinedload, selectinload, load_only, aliased
from sqlalchemy import case, distinct, extract, func, not_, or_, cast, String, and_
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP, getcontext
from docx import Document
from urllib.parse import urlencode
from collections import defaultdict, Counter
from docx.shared import Pt
from docx.oxml.ns import qn
from src.decorators.formatar_datas import formatar_data_extenso, formatar_data_sem_zero
from src.decorators.email_utils import send_reset_password_email, verify_password_reset_token
import re
import plotly.graph_objs as go
import plotly.io as pio
from collections import defaultdict
from src.utils.sa_serialize import sa_to_dict
from sqlalchemy.inspection import inspect as sa_inspect
from src.security.perms import has_perm
from src.authz import is_super_or_perm, can_ferias_bypass_janela, is_super, require_perm
from src.utils.cadastro_status import cadastro_esta_completo
from src.utils.painel import (
    _obter_obm_principal,
    listar_situacoes_atualizacao,
    obter_resumo_atualizacao_cadastral,
    obter_militares_atualizacao_cadastral,
    serializar_militar_atualizacao,
    listar_obms_atualizacao,
    listar_postos_grad_atualizacao,
    obter_detalhes_militar_atualizacao,)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.services.importar_militares import (
    ler_planilha,
    colunas_reconhecidas,
    colunas_nao_reconhecidas,
    analisar_importacao,
    importar_dataframe,
    salvar_historico_importacao,
)
from src.services.militar_situacao_service import (
    parse_date_flex,
    sincronizar_blocos_funcionais,
)
import time
from src.utils.utils import registrar_log_download
from dateutil.relativedelta import relativedelta


# --- CACHE DO PAINEL ---
CACHE_PAINEL = {
    "dados": None,
    "ultima_atualizacao": 0
}
TEMPO_CACHE_SEGUNDOS = 300  # 5 minutos


@app.route("/painel-efetivo/limpar-cache")
@login_required  # Opcional: apenas quem tá logado pode limpar
def limpar_cache_painel():
    try:
        CACHE_PAINEL["dados"] = None
        CACHE_PAINEL["ultima_atualizacao"] = 0
        flash("Cache do painel limpo com sucesso! Os dados serão atualizados na próxima requisição.", "success")
        return redirect(url_for("painel_efetivo_publico"))
    except Exception as e:
        print(f"Erro ao limpar cache: {e}")
        return "Erro ao limpar cache", 500


# @app.route("/painel-efetivo")
# def painel_efetivo_publico():
#     try:
#         estatisticas = obter_estatisticas_militares()

#         q = (request.args.get("q") or "").strip()
#         status = (request.args.get("status") or "").strip()
#         obm_id = request.args.get("obm_id", type=int)
#         posto_grad_id = request.args.get("posto_grad_id", type=int)
#         modalidade_id = request.args.get("modalidade_id", type=int)
#         page = request.args.get("page", default=1, type=int)
#         per_page = 50

#         resumo_atualizacao = obter_resumo_atualizacao_cadastral(
#             obm_id=obm_id,
#             posto_grad_id=posto_grad_id,
#             modalidade_id=modalidade_id,
#         )

#         militares, total_filtrado = obter_militares_atualizacao_cadastral(
#             q=q,
#             status=status,
#             obm_id=obm_id,
#             posto_grad_id=posto_grad_id,
#             modalidade_id=modalidade_id,
#             page=page,
#             per_page=per_page,
#         )

#         obms = listar_obms_atualizacao()
#         postos_grad = listar_postos_grad_atualizacao()
#         situacoes = listar_situacoes_atualizacao()
#         atualizado_em = datetime.now(
#             ZoneInfo("America/Manaus")
#         ).strftime("%d/%m/%Y %H:%M:%S")

#         return render_template(
#             "painel_efetivo_publico.html",
#             **estatisticas,
#             **resumo_atualizacao,
#             militares=militares,
#             obms=obms,
#             postos_grad=postos_grad,
#             situacoes=situacoes,
#             q=q,
#             status=status,
#             obm_id=obm_id,
#             posto_grad_id=posto_grad_id,
#             modalidade_id=modalidade_id,
#             atualizado_em=atualizado_em,
#             page=page,
#             per_page=per_page,
#             total_filtrado=total_filtrado,
#             total_paginas=(total_filtrado + per_page - 1) // per_page,
#         )
#     except Exception as e:
#         print(f"Erro ao carregar painel público: {e}")
#         return jsonify({"error": "Erro ao carregar o painel"}), 500


@app.route("/painel-efetivo/api/militar/<int:militar_id>")
def painel_efetivo_publico_militar_detalhe(militar_id):
    try:
        dados = obter_detalhes_militar_atualizacao(militar_id)
        if not dados:
            return jsonify({"ok": False, "error": "Militar não encontrado"}), 404

        return jsonify({
            "ok": True,
            "militar": dados
        })
    except Exception as e:
        print(f"Erro ao carregar detalhes do militar: {e}")
        return jsonify({"ok": False, "error": "Erro ao carregar detalhes"}), 500


@app.route("/painel-efetivo/exportar-excel")
def painel_efetivo_publico_exportar_excel():
    try:
        q = (request.args.get("q") or "").strip()
        status = (request.args.get("status") or "").strip()
        obm_id = request.args.get("obm_id", type=int)
        posto_grad_id = request.args.get("posto_grad_id", type=int)
        modalidade_id = request.args.get("modalidade_id", type=int)

        militares, _ = obter_militares_atualizacao_cadastral(
            q=q,
            status=status,
            obm_id=obm_id,
            posto_grad_id=posto_grad_id,
            modalidade_id=modalidade_id,
            page=1,
            per_page=100000
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Atualizacao Cadastral"

        headers = [
            "Nome",
            "Posto/Grad",
            "OBM",
            "Situação",
            "Status",
            "Última auditoria",
            "Ação auditoria",
            "Observação auditoria",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="0B4A7D")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9E3EF"),
            right=Side(style="thin", color="D9E3EF"),
            top=Side(style="thin", color="D9E3EF"),
            bottom=Side(style="thin", color="D9E3EF"),
        )

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for militar in militares:
            auditoria = None
            if getattr(militar, "auditorias_atualizacao", None):
                auditoria = sorted(
                    militar.auditorias_atualizacao,
                    key=lambda a: a.criado_em or datetime.min,
                    reverse=True
                )[0]

            status_label = "Atualizado" if bool(
                militar.cadastro_atualizado) else "Pendente"

            ws.append([
                militar.nome_completo or "-",
                militar.posto_grad.sigla if militar.posto_grad else "-",
                _obter_obm_principal(militar),
                militar.situacao.condicao if militar.situacao else "-",
                status_label,
                auditoria.criado_em.strftime(
                    "%d/%m/%Y %H:%M") if auditoria and auditoria.criado_em else "-",
                auditoria.acao if auditoria and auditoria.acao else "-",
                auditoria.observacao if auditoria and auditoria.observacao else "-",
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        widths = {
            "A": 40,
            "B": 15,
            "C": 18,
            "D": 20,
            "E": 14,
            "F": 22,
            "G": 22,
            "H": 50,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{ws.max_row}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        data_exportacao = datetime.now(
            ZoneInfo("America/Manaus")).strftime("%Y-%m-%d_%H-%M-%S")

        registrar_log_download(
            nome_relatorio="Painel Público - Atualização Cadastral",
            colunas_lista=headers,
            filtros_dict={"status": "Atualizado" if status == "atualizado" else "Pendente" if status == "pendente" else "Todos",
                          "obm": Obm.query.get(obm_id).sigla if obm_id else "Todas" if obm_id is not None else "N/A",
                          "posto_grad": PostoGrad.query.get(posto_grad_id).sigla if posto_grad_id else "Todos" if posto_grad_id is not None else "N/A",
                          "situacao": Modalidade.query.get(modalidade_id).descricao if modalidade_id else "Todas" if modalidade_id is not None else "N/A",
                          "q": q or "N/A"},
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=f"militares_atualizacao_cadastral_{data_exportacao}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Erro ao exportar Excel do painel público: {e}")
        flash("Erro ao exportar o Excel.", "danger")
        return redirect(url_for("painel_efetivo_publico"))
