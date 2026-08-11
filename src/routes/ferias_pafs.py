
from flask_login import login_required
from flask import abort, request, jsonify, make_response
import base64
import matplotlib.pyplot as plt
from flask import render_template, request, jsonify, make_response, \
    Response
from flask_login import login_required, current_user
from src import app, database
from src.models import (Militar, PostoGrad, Quadro, Obm, User, MilitarObmFuncao,
                        Paf,
                        Meses, now_manaus_naive)
from src.decorators.control import checar_ocupacao, militar_esta_no_escopo, obms_permitidas_para_usuario
from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import joinedload, selectinload, aliased
from sqlalchemy import case, func, and_
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from src.utils.sa_serialize import sa_to_dict
from sqlalchemy.inspection import inspect as sa_inspect
from src.authz import is_super_or_perm, can_ferias_bypass_janela, is_super
from openpyxl import Workbook
from openpyxl.styles import Font
from src.utils.utils import registrar_log_download

from src.routes.helpers import (
    parse_date,
    validate_vacation_period,
    paf_ano_vigente,
    first_day_next_month,
)


@app.route('/ferias_dados', methods=['GET', 'POST'])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER', 'DIRETOR DRH', 'DRH')
def ferias_dados():
    draw = request.form.get('draw', type=int)
    start = request.form.get('start', type=int)
    length = request.form.get('length', type=int)
    search_value = request.form.get('search[value]', type=str)

    # ✅ ano vem do frontend; fallback 2026
    ano = request.form.get('ano', type=int) or 2026

    Usuario = aliased(User)
    query = (
        database.session.query(Militar, Paf, Usuario)
        .outerjoin(
            Paf,
            and_(
                Militar.id == Paf.militar_id,
                Paf.ano_referencia == ano
            )
        )
        .outerjoin(Usuario, Usuario.id == Paf.usuario_id)
    )

    if search_value:
        query = query.filter(
            (Militar.nome_completo.ilike(f'%{search_value}%')) |
            (Militar.matricula.ilike(f'%{search_value}%')) |
            (Militar.quadro.has(quadro=search_value))
        )

    total_records = query.count()

    militares_pafs = query.offset(start).limit(length).all()

    data = []
    for militar, paf, usuario in militares_pafs:
        data.append({
            "posto_grad": militar.posto_grad.sigla if militar.posto_grad else "",
            "nome_completo": militar.nome_completo,
            "matricula": militar.matricula,
            "quadro": militar.quadro.quadro if militar.quadro else "",

            "mes_usufruto": paf.mes_usufruto if paf else "",
            "qtd_dias_1": paf.qtd_dias_primeiro_periodo if paf else "",
            "inicio_1": str(paf.primeiro_periodo_ferias) if paf and paf.primeiro_periodo_ferias else "",
            "fim_1": str(paf.fim_primeiro_periodo) if paf and paf.fim_primeiro_periodo else "",
            "qtd_dias_2": paf.qtd_dias_segundo_periodo if paf else "",
            "inicio_2": str(paf.segundo_periodo_ferias) if paf and paf.segundo_periodo_ferias else "",
            "fim_2": str(paf.fim_segundo_periodo) if paf and paf.fim_segundo_periodo else "",
            "qtd_dias_3": paf.qtd_dias_terceiro_periodo if paf else "",
            "inicio_3": str(paf.terceiro_periodo_ferias) if paf and paf.terceiro_periodo_ferias else "",
            "fim_3": str(paf.fim_terceiro_periodo) if paf and paf.fim_terceiro_periodo else "",
            "alterado_por": (usuario.nome if (paf and usuario and getattr(usuario, "nome", None))
                             else (usuario.email if (paf and usuario and getattr(usuario, "email", None)) else "")),
            "alterado_em": (paf.data_alteracao.strftime("%d/%m/%Y %H:%M")
                            if (paf and paf.data_alteracao) else ""),

            "id": militar.id
        })

    return jsonify({
        "draw": draw,
        "recordsTotal": total_records,
        "recordsFiltered": total_records,
        "data": data
    })


@app.route('/ferias', methods=['GET'])
@login_required
@checar_ocupacao('SUPER USER', 'DRH')
def exibir_ferias():
    if not is_super_or_perm("NAV_FERIAS_SUPER"):
        abort(403)
    ano_vigente = 2026
    # se quiser depois, a gente busca do banco automaticamente
    anos_disponiveis = [2025, 2026]
    return render_template(
        'ferias.html',
        ano_atual=datetime.now().year,
        ano_vigente=ano_vigente,
        anos_disponiveis=anos_disponiveis
    )


@app.route('/pafs/nao_preenchidos')
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER')
def pafs_nao_preenchidos():
    subquery_pafs = database.session.query(Paf.militar_id).subquery()

    prioridade_obm = case(
        (Obm.sigla == 'GAB SUBCMT-GERAL', 1),
        else_=2
    )

    ordem_posto = case(
        (PostoGrad.sigla == 'CEL', 1),
        (PostoGrad.sigla == 'TC', 2),
        (PostoGrad.sigla == 'MAJ', 3),
        (PostoGrad.sigla == 'CAP', 4),
        (PostoGrad.sigla == '1 TEN', 5),
        (PostoGrad.sigla == '2 TEN', 6),
        (PostoGrad.sigla == 'AL OF', 7),
        (PostoGrad.sigla == 'ALUNO OFICIAL', 8),
        (PostoGrad.sigla == 'SUBTENENTE', 9),
        (PostoGrad.sigla == '1 SGT', 10),
        (PostoGrad.sigla == '2 SGT', 11),
        (PostoGrad.sigla == '3 SGT', 12),
        (PostoGrad.sigla == 'AL SGT', 13),
        (PostoGrad.sigla == 'CB', 14),
        (PostoGrad.sigla == 'SD', 15),
        (PostoGrad.sigla == 'AL SD', 16),
        else_=99
    )

    sub_militares = (
        database.session.query(
            Militar.id.label("militar_id"),
            Militar.nome_completo,
            PostoGrad.sigla.label("posto_grad"),
            Quadro.quadro.label("quadro"),
            Obm.sigla.label("obm"),
            ordem_posto.label("ordem"),
            func.row_number().over(
                partition_by=Militar.id,
                order_by=[prioridade_obm.asc(), MilitarObmFuncao.id.desc()]
            ).label("linha")
        )
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .join(Obm, Obm.id == MilitarObmFuncao.obm_id)
        .join(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .join(Quadro, Quadro.id == Militar.quadro_id)
        .filter(MilitarObmFuncao.data_fim.is_(None))  # OBMs ativas
        .filter(~Militar.id.in_(subquery_pafs))  # sem PAF
        .subquery()
    )

    militares_sem_paf = (
        database.session.query(
            sub_militares.c.nome_completo,
            sub_militares.c.posto_grad,
            sub_militares.c.quadro,
            sub_militares.c.obm
        )
        .filter(sub_militares.c.linha == 1)
        .order_by(sub_militares.c.ordem, sub_militares.c.obm)
        .all()
    )

    return render_template("pafs_nao_preenchidos.html", militares=militares_sem_paf)


@app.route("/debug/militar/<int:militar_id>/full")
@login_required
def debug_militar_full(militar_id):
    # profundidade padrão (pode diminuir se ainda ficar pesado)
    depth = request.args.get("depth", default=4, type=int)

    # monta options de eager load para TODAS as relationships do Militar
    mapper = sa_inspect(Militar)
    rel_options = [
        selectinload(getattr(Militar, rel.key))
        for rel in mapper.relationships
    ]

    militar = (
        database.session.query(Militar)
        .options(*rel_options)
        .get(militar_id)
    )

    if not militar:
        return jsonify({"error": "Militar não encontrado"}), 404

    payload = sa_to_dict(militar, depth=depth, root_class=Militar)

    return jsonify(payload)


@app.route('/ferias-chefe', methods=['GET'])
@login_required
@checar_ocupacao('DIRETOR DRH', 'DIRETOR', 'CHEFE', 'SUPER USER', 'ATUALIZACAO CADASTRAL')
def exibir_ferias_chefe():
    dia_atual = datetime.now().day

    permitidas = obms_permitidas_para_usuario(current_user)
    lista_obms = Obm.query.filter(Obm.id.in_(
        sorted(permitidas))).order_by(Obm.sigla.asc()).all()

    ano_vigente = paf_ano_vigente()
    return render_template(
        'ferias_chefe2.html',
        lista_obms=lista_obms,
        ano_atual=datetime.now().year,
        dia_atual=dia_atual,
        ano_vigente=ano_vigente,
    )


@app.route('/pafs/tabela/<int:obm_id>', methods=['GET'])
@login_required
@checar_ocupacao('DIRETOR DRH', 'DIRETOR', 'CHEFE', 'SUPER USER', 'ATUALIZACAO CADASTRAL')
def carregar_tabela_obm(obm_id):
    if getattr(current_user, "funcao_user_id", None) != 6:
        permitidas = obms_permitidas_para_usuario(current_user)
        if obm_id not in permitidas:
            return "<div class='alert alert-danger'>Sem permissão para esta OBM.</div>", 403

    ano = int(request.args.get("ano") or datetime.now().year)

    meses = {
        "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4, "Maio": 5, "Junho": 6,
        "Julho": 7, "Agosto": 8, "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12
    }
    current_date = datetime.now().date()
    is_super = (getattr(current_user, "funcao_user_id", None) == 6)

    min_global = first_day_next_month(current_date)

    if is_super:
        min_iso = f"{ano}-01-01"
        min_year = 0
        min_month = 1
        bloqueio_mes_atual = False
    else:
        min_year = min_global.year
        min_month = min_global.month
        bloqueio_mes_atual = True

        if ano < min_year:
            min_iso = f"{ano}-12-31"
        elif ano == min_year:
            min_iso = min_global.isoformat()
        else:
            min_iso = f"{ano}-01-01"

    current_month = datetime.now().month
    obm = Obm.query.get(obm_id)
    if not obm:
        return "<div class='alert alert-danger'>OBM não encontrada</div>", 404

    militares_pafs = (
        database.session.query(Militar, Paf)
        .outerjoin(Paf, database.and_(
            Paf.militar_id == Militar.id,
            Paf.ano_referencia == ano
        ))
        .options(joinedload(Militar.obm_funcoes))
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
            Militar.inativo.is_(False)  # <-- NOVO FILTRO
        )
        .all()
    )

    return render_template(
        'partial_tabela_obm.html',
        obm=obm,
        militares_pafs=militares_pafs,
        meses=meses,
        current_month=current_month,
        current_date=current_date,
        ano=ano,
        min_iso=min_iso,
        min_year=min_year,
        min_month=min_month,
        bloqueio_mes_atual=bloqueio_mes_atual,
        is_super=is_super  # 👇 Injetando quem é Super User para a tabela
    )


# 2. ADICIONE ESTA NOVA ROTA NO SEU ARQUIVO DE ROTAS:


@app.route('/pafs/toggle_excecao', methods=['POST'])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER')
def toggle_excecao():
    militar_id = request.form.get('militar_id', type=int)
    ano = request.form.get('ano', type=int)
    excecao = request.form.get('excecao') == 'true'

    paf = Paf.query.filter_by(militar_id=militar_id,
                              ano_referencia=ano).first()

    # Se o PAF não existir ainda, cria um "vazio" só para registrar a flag da exceção
    if not paf:
        paf = Paf(militar_id=militar_id, ano_referencia=ano,
                  usuario_id=current_user.id)
        database.session.add(paf)

    paf.excecao_virada_ano = excecao
    paf.usuario_id = current_user.id

    try:
        from datetime import datetime
        # Assumindo que você tem now_manaus_naive() na sua aplicação
        paf.data_alteracao = now_manaus_naive()
    except:
        paf.data_alteracao = datetime.now()

    database.session.commit()

    return jsonify({"status": "success", "message": "Status de exceção atualizado."})


@app.route("/exportar-pafs-obm/<int:obm_id>")
@login_required
@checar_ocupacao('DIRETOR DRH', 'DIRETOR', 'CHEFE', 'SUPER USER', 'ATUALIZACAO CADASTRAL')
def exportar_pafs_obm(obm_id):
    # ✅ permissão por OBM (SUPER USER = id 6 no teu sistema)
    if getattr(current_user, "funcao_user_id", None) != 6:
        permitidas = obms_permitidas_para_usuario(current_user)
        if obm_id not in permitidas:
            return "Sem permissão para exportar esta OBM.", 403

    obm = Obm.query.get_or_404(obm_id)

    # ✅ ano: se vier por querystring, usa; senão, pega o último ano da tabela paf
    ano = request.args.get("ano", type=int)
    if not ano:
        ano = database.session.query(func.max(Paf.ano_referencia)).scalar()
    if not ano:
        return "Não há PAFs cadastrados para exportação.", 404

    # ✅ efetivo ativo da OBM + PAF do ano
    militares_pafs = (
        database.session.query(Militar, Paf)
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None)
        )
        .outerjoin(Paf, and_(
            Paf.militar_id == Militar.id,
            Paf.ano_referencia == ano
        ))
        .order_by(Militar.nome_completo.asc())
        .all()
    )

    # ==== Excel ====
    wb = Workbook()
    ws = wb.active
    ws.title = f"{obm.sigla} {ano}"

    colunas = [
        "OBM", "Ano", "Posto/Grad", "Nome", "Matrícula", "Quadro", "Mês Usufruto",
        "Qtd. Dias 1º", "Início 1º", "Fim 1º",
        "Qtd. Dias 2º", "Início 2º", "Fim 2º",
        "Qtd. Dias 3º", "Início 3º", "Fim 3º"
    ]
    ws.append(colunas)

    for col_num, col_name in enumerate(colunas, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = col_name
        c.font = Font(bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    def fmt(dt):
        return dt.strftime("%d/%m/%Y") if dt else ""

    for militar, paf in militares_pafs:
        ws.append([
            obm.sigla,
            ano,
            militar.posto_grad.sigla if militar.posto_grad else "",
            militar.nome_completo,
            militar.matricula,
            militar.quadro.quadro if militar.quadro else "",
            paf.mes_usufruto if paf else "",

            paf.qtd_dias_primeiro_periodo if paf else "",
            fmt(paf.primeiro_periodo_ferias) if paf else "",
            fmt(paf.fim_primeiro_periodo) if paf else "",

            paf.qtd_dias_segundo_periodo if paf else "",
            fmt(paf.segundo_periodo_ferias) if paf else "",
            fmt(paf.fim_segundo_periodo) if paf else "",

            paf.qtd_dias_terceiro_periodo if paf else "",
            fmt(paf.terceiro_periodo_ferias) if paf else "",
            fmt(paf.fim_terceiro_periodo) if paf else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    registrar_log_download(
        nome_relatorio=f"Plano Anual de Férias (PAF) - {obm.sigla}",
        colunas_lista=colunas,  # Puxa a lista de colunas dinâmica
        filtros_dict={
            "ano_referencia": ano,
            "obm_id": obm_id,
            "obm_sigla": obm.sigla
        }
    )

    filename = f"pafs_{obm.sigla}_{ano}.xlsx"
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@app.route('/grafico-ferias/<int:obm_id>', methods=['GET'])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER')
def grafico_ferias(obm_id):
    # Lista de OBMs adicionais para obm_id_1 == 16
    obms_adicionais = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25]

    # Seleciona militares da OBM específica
    militares = (
        database.session.query(Militar, Paf)
        .outerjoin(Paf, Paf.militar_id == Militar.id)
        .join(MilitarObmFuncao, Militar.id == MilitarObmFuncao.militar_id)
        .filter(
            (MilitarObmFuncao.obm_id == obm_id) |
            (MilitarObmFuncao.obm_id.in_(obms_adicionais) if obm_id == 16 else False)
        )
        .all()
    )

    # Contar número de militares de férias por mês
    ferias_por_mes = {mes.id: 0 for mes in Meses.query.all()}
    for militar, paf in militares:
        if paf:
            if paf.primeiro_periodo_ferias:
                mes = paf.primeiro_periodo_ferias.month
                ferias_por_mes[mes] += 1
            if paf.segundo_periodo_ferias:
                mes = paf.segundo_periodo_ferias.month
                ferias_por_mes[mes] += 1
            if paf.terceiro_periodo_ferias:
                mes = paf.terceiro_periodo_ferias.month
                ferias_por_mes[mes] += 1

    # Gerar gráfico
    labels = [mes.mes for mes in Meses.query.all()]
    values = [ferias_por_mes[mes.id] for mes in Meses.query.all()]

    plt.figure(figsize=(10, 6))
    plt.bar(labels, values, color='skyblue')
    plt.xlabel('Mês')
    plt.ylabel('Número de Militares de Férias')
    plt.title('Militares de Férias por Mês')
    plt.xticks(rotation=25)

    # Salvar gráfico em um buffer de memória
    buf = BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    image_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()

    return Response(response=image_base64, status=200, mimetype='text/plain')


@app.route('/pafs/update', methods=['POST'])
@login_required
def update_paf():
    hoje = datetime.now().day

    if (hoje < 10 or hoje > 20) and not can_ferias_bypass_janela():
        return jsonify({"message": "Alterações só são permitidas de 10 a 20 de cada mês."}), 403

    # permissão de ação (importante!)
    if not (is_super() or is_super_or_perm("FERIAS_UPDATE") or is_super_or_perm("FERIAS_SUPER")):
        return jsonify({"error": "Sem permissão para atualizar PAF."}), 403

    data = request.form
    militar_id = int(data.get('militar_id') or 0)
    ano = int(data.get('ano_referencia') or datetime.now().year)

    if not militar_id:
        return jsonify({"error": "militar_id inválido"}), 400

    # ✅ escopo OBM: só exige para quem NÃO é super real e NÃO tem super de férias
    if not (is_super() or is_super_or_perm("FERIAS_SUPER")):
        permitidas = obms_permitidas_para_usuario(current_user)
        if not militar_esta_no_escopo(militar_id, permitidas):
            return jsonify({"error": "Sem permissão para alterar PAF deste militar."}), 403

    # ✅ novo: busca por militar + ano
    paf = Paf.query.filter_by(militar_id=militar_id,
                              ano_referencia=ano).first()
    if not paf:
        paf = Paf(militar_id=militar_id, ano_referencia=ano)
        database.session.add(paf)

    mes_usufruto = data.get('mes_usufruto')

    qtd_dias_primeiro_periodo = int(data.get('qtd_dias_1') or 0)
    primeiro_periodo_inicio = parse_date(data.get('inicio_1'))
    primeiro_periodo_fim = parse_date(data.get('fim_1'))

    qtd_dias_segundo_periodo = int(data.get('qtd_dias_2') or 0)
    segundo_periodo_inicio = parse_date(data.get('inicio_2'))
    segundo_periodo_fim = parse_date(data.get('fim_2'))

    qtd_dias_terceiro_periodo = int(data.get('qtd_dias_3') or 0)
    terceiro_periodo_inicio = parse_date(data.get('inicio_3'))
    terceiro_periodo_fim = parse_date(data.get('fim_3'))

    # validação (mantém tua lógica)
    try:
        if primeiro_periodo_inicio:
            validate_vacation_period(
                primeiro_periodo_inicio, qtd_dias_primeiro_periodo)
        if segundo_periodo_inicio:
            validate_vacation_period(
                segundo_periodo_inicio, qtd_dias_segundo_periodo)
        if terceiro_periodo_inicio:
            validate_vacation_period(
                terceiro_periodo_inicio, qtd_dias_terceiro_periodo)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    paf.mes_usufruto = mes_usufruto
    paf.qtd_dias_primeiro_periodo = qtd_dias_primeiro_periodo
    paf.primeiro_periodo_ferias = primeiro_periodo_inicio
    paf.fim_primeiro_periodo = primeiro_periodo_fim

    paf.qtd_dias_segundo_periodo = qtd_dias_segundo_periodo
    paf.segundo_periodo_ferias = segundo_periodo_inicio
    paf.fim_segundo_periodo = segundo_periodo_fim

    paf.qtd_dias_terceiro_periodo = qtd_dias_terceiro_periodo
    paf.terceiro_periodo_ferias = terceiro_periodo_inicio
    paf.fim_terceiro_periodo = terceiro_periodo_fim

    paf.usuario_id = current_user.id
    paf.data_alteracao = now_manaus_naive()

    database.session.commit()
    return jsonify({"message": "Dados salvos com sucesso!"})
