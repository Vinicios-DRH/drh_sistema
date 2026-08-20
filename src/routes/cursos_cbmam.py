"""Cursos CBMAM: administração pela BM-3 (cadastro de cursos, abertura de
edições/inscrições e análise dos pedidos) e autoatendimento do militar
(inscrição em cursos abertos, acompanhamento do parecer).

Regras de negócio em src.services.cursos_cbmam_service; controle de acesso
em src.authz (can_manage_cursos_cbmam / pode_ver_autoatendimento_militar).
"""
from io import BytesIO

from flask import abort, current_app, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import app, database
from src.authz import can_manage_cursos_cbmam, is_super, pode_ver_autoatendimento_militar
from src.formatar_cpf import get_militar_por_user
from src.models import Militar, PostoGrad
from src.services.cursos_cbmam_service import (
    analisar_solicitacao,
    atualizar_curso_andamento,
    cancelar_curso_andamento,
    criar_curso,
    criar_curso_andamento,
    criar_solicitacao_inscricao,
    listar_auditoria_andamento,
    listar_auditoria_solicitacoes_do_andamento,
    listar_cursos_andamento,
    listar_cursos_base,
    listar_cursos_disponiveis_para_militar,
    listar_militares_inscritos_para_relatorio,
    listar_minhas_solicitacoes,
    listar_solicitacoes_para_analise,
    obter_curso_andamento,
    obter_solicitacao,
    reativar_curso_andamento,
)
from src.services.militar_situacao_service import parse_date_flex
from src.utils.utils import registrar_log_download


def _resolver_militar_atual():
    mid = getattr(current_user, "militar_id", None)
    if mid:
        militar = Militar.query.get(mid)
        if militar:
            return militar
    return get_militar_por_user(current_user)


# ---------------------------------------------------------------------------
# BM-3 — administração
# ---------------------------------------------------------------------------

@app.route("/cursos-cbmam/admin")
@login_required
def cursos_cbmam_admin():
    if not can_manage_cursos_cbmam():
        abort(403)

    return render_template(
        "cursos_cbmam_admin.html",
        cursos_base=listar_cursos_base(),
        andamentos=listar_cursos_andamento(),
        postos_grad=PostoGrad.query.order_by(PostoGrad.id.asc()).all(),
        solicitacoes_pendentes_qtd=len(listar_solicitacoes_para_analise(apenas_pendentes=True)),
    )


@app.post("/cursos-cbmam/admin/cursos")
@login_required
def cursos_cbmam_criar_curso():
    if not can_manage_cursos_cbmam():
        abort(403)

    try:
        criar_curso(
            nome=request.form.get("nome"),
            descricao=request.form.get("descricao"),
        )
        database.session.commit()
        flash("Curso cadastrado no catálogo.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")
    except Exception as e:
        database.session.rollback()
        current_app.logger.exception("Erro ao criar curso CBMAM")
        flash(f"Erro ao cadastrar curso: {str(e)}", "alert-danger")

    return redirect(url_for("cursos_cbmam_admin"))


@app.post("/cursos-cbmam/admin/andamentos")
@login_required
def cursos_cbmam_criar_andamento():
    if not can_manage_cursos_cbmam():
        abort(403)

    try:
        criar_curso_andamento(
            curso_id=request.form.get("curso_id"),
            data_inicio=parse_date_flex(request.form.get("data_inicio")),
            data_fim=parse_date_flex(request.form.get("data_fim")),
            data_limite_inscricao=parse_date_flex(request.form.get("data_limite_inscricao")),
            destinado_a=request.form.get("destinado_a"),
            posto_grad_ids=request.form.getlist("posto_grad_ids"),
            criado_por_user_id=current_user.id,
        )
        database.session.commit()
        flash("Inscrições abertas! O curso já está disponível para os militares elegíveis.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")
    except Exception as e:
        database.session.rollback()
        current_app.logger.exception("Erro ao abrir edição de curso CBMAM")
        flash(f"Erro ao abrir inscrições: {str(e)}", "alert-danger")

    return redirect(url_for("cursos_cbmam_admin"))


@app.post("/cursos-cbmam/admin/andamentos/<int:andamento_id>/editar")
@login_required
def cursos_cbmam_editar_andamento(andamento_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    try:
        atualizar_curso_andamento(
            andamento,
            data_inicio=parse_date_flex(request.form.get("data_inicio")),
            data_fim=parse_date_flex(request.form.get("data_fim")),
            data_limite_inscricao=parse_date_flex(request.form.get("data_limite_inscricao")),
            destinado_a=request.form.get("destinado_a"),
            posto_grad_ids=request.form.getlist("posto_grad_ids"),
            editado_por_user_id=current_user.id,
        )
        database.session.commit()
        flash("Edição do curso atualizada.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")
    except Exception as e:
        database.session.rollback()
        current_app.logger.exception("Erro ao editar edição de curso CBMAM")
        flash(f"Erro ao atualizar: {str(e)}", "alert-danger")

    return redirect(url_for("cursos_cbmam_admin_andamento", andamento_id=andamento_id))


@app.post("/cursos-cbmam/admin/andamentos/<int:andamento_id>/cancelar")
@login_required
def cursos_cbmam_cancelar_andamento(andamento_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    try:
        cancelar_curso_andamento(andamento, cancelado_por_user_id=current_user.id)
        database.session.commit()
        flash("Edição cancelada. As inscrições já recebidas continuam registradas.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")

    return redirect(url_for("cursos_cbmam_admin_andamento", andamento_id=andamento_id))


@app.post("/cursos-cbmam/admin/andamentos/<int:andamento_id>/reativar")
@login_required
def cursos_cbmam_reativar_andamento(andamento_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    try:
        reativar_curso_andamento(andamento, reativado_por_user_id=current_user.id)
        database.session.commit()
        flash("Cancelamento desfeito.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")

    return redirect(url_for("cursos_cbmam_admin_andamento", andamento_id=andamento_id))


@app.route("/cursos-cbmam/admin/andamentos/<int:andamento_id>/relatorio-excel")
@login_required
def cursos_cbmam_relatorio_excel(andamento_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    linhas = listar_militares_inscritos_para_relatorio(andamento_id)

    headers = [
        "Posto/Grad", "Quadro", "Nome Completo", "Nome de Guerra",
        "OBM 1", "OBM 2", "Telefone", "Telefone de Emergência",
        "Contato de Emergência", "Situação da Inscrição",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscritos"
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0B4A7D")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for linha in linhas:
        ws.append([
            linha["posto_grad"], linha["quadro"], linha["nome_completo"], linha["nome_guerra"],
            linha["obm_1"], linha["obm_2"], linha["telefone"], linha["telefone_emergencia"],
            linha["contato_emergencia_nome"], linha["situacao_inscricao"],
        ])

    larguras = [10, 12, 32, 20, 12, 12, 16, 18, 26, 18]
    for col_num, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = largura
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_curso = andamento.curso.nome if andamento.curso else "curso"
    registrar_log_download(
        nome_relatorio=f"Cursos CBMAM — Inscritos ({nome_curso})",
        colunas_lista=headers,
        filtros_dict={"curso_andamento_id": andamento_id},
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=f"inscritos_{andamento_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/cursos-cbmam/admin/andamentos/<int:andamento_id>")
@login_required
def cursos_cbmam_admin_andamento(andamento_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    return render_template(
        "cursos_cbmam_admin_andamento.html",
        andamento=andamento,
        solicitacoes=listar_solicitacoes_para_analise(curso_andamento_id=andamento_id),
        postos_grad=PostoGrad.query.order_by(PostoGrad.id.asc()).all(),
        auditoria_andamento=listar_auditoria_andamento(andamento_id),
        auditoria_solicitacoes=listar_auditoria_solicitacoes_do_andamento(andamento_id),
    )


@app.post("/cursos-cbmam/admin/solicitacoes/<int:solicitacao_id>/analisar")
@login_required
def cursos_cbmam_analisar_solicitacao(solicitacao_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    solicitacao = obter_solicitacao(solicitacao_id)
    if not solicitacao:
        abort(404)

    deferido = request.form.get("deferido") == "1"
    analisar_solicitacao(
        solicitacao,
        deferido=deferido,
        observacao=request.form.get("observacao_analise"),
        analisado_por_user_id=current_user.id,
    )
    database.session.commit()

    flash(
        "Inscrição deferida." if deferido else "Inscrição indeferida.",
        "alert-success" if deferido else "alert-warning",
    )
    return redirect(url_for("cursos_cbmam_admin_andamento", andamento_id=solicitacao.curso_andamento_id))


@app.route("/cursos-cbmam/admin/solicitacoes/<int:solicitacao_id>/arquivo")
@login_required
def cursos_cbmam_admin_arquivo(solicitacao_id):
    if not can_manage_cursos_cbmam():
        abort(403)

    solicitacao = obter_solicitacao(solicitacao_id)
    if not solicitacao:
        abort(404)

    return redirect(solicitacao.url_arquivo)


# ---------------------------------------------------------------------------
# Militar — autoatendimento
# ---------------------------------------------------------------------------

@app.route("/meus-cursos")
@login_required
def meus_cursos():
    if not (is_super() or pode_ver_autoatendimento_militar()):
        abort(403)

    militar = _resolver_militar_atual()
    if not militar:
        flash("Não foi possível localizar seus dados de militar.", "alert-warning")
        return redirect(url_for("home"))

    return render_template(
        "meus_cursos.html",
        militar=militar,
        cursos_disponiveis=listar_cursos_disponiveis_para_militar(militar),
        minhas_solicitacoes=listar_minhas_solicitacoes(militar.id),
    )


@app.post("/meus-cursos/<int:andamento_id>/inscrever")
@login_required
def meus_cursos_inscrever(andamento_id):
    if not (is_super() or pode_ver_autoatendimento_militar()):
        abort(403)

    militar = _resolver_militar_atual()
    if not militar:
        flash("Não foi possível localizar seus dados de militar.", "alert-warning")
        return redirect(url_for("home"))

    andamento = obter_curso_andamento(andamento_id)
    if not andamento:
        abort(404)

    try:
        criar_solicitacao_inscricao(andamento, militar, request.files.get("arquivo_inscricao"))
        database.session.commit()
        flash("Inscrição enviada! Acompanhe o parecer aqui mesmo.", "alert-success")
    except ValueError as e:
        database.session.rollback()
        flash(str(e), "alert-warning")
    except Exception as e:
        database.session.rollback()
        current_app.logger.exception("Erro ao enviar inscrição de curso CBMAM")
        flash(f"Erro ao enviar inscrição: {str(e)}", "alert-danger")

    return redirect(url_for("meus_cursos"))


@app.route("/meus-cursos/solicitacoes/<int:solicitacao_id>/arquivo")
@login_required
def meus_cursos_arquivo(solicitacao_id):
    if not (is_super() or pode_ver_autoatendimento_militar()):
        abort(403)

    militar = _resolver_militar_atual()
    solicitacao = obter_solicitacao(solicitacao_id)
    if not solicitacao or not militar or solicitacao.militar_id != militar.id:
        abort(404)

    return redirect(solicitacao.url_arquivo)
