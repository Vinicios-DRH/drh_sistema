import unicodedata

from flask import current_app
import io
import math
from zoneinfo import ZoneInfo
from flask_wtf.csrf import validate_csrf
from flask_login import login_required
from flask import abort, json, request, jsonify, make_response, current_app
from random import shuffle
import os
import zipfile
import qrcode
import pytz
import pandas as pd
import base64
import matplotlib.pyplot as plt
import requests
import urllib
from src.decorators.utils_acumulo import b2_bucket_name, b2_client, b2_delete_all_versions, b2_upload_fileobj
from src.identificacao import buscar_pessoa_por_cpf, normaliza_matricula
from src.formatar_cpf import cadete_restantes, formatar_cpf, get_militar_por_user, is_cadete
from flask import render_template, redirect, url_for, request, flash, jsonify, session, send_file, make_response, \
    Response, stream_with_context
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import validate_csrf, generate_csrf
from werkzeug.utils import secure_filename
from src import app, database, bcrypt
from src.forms import (AtualizacaoCadastralForm, ControleConvocacaoForm, CriarSenhaForm, FichaAlunosForm, FormEsqueciSenha, FormFiltroMilitar, FormMilitarInativo, FormResetarSenhaPublica, FormViatura,
                       IdentificacaoForm, ImpactoForm, FormLogin, FormMilitar, FormCriarUsuario, FormMotoristas, FormFiltroMotorista, LtsAlunoForm, RecompensaAlunoForm,
                       RestricaoAlunoForm, SancaoAlunoForm, TabelaVencimentoForm, InativarAlunoForm, MatriculaConfirmForm)
from src.models import (ControleConvocacao, Convocacao, DocumentoMilitar, EfetivoDiarioOBM, HistoricoEfetivoDiario, ImportacaoMilitarHistorico, LogAcesso, LtsAlunos, Militar, MilitaresInativos, NomeConvocado, PostoGrad, Quadro, Obm, Localidade, Funcao, RecompensaAluno, RestricaoAluno, SancaoAluno, SegundoVinculo, SituacaoConvocacao, User, FuncaoUser, PublicacaoBg,
                        EstadoCivil, Especialidade, Destino, Motivo, Modalidade, Punicao, Comportamento, MilitarObmFuncao,
                        FuncaoGratificada,
                        MilitaresAgregados, MilitaresADisposicao, LicencaEspecial, LicencaParaTratamentoDeSaude, Paf,
                        Meses, Motoristas, Categoria, TabelaVencimento, ValorDetalhadoPostoGrad, FichaAlunos, AlunoInativo, Viaturas, ViaturaMilitar, MilitarGraduacao, MilitarContatoEmergencia, MilitarConjuge, LogExportacaoExcel, Curso, MilitarCurso, AuditoriaAtualizacaoCadastral, now_manaus_naive)
from src.querys import dados_para_mapa, obter_estatisticas_militares, login_usuario
from src.decorators.control import checar_ocupacao, militar_esta_no_escopo, obms_permitidas_para_usuario, sync_user_admin_obms_from_militar
from src.decorators.business_logic import processar_militares_a_disposicao, processar_militares_agregados, \
    processar_militares_le, processar_militares_lts
from datetime import datetime, date, timedelta
from io import BytesIO
from sqlalchemy.orm import joinedload, selectinload, load_only, aliased
from sqlalchemy import case, distinct, extract, func, not_, or_, cast, String, and_
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP, getcontext
from docx import Document
from urllib.parse import urlencode
from collections import defaultdict, Counter
from docx.shared import Pt
from docx.oxml.ns import qn
from src.decorators.formatar_datas import formatar_data_extenso, formatar_data_sem_zero
from src.decorators.email_utils import send_reset_password_email, verify_password_reset_token
import re
import plotly.graph_objs as go
import plotly.io as pio
from collections import defaultdict
from src.utils.sa_serialize import sa_to_dict
from sqlalchemy.inspection import inspect as sa_inspect
from src.security.perms import has_perm
from src.authz import is_super_or_perm, can_ferias_bypass_janela, is_super, require_perm
from src.utils.cadastro_status import cadastro_esta_completo
from src.utils.painel import (
    _obter_obm_principal,
    listar_situacoes_atualizacao,
    obter_resumo_atualizacao_cadastral,
    obter_militares_atualizacao_cadastral,
    serializar_militar_atualizacao,
    listar_obms_atualizacao,
    listar_postos_grad_atualizacao,
    obter_detalhes_militar_atualizacao,)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.services.importar_militares import (
    ler_planilha,
    colunas_reconhecidas,
    colunas_nao_reconhecidas,
    analisar_importacao,
    importar_dataframe,
    salvar_historico_importacao,
)
from src.services.militar_situacao_service import (
    parse_date_flex,
    sincronizar_blocos_funcionais,
)
import time
from src.utils.utils import registrar_log_download
from dateutil.relativedelta import relativedelta

from src.routes_helpers import build_tabela_militares_query, get_status_sets


@app.route("/export-excel", methods=["POST"])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'MAPA DA FORÇA', 'SUPER USER', 'DRH', 'DIRETOR DRH')
def export_excel():
    try:
        # Pega as colunas selecionadas pelo usuário no modal
        colunas_selecionadas = request.form.getlist('colunas')

        # Prevenção caso o usuário envie vazio
        if not colunas_selecionadas:
            return jsonify({'error': 'Nenhuma coluna selecionada para exportação.'}), 400

        today = date.today()

        query = (
            build_tabela_militares_query()
            .order_by(Militar.nome_completo.asc())
        )

        militares_filtrados = query.all()

        agregados_ids, adisposicao_ids = get_status_sets(query, today)

        rows = []
        for militar in militares_filtrados:
            obm_funcoes_ativas = sorted(
                [of for of in militar.obm_funcoes if of.data_fim is None],
                key=lambda of: of.data_criacao or date.min,
                reverse=True
            )

            obms = [
                of.obm.sigla if of.obm else 'OBM não encontrada' for of in obm_funcoes_ativas]
            funcoes = [
                of.funcao.ocupacao if of.funcao else 'Função não encontrada' for of in obm_funcoes_ativas]

            destino_txt = 'N/A'
            try:
                if getattr(militar, 'destino', None):
                    destino_txt = getattr(
                        militar.destino, 'local', None) or 'N/A'
                elif getattr(militar, 'destino_id', None):
                    d = Destino.query.get(militar.destino_id)
                    destino_txt = getattr(d, 'local', None) or str(
                        militar.destino_id)
            except Exception:
                pass

            inclusao_fmt = militar.inclusao.strftime(
                '%d/%m/%Y') if militar.inclusao else 'N/A'

            situacao = (militar.situacao or "").strip().upper()

            modalidade = (
                (militar.modalidade.descricao or "").strip().upper()
                if militar.modalidade
                else ""
            )

            destino_txt = (
                militar.destino.local
                if militar.destino and militar.destino.local
                else "N/A"
            )

            # Agregado
            agregado_exibe = (
                "SIM"
                if situacao == "AGREGADO"
                else "NÃO"
            )

            # À disposição
            adisposicao_exibe = (
                "SIM"
                if modalidade == "À DISPOSIÇÃO"
                else "NÃO"
            )

            modalidade_exibe = (
                militar.modalidade.descricao
                if militar.modalidade
                else "N/A"
            )

            sexo_raw = (militar.sexo or '').strip().lower()
            sexo_exibe = (
                'Masculino' if sexo_raw.startswith('m')
                else 'Feminino' if sexo_raw.startswith('f')
                else (militar.sexo or 'N/A')
            )

            linha_completa = {
                'Nome Completo': militar.nome_completo or 'N/A',
                'Nome de Guerra': militar.nome_guerra or 'N/A',
                'Posto/Graduação': militar.posto_grad.sigla if militar.posto_grad else 'N/A',
                'Quadro': militar.quadro.quadro if militar.quadro else 'N/A',
                'Sexo': sexo_exibe,
                'Raça/Cor': militar.raca or 'N/A',
                'CPF': militar.cpf or 'N/A',
                'RG': militar.rg or 'N/A',
                'Matrícula': militar.matricula or 'N/A',
                'Inclusão': inclusao_fmt,
                'Especialidade': militar.especialidade.ocupacao if militar.especialidade else 'N/A',
                'Localidade': militar.localidade.sigla if militar.localidade else 'N/A',

                'Situação': militar.situacao or 'N/A',
                'Modalidade': modalidade_exibe,
                'Agregado': agregado_exibe,
                'À Disposição': adisposicao_exibe,

                'Destino': destino_txt,

                'OBM 1': obms[0] if len(obms) > 0 else 'N/A',
                'Função 1': funcoes[0] if len(funcoes) > 0 else 'N/A',
                'OBM 2': obms[1] if len(obms) > 1 else 'N/A',
                'Função 2': funcoes[1] if len(funcoes) > 1 else 'N/A',

                'Data de Nascimento': militar.data_nascimento.strftime('%d/%m/%Y') if militar.data_nascimento else 'N/A',
                'Graduação': militar.graduacao or 'N/A',
                'Grau de Instrução': militar.grau_instrucao or 'N/A',
                'Pós-Graduação': militar.pos_graduacao or 'N/A',
                'Mestrado': militar.mestrado or 'N/A',
                'Doutorado': militar.doutorado or 'N/A',
            }

            # Filtra o dicionário mantendo APENAS as colunas que o usuário selecionou
            linha_filtrada = {col: linha_completa.get(
                col, 'N/A') for col in colunas_selecionadas}
            rows.append(linha_filtrada)

        # Monta o DataFrame apenas com as colunas selecionadas
        df = pd.DataFrame(rows, columns=colunas_selecionadas)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Militares')

            workbook = writer.book
            worksheet = writer.sheets['Militares']

            header_format = workbook.add_format({
                'bg_color': '#0b2e4f',
                'font_color': '#FFFFFF',
                'bold': True,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            body_format = workbook.add_format({
                'text_wrap': True,
                'valign': 'top',
                'border': 1
            })

            largura_colunas = {
                'Nome Completo': 38,
                'Nome de Guerra': 24,
                'Posto/Graduação': 18,
                'Quadro': 18,
                'Sexo': 14,
                'Raça/Cor': 16,
                'CPF': 16,
                'RG': 16,
                'Matrícula': 16,
                'Inclusão': 14,
                'Especialidade': 24,
                'Localidade': 14,

                'Situação': 18,
                'Modalidade': 18,
                'Agregado': 14,
                'À Disposição': 16,

                'Destino': 24,
                'OBM 1': 16,
                'Função 1': 24,
                'OBM 2': 16,
                'Função 2': 24,
                'Data de Nascimento': 16,
                'Graduação': 18,
                'Grau de Instrução': 22,
                'Pós-Graduação': 20,
                'Mestrado': 18,
                'Doutorado': 18,
            }

            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                # Define a largura com base no nome da coluna (default 20 se não achar)
                width = largura_colunas.get(value, 20)
                worksheet.set_column(col_num, col_num, width, body_format)

            worksheet.freeze_panes(1, 0)

            if len(df) > 0:
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

        output.seek(0)

        # --- AUDITORIA ---
        # Registra a ação no banco de dados antes de enviar o arquivo
        # Salva os parâmetros da URL como string
        filtros_usados = str(request.args.to_dict())

        registrar_log_download(
            nome_relatorio="Exportação de Militares Filtrados",
            colunas_lista=colunas_selecionadas,
            filtros_dict={"status": "; ".join(filtros_usados)},
        )
        # -----------------

        return send_file(
            output,
            as_attachment=True,
            download_name='militares_filtrados.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        app.logger.error(f"Erro ao exportar Excel: {str(e)}")
        # Em caso de erro, importante dar um rollback
        database.session.rollback()
        return jsonify({
            'error': 'Ocorreu um erro ao exportar o Excel.',
            'details': str(e)
        }), 500


@app.route("/auditoria/exportacoes", methods=["GET"])
@login_required
@checar_ocupacao('SUPER USER')
def auditoria_exportacoes():
    try:
        # Busca os logs ordenados do mais recente para o mais antigo
        # Assumindo que você fez o relationship 'usuario' na model LogExportacaoExcel
        logs = LogExportacaoExcel.query.order_by(
            LogExportacaoExcel.data_download.desc()).all()

        return render_template("auditoria_exportacoes.html", logs=logs)
    except Exception as e:
        app.logger.error(f"Erro ao carregar auditoria: {str(e)}")
        flash("Erro ao carregar os dados de auditoria.", "error")
        return redirect(url_for('home'))


@app.route("/auditoria/exportacoes/exportar-excel", methods=["GET"])
@login_required
@checar_ocupacao('SUPER USER')
def exportar_auditoria_excel():
    try:
        logs = LogExportacaoExcel.query.order_by(
            LogExportacaoExcel.data_download.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Log de Exportações"

        # Cabeçalhos
        headers = [
            "ID Log", "Relatório Baixado", "Data e Hora", "Operador",
            "CPF do Operador", "Endereço IP", "Qtd. Colunas", "Filtros Usados"
        ]
        ws.append(headers)

        # Estilo do Header
        header_fill = PatternFill("solid", fgColor="0B4A7D")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados
        for log in logs:
            cpf_user = log.usuario.cpf_norm if log.usuario else "N/A"
            nome_user = log.usuario.nome if log.usuario else "Usuário Deletado"
            data_str = log.data_download.strftime(
                '%d/%m/%Y %H:%M:%S') if log.data_download else "N/A"
            qtd_colunas = len(log.colunas_selecionadas.split(
                ';')) if log.colunas_selecionadas else 0

            ws.append([
                log.id,
                log.nome_relatorio or "Relatório Genérico",
                data_str,
                nome_user,
                cpf_user,
                log.ip_address,
                qtd_colunas,
                log.filtros_aplicados
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ----------------------------------------------------------------- #
        # O PULO DO GATO: Registra o download DESTA auditoria no banco!     #
        # ----------------------------------------------------------------- #
        registrar_log_download(
            nome_relatorio="Auditoria de Exportações (Logs)",
            colunas_lista=headers,
            filtros_dict={"busca": "Todos os registros de log no banco"}
        )

        data_exportacao = datetime.now(
            ZoneInfo("America/Manaus")).strftime("%Y%m%d_%H%M%S")

        return send_file(
            output,
            as_attachment=True,
            download_name=f"auditoria_downloads_{data_exportacao}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        app.logger.error(f"Erro ao exportar excel de auditoria: {str(e)}")
        flash("Erro ao gerar o arquivo Excel.", "error")
        return redirect(url_for('auditoria_exportacoes'))


@app.route("/exportar-excel/<string:tabela>")
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'MAPA DA FORÇA', 'DRH', 'SUPER USER', 'DIRETOR DRH')
def exportar_excel(tabela):
    # Mapeamento das tabelas para consultas
    tabela_mapping = {
        'militares_agregados': {
            'model': MilitaresAgregados,
            'columns': [
                'posto_grad.sigla', 'quadro.quadro', 'militar.nome_completo',
                'destino.local', 'situacao.condicao', 'inicio_periodo',
                'fim_periodo_agregacao', 'status', 'publicacao_bg.boletim_geral'
            ]
        },
        'militares_a_disposicao': {
            'model': MilitaresADisposicao,
            'columns': [
                'posto_grad.sigla', 'quadro.quadro', 'militar.nome_completo',
                'destino.local', 'situacao.condicao', 'inicio_periodo',
                'fim_periodo_disposicao', 'status', 'publicacao_bg.boletim_geral'
            ]
        }
    }

    # Verifica se a tabela especificada está no mapeamento
    if tabela not in tabela_mapping:
        return jsonify({'error': 'Tabela não encontrada'}), 404

    # Obter modelo e colunas da tabela especificada
    tabela_info = tabela_mapping[tabela]
    modelo = tabela_info['model']
    colunas = tabela_info['columns']

    # Consultar os dados da tabela
    dados = modelo.query.all()

    # Construir dados para o DataFrame do Pandas
    export_data = []
    for item in dados:
        export_row = {
            'Posto/Graduação': getattr(item.posto_grad, 'sigla', 'N/A'),
            'Quadro': getattr(item.quadro, 'quadro', 'N/A'),
            'Nome Completo': getattr(item.militar, 'nome_completo', 'N/A'),
            'Destino': getattr(item.destino, 'local', 'N/A'),
            'Situação': getattr(item.situacao, 'condicao', 'N/A'),
            'A contar de': item.inicio_periodo.strftime('%d/%m/%Y') if item.inicio_periodo else 'N/A',
            'Término': (
                item.fim_periodo_agregacao.strftime('%d/%m/%Y') if tabela == 'militares_agregados' else
                item.fim_periodo_disposicao.strftime(
                    '%d/%m/%Y') if item.fim_periodo_disposicao else 'N/A'
            ),
            'Status': item.status,
            'Documento Autorizador': getattr(item.publicacao_bg, 'boletim_geral', 'N/A')
        }
        export_data.append(export_row)

    # Criar DataFrame
    df = pd.DataFrame(export_data)

    # Gerar arquivo Excel em memória
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
        workbook = writer.book
        worksheet = writer.sheets['Dados']

        # Definir formatos personalizados
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1F4E78',  # Azul escuro
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter'
        })

        cell_centered_format = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter'})
        cell_left_format = workbook.add_format(
            {'align': 'left', 'valign': 'vcenter'})

        # Ajustar largura das colunas
        column_widths = {
            'Posto/Graduação': 15,
            'Quadro': 15,
            'Nome Completo': 30,
            'Destino': 20,
            'Situação': 15,
            'A contar de': 12,
            'Término': 12,
            'Status': 15,
            'Documento Autorizador': 30
        }

        # Aplicar largura das colunas e formatação
        for col_num, column_title in enumerate(df.columns):
            width = column_widths.get(column_title, 15)  # Valor padrão 15
            worksheet.set_column(col_num, col_num, width)

            # Aplicar estilo ao cabeçalho
            worksheet.write(0, col_num, column_title, header_format)

            # Aplicar centralização condicional para colunas específicas
            if column_title in ['Nome Completo', 'Documento Autorizador']:
                worksheet.set_column(col_num, col_num, width, cell_left_format)
            else:
                worksheet.set_column(
                    col_num, col_num, width, cell_centered_format)

    output.seek(0)

    # Enviar arquivo Excel para download
    nome_arquivo = f'{tabela}.xlsx'
    return send_file(output, as_attachment=True, download_name=nome_arquivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route("/exportar-pafs/<string:tabela>")
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'MAPA DA FORÇA', 'DRH', 'SUPER USER', 'DIRETOR DRH')
def exportar_pafs(tabela):
    if tabela != "pafs":
        return "Tabela inválida", 400

    # ✅ pega o último ano que existe na tabela paf
    ultimo_ano = database.session.query(func.max(Paf.ano_referencia)).scalar()
    if not ultimo_ano:
        return "Não há PAFs cadastrados para exportação.", 404

    # ✅ traz todos os militares, mas só o PAF do último ano
    militares_pafs = (
        database.session.query(Militar, Paf)
        .outerjoin(Paf, and_(
            Paf.militar_id == Militar.id,
            Paf.ano_referencia == ultimo_ano
        ))
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"PAFs {ultimo_ano}"

    colunas = [
        "Ano", "Posto/Grad", "Nome", "Matrícula", "Quadro", "Mês Usufruto",
        "Qtd. Dias 1º Período", "1º Período de Férias", "Fim 1º Período",
        "Qtd. Dias 2º Período", "2º Período de Férias", "Fim 2º Período",
        "Qtd. Dias 3º Período", "3º Período de Férias", "Fim 3º Período"
    ]
    ws.append(colunas)

    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_name
        cell.font = Font(bold=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    for militar, paf in militares_pafs:
        ws.append([
            ultimo_ano,
            militar.posto_grad.sigla if militar.posto_grad else "",
            militar.nome_completo,
            militar.matricula,
            militar.quadro.quadro if militar.quadro else "",
            paf.mes_usufruto if paf else "",
            paf.qtd_dias_primeiro_periodo if paf else "",
            paf.primeiro_periodo_ferias.strftime(
                "%d/%m/%Y") if paf and paf.primeiro_periodo_ferias else "",
            paf.fim_primeiro_periodo.strftime(
                "%d/%m/%Y") if paf and paf.fim_primeiro_periodo else "",
            paf.qtd_dias_segundo_periodo if paf else "",
            paf.segundo_periodo_ferias.strftime(
                "%d/%m/%Y") if paf and paf.segundo_periodo_ferias else "",
            paf.fim_segundo_periodo.strftime(
                "%d/%m/%Y") if paf and paf.fim_segundo_periodo else "",
            paf.qtd_dias_terceiro_periodo if paf else "",
            paf.terceiro_periodo_ferias.strftime(
                "%d/%m/%Y") if paf and paf.terceiro_periodo_ferias else "",
            paf.fim_terceiro_periodo.strftime(
                "%d/%m/%Y") if paf and paf.fim_terceiro_periodo else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    registrar_log_download(
        nome_relatorio="Plano Anual de Férias (PAF Geral)",
        colunas_lista=colunas,  # Usa a própria variável 'colunas' que você já declarou lá em cima
        filtros_dict={"ano_referencia": ultimo_ano,
                      "escopo": "Toda a Corporação"}
    )

    response = make_response(output.read())
    response.headers[
        "Content-Disposition"] = f"attachment; filename=pafs_militares_{ultimo_ano}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
