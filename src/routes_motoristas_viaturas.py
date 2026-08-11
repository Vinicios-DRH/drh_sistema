import unicodedata

from flask import current_app
import io
import math
from zoneinfo import ZoneInfo
from flask_wtf.csrf import validate_csrf
from flask_login import login_required
from flask import abort, json, request, jsonify, make_response, current_app
from random import shuffle
import os
import zipfile
import qrcode
import pytz
import pandas as pd
import base64
import matplotlib.pyplot as plt
import requests
import urllib
from src.decorators.utils_acumulo import b2_bucket_name, b2_client, b2_delete_all_versions, b2_upload_fileobj
from src.identificacao import buscar_pessoa_por_cpf, normaliza_matricula
from src.formatar_cpf import cadete_restantes, formatar_cpf, get_militar_por_user, is_cadete
from flask import render_template, redirect, url_for, request, flash, jsonify, session, send_file, make_response, \
    Response, stream_with_context
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import validate_csrf, generate_csrf
from werkzeug.utils import secure_filename
from src import app, database, bcrypt
from src.forms import (AtualizacaoCadastralForm, ControleConvocacaoForm, CriarSenhaForm, FichaAlunosForm, FormEsqueciSenha, FormFiltroMilitar, FormMilitarInativo, FormResetarSenhaPublica, FormViatura,
                       IdentificacaoForm, ImpactoForm, FormLogin, FormMilitar, FormCriarUsuario, FormMotoristas, FormFiltroMotorista, LtsAlunoForm, RecompensaAlunoForm,
                       RestricaoAlunoForm, SancaoAlunoForm, TabelaVencimentoForm, InativarAlunoForm, MatriculaConfirmForm)
from src.models import (ControleConvocacao, Convocacao, DocumentoMilitar, EfetivoDiarioOBM, HistoricoEfetivoDiario, ImportacaoMilitarHistorico, LogAcesso, LtsAlunos, Militar, MilitaresInativos, NomeConvocado, PostoGrad, Quadro, Obm, Localidade, Funcao, RecompensaAluno, RestricaoAluno, SancaoAluno, SegundoVinculo, SituacaoConvocacao, User, FuncaoUser, PublicacaoBg,
                        EstadoCivil, Especialidade, Destino, Motivo, Modalidade, Punicao, Comportamento, MilitarObmFuncao,
                        FuncaoGratificada,
                        MilitaresAgregados, MilitaresADisposicao, LicencaEspecial, LicencaParaTratamentoDeSaude, Paf,
                        Meses, Motoristas, Categoria, TabelaVencimento, ValorDetalhadoPostoGrad, FichaAlunos, AlunoInativo, Viaturas, ViaturaMilitar, MilitarGraduacao, MilitarContatoEmergencia, MilitarConjuge, LogExportacaoExcel, Curso, MilitarCurso, AuditoriaAtualizacaoCadastral, now_manaus_naive)
from src.querys import dados_para_mapa, obter_estatisticas_militares, login_usuario
from src.decorators.control import checar_ocupacao, militar_esta_no_escopo, obms_permitidas_para_usuario, sync_user_admin_obms_from_militar
from src.decorators.business_logic import processar_militares_a_disposicao, processar_militares_agregados, \
    processar_militares_le, processar_militares_lts
from datetime import datetime, date, timedelta
from io import BytesIO
from sqlalchemy.orm import joinedload, selectinload, load_only, aliased
from sqlalchemy import case, distinct, extract, func, not_, or_, cast, String, and_
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP, getcontext
from docx import Document
from urllib.parse import urlencode
from collections import defaultdict, Counter
from docx.shared import Pt
from docx.oxml.ns import qn
from src.decorators.formatar_datas import formatar_data_extenso, formatar_data_sem_zero
from src.decorators.email_utils import send_reset_password_email, verify_password_reset_token
import re
import plotly.graph_objs as go
import plotly.io as pio
from collections import defaultdict
from src.utils.sa_serialize import sa_to_dict
from sqlalchemy.inspection import inspect as sa_inspect
from src.security.perms import has_perm
from src.authz import is_super_or_perm, can_ferias_bypass_janela, is_super, require_perm
from src.utils.cadastro_status import cadastro_esta_completo
from src.utils.painel import (
    _obter_obm_principal,
    listar_situacoes_atualizacao,
    obter_resumo_atualizacao_cadastral,
    obter_militares_atualizacao_cadastral,
    serializar_militar_atualizacao,
    listar_obms_atualizacao,
    listar_postos_grad_atualizacao,
    obter_detalhes_militar_atualizacao,)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.services.importar_militares import (
    ler_planilha,
    colunas_reconhecidas,
    colunas_nao_reconhecidas,
    analisar_importacao,
    importar_dataframe,
    salvar_historico_importacao,
)
from src.services.militar_situacao_service import (
    parse_date_flex,
    sincronizar_blocos_funcionais,
)
import time
from src.utils.utils import registrar_log_download
from dateutil.relativedelta import relativedelta



@app.route('/adicionar-motorista', methods=['GET', 'POST'])
@login_required
def adicionar_motorista():
    form_motorista = FormMotoristas()

    militares_query = (
        database.session.query(
            Militar.id,
            Militar.nome_completo,
            Militar.matricula,
            PostoGrad.sigla.label("posto_grad_sigla"),
            Obm.sigla.label("obm_sigla")
        )
        .outerjoin(PostoGrad, Militar.posto_grad_id == PostoGrad.id)
        .outerjoin(MilitarObmFuncao, (MilitarObmFuncao.militar_id == Militar.id) & (MilitarObmFuncao.data_fim == None))
        .outerjoin(Obm, MilitarObmFuncao.obm_id == Obm.id)
        .order_by(Militar.nome_completo)
        .all()
    )

    form_motorista.nome_completo.choices = [
        (militar.id, militar.nome_completo) for militar in militares_query if militar.id is not None
    ]

    militares = {
        militar.id: {
            'matricula': militar.matricula,
            'obm_id_1': militar.obm_sigla,
            'posto_grad_id': militar.posto_grad_sigla
        }
        for militar in militares_query
    }

    form_motorista.categoria_id.choices = [
        (0, '-- Selecione uma categoria --')
    ] + [(categoria.id, categoria.sigla) for categoria in Categoria.query.all()]

    if form_motorista.validate_on_submit():
        try:
            novo_motorista = Motoristas(
                militar_id=form_motorista.nome_completo.data,
                categoria_id=form_motorista.categoria_id.data or None,
                boletim_geral=form_motorista.boletim_geral.data,
                siged=form_motorista.siged.data,
                vencimento_cnh=form_motorista.vencimento_cnh.data,  # <-- corrigido
                usuario_id=current_user.id,
                desclassificar="NÃO",
                created=datetime.utcnow()
            )

            if form_motorista.cnh_imagem.data and form_motorista.cnh_imagem.data.filename != '':
                file = form_motorista.cnh_imagem.data
                ext = file.filename.split('.')[-1]
                timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')

                nome_militar = next(
                    (m.nome_completo for m in militares_query if m.id ==
                     form_motorista.nome_completo.data),
                    'motorista'
                ).replace(" ", "_")

                nome_arquivo = secure_filename(
                    f"{nome_militar}_cnh_{timestamp}.{ext}")
                file_bytes = file.read()

                app.supabase.storage.from_('motoristas').upload(
                    path=nome_arquivo,
                    file=file_bytes,
                    file_options={"content-type": file.mimetype}
                )

                novo_motorista.cnh_imagem = nome_arquivo

            database.session.add(novo_motorista)
            database.session.commit()
            flash('Motorista cadastrado com sucesso!', 'success')
            return redirect(url_for('adicionar_motorista'))

        except Exception as e:
            database.session.rollback()
            flash(f'Erro ao cadastrar motorista: {str(e)}', 'danger')

    return render_template('adicionar_motorista.html', form_motorista=form_motorista, militares=militares)


@app.route('/motoristas', methods=['GET', 'POST'])
@login_required
def motoristas():
    form_filtro = FormFiltroMotorista()

    form_filtro.obm_id.choices = [
        ('', '-- Selecione OBM --')
    ] + [(str(obm.id), obm.sigla) for obm in Obm.query.order_by(Obm.sigla).all()]

    form_filtro.posto_grad_id.choices = [
        ('', '-- Selecione Posto/Grad --')
    ] + [(str(posto.id), posto.sigla) for posto in PostoGrad.query.order_by(PostoGrad.sigla).all()]

    form_filtro.categoria_id.choices = [
        ('', '-- Selecione uma categoria --')
    ] + [(str(categoria.id), categoria.sigla) for categoria in Categoria.query.order_by(Categoria.sigla).all()]

    form_filtro.viatura_id.choices = [
        ('', '-- Selecione uma viatura --')
    ] + [
        (
            str(viatura.id),
            f"{viatura.prefixo or 'S/PREFIXO'} - {viatura.placa or 'S/PLACA'} - {viatura.marca_modelo or 'S/MODELO'}"
        )
        for viatura in Viaturas.query.order_by(Viaturas.prefixo.asc()).all()
    ]

    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str)
    obm_id = request.args.get('obm_id', '', type=str)
    posto_grad_id = request.args.get('posto_grad_id', '', type=str)
    categoria_id = request.args.get('categoria_id', '', type=str)
    viatura_id = request.args.get('viatura_id', '', type=str)

    # mantém os valores selecionados no form
    form_filtro.obm_id.data = obm_id
    form_filtro.posto_grad_id.data = posto_grad_id
    form_filtro.categoria_id.data = categoria_id
    form_filtro.viatura_id.data = viatura_id

    query = (
        Motoristas.query
        .join(Militar)
        .options(
            joinedload(Motoristas.militar).joinedload(Militar.posto_grad),
            joinedload(Motoristas.militar).joinedload(Militar.quadro),
            joinedload(Motoristas.categoria),
        )
        .filter(
            or_(
                Motoristas.desclassificar.is_(None),
                Motoristas.desclassificar != 'SIM'
            )
        )
    )

    # Filtro por OBM
    if obm_id:
        subquery = (
            MilitarObmFuncao.query
            .filter_by(obm_id=obm_id)
            .with_entities(MilitarObmFuncao.militar_id)
        )
        query = query.filter(Motoristas.militar_id.in_(subquery))

    # Filtro por Posto/Graduação
    if posto_grad_id:
        query = query.filter(Militar.posto_grad_id == posto_grad_id)

    # Filtro por Categoria
    if categoria_id:
        query = query.filter(Motoristas.categoria_id == categoria_id)

    # Filtro por Viatura
    if viatura_id:
        subquery_viatura = (
            ViaturaMilitar.query
            .filter_by(viatura_id=viatura_id)
            .with_entities(ViaturaMilitar.militar_id)
        )
        query = query.filter(Motoristas.militar_id.in_(subquery_viatura))

    # Filtro por Nome
    if search:
        query = query.filter(Militar.nome_completo.ilike(f'%{search}%'))

    # Apenas registros ativos
    query = query.filter(Motoristas.modified.is_(None))

    # Evita duplicidade caso futuramente haja joins extras
    # query = query.distinct(Motoristas.id)

    # Paginação
    motoristas_paginados = query.order_by(
        Militar.nome_completo.asc()
    ).paginate(page=page, per_page=per_page)

    # Contagem de militares e motoristas válidos
    total_militares = Militar.query.count()
    total_motoristas = Motoristas.query.filter(
        Motoristas.modified.is_(None),
        or_(
            Motoristas.desclassificar.is_(None),
            Motoristas.desclassificar != 'SIM'
        )
    ).count()

    # Gráfico: Percentual de militares que são motoristas
    labels_motoristas = ['Motoristas', 'Não são motoristas']
    values_motoristas = [total_motoristas, total_militares - total_motoristas]
    fig_motoristas = go.Figure(
        data=[go.Pie(labels=labels_motoristas,
                     values=values_motoristas, hole=0.4)]
    )
    grafico_motoristas = pio.to_json(fig_motoristas)

    # Gráfico: Motoristas por categoria
    categorias = database.session.query(
        Categoria.sigla,
        database.func.count(Motoristas.id)
    ).join(Motoristas).filter(
        Motoristas.modified.is_(None),
        or_(
            Motoristas.desclassificar.is_(None),
            Motoristas.desclassificar != 'SIM'
        )
    ).group_by(Categoria.sigla).all()

    labels_categorias = [c[0] for c in categorias]
    values_categorias = [c[1] for c in categorias]
    fig_categorias = go.Figure(
        data=[go.Pie(labels=labels_categorias,
                     values=values_categorias, hole=0.4)]
    )
    grafico_categorias = pio.to_json(fig_categorias)

    # Gráfico: Motoristas por OBM
    obms = database.session.query(
        Obm.sigla,
        database.func.count(Motoristas.id)
    ).join(
        MilitarObmFuncao, Obm.id == MilitarObmFuncao.obm_id
    ).join(
        Motoristas, MilitarObmFuncao.militar_id == Motoristas.militar_id
    ).filter(
        Motoristas.modified.is_(None),
        or_(
            Motoristas.desclassificar.is_(None),
            Motoristas.desclassificar != 'SIM'
        )
    ).group_by(Obm.sigla).all()

    labels_obms = [obm[0] for obm in obms]
    values_obms = [obm[1] for obm in obms]
    fig_obms = go.Figure(
        data=[go.Pie(labels=labels_obms, values=values_obms, hole=0.4)]
    )
    grafico_obms = pio.to_json(fig_obms)

    return render_template(
        'motoristas.html',
        motoristas=motoristas_paginados,
        search=search,
        form_filtro=form_filtro,
        grafico_motoristas=grafico_motoristas,
        grafico_categorias=grafico_categorias,
        grafico_obms=grafico_obms
    )


@app.route('/viaturas/nova', methods=['GET', 'POST'])
@login_required
def cadastrar_viatura():
    form = FormViatura()

    form.obm_id.choices = [('', '-- Selecione a OBM --')] + [
        (str(obm.id), obm.sigla) for obm in Obm.query.order_by(Obm.sigla.asc()).all()
    ]

    if form.validate_on_submit():
        prefixo = (form.prefixo.data or '').strip().upper()
        placa = (form.placa.data or '').strip().upper()
        marca_modelo = (form.marca_modelo.data or '').strip()
        obm_id = form.obm_id.data

        viatura_existente_prefixo = Viaturas.query.filter_by(
            prefixo=prefixo).first()
        if viatura_existente_prefixo:
            flash('Já existe uma viatura cadastrada com esse prefixo.', 'warning')
            return render_template('viaturas/cadastrar_viatura.html', form=form)

        if placa:
            viatura_existente_placa = Viaturas.query.filter_by(
                placa=placa).first()
            if viatura_existente_placa:
                flash('Já existe uma viatura cadastrada com essa placa.', 'warning')
                return render_template('viaturas/cadastrar_viatura.html', form=form)

        nova_viatura = Viaturas(
            prefixo=prefixo,
            placa=placa,
            marca_modelo=marca_modelo,
            obm_id=int(obm_id) if obm_id else None
        )

        database.session.add(nova_viatura)
        database.session.commit()

        flash('Viatura cadastrada com sucesso.', 'success')
        return redirect(url_for('listar_viaturas'))

    return render_template('viaturas/cadastrar_viatura.html', form=form)


@app.route('/viaturas-lista')
@login_required
def listar_viaturas():
    search = request.args.get('search', '', type=str).strip()

    query = Viaturas.query.options(joinedload(Viaturas.obm))

    if search:
        query = query.filter(
            database.or_(
                Viaturas.prefixo.ilike(f'%{search}%'),
                Viaturas.placa.ilike(f'%{search}%'),
                Viaturas.marca_modelo.ilike(f'%{search}%')
            )
        )

    viaturas = query.order_by(Viaturas.prefixo.asc()).all()

    return render_template('viaturas/listar_viaturas.html', viaturas=viaturas, search=search)


@app.route('/atualizar-motorista/<int:motorista_id>', methods=['GET', 'POST'])
@login_required
def atualizar_motorista(motorista_id):
    motorista = Motoristas.query.get_or_404(motorista_id)

    form_motorista = FormMotoristas(obj=motorista)

    militar_atual = (motorista.militar.id, motorista.militar.nome_completo)
    form_motorista.nome_completo.choices = [militar_atual]
    form_motorista.nome_completo.data = motorista.militar.id

    form_motorista.categoria_id.choices = [
        (categoria.id, categoria.sigla) for categoria in Categoria.query.all()
    ]
    form_motorista.categoria_id.data = motorista.categoria_id

    form_motorista.matricula.data = motorista.militar.matricula
    form_motorista.posto_grad_id.data = (
        motorista.militar.posto_grad.sigla if motorista.militar.posto_grad else None
    )
    form_motorista.obm_id_1.data = (
        motorista.militar.obm_funcoes[0].obm.sigla if motorista.militar.obm_funcoes else None
    )

    if request.method == 'POST':
        if request.form.get('action') == 'desclassificar':
            try:
                # fecha o registro atual
                motorista.modified = datetime.utcnow()

                # cria o novo registro como DESCLASSIFICADO
                novo_motorista = Motoristas(
                    militar_id=motorista.militar_id,
                    categoria_id=motorista.categoria_id,
                    boletim_geral=motorista.boletim_geral,
                    siged=motorista.siged,
                    usuario_id=motorista.usuario_id,
                    vencimento_cnh=motorista.vencimento_cnh,
                    cnh_imagem=motorista.cnh_imagem,
                    created=datetime.utcnow(),
                    desclassificar='SIM',
                    desclassificar_por=current_user.id,
                    desclassificar_em=datetime.utcnow()
                )

                database.session.add(novo_motorista)
                database.session.commit()

                flash('Motorista desclassificado com sucesso!', 'warning')
                return redirect(url_for('motoristas_desclassificados'))

            except Exception as e:
                database.session.rollback()
                flash(f'Erro ao desclassificar motorista: {str(e)}', 'danger')

        if form_motorista.validate_on_submit():
            try:
                motorista.modified = datetime.utcnow()

                # cria novo registro ATIVO/ATUALIZADO
                novo_motorista = Motoristas(
                    militar_id=motorista.militar_id,
                    categoria_id=form_motorista.categoria_id.data,
                    boletim_geral=form_motorista.boletim_geral.data,
                    siged=form_motorista.siged.data,
                    usuario_id=current_user.id,
                    vencimento_cnh=form_motorista.vencimento_cnh.data,
                    created=datetime.utcnow(),
                    desclassificar='NÃO',
                    desclassificar_por=None,
                    desclassificar_em=None
                )

                if form_motorista.cnh_imagem.data and form_motorista.cnh_imagem.data.filename != '':
                    file = form_motorista.cnh_imagem.data
                    ext = file.filename.split('.')[-1]
                    timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
                    nome_formatado = motorista.militar.nome_completo.replace(
                        " ", "_")
                    nome_arquivo = secure_filename(
                        f"{nome_formatado}_cnh_{timestamp}.{ext}")
                    file_bytes = file.read()

                    app.supabase.storage.from_('motoristas').upload(
                        path=nome_arquivo,
                        file=file_bytes,
                        file_options={"content-type": file.mimetype}
                    )

                    novo_motorista.cnh_imagem = nome_arquivo
                else:
                    novo_motorista.cnh_imagem = motorista.cnh_imagem

                database.session.add(novo_motorista)
                database.session.commit()

                flash('Motorista atualizado com sucesso!', 'success')
                return redirect(url_for('motoristas'))

            except Exception as e:
                database.session.rollback()
                flash(f'Erro ao atualizar motorista: {str(e)}', 'danger')

    return render_template(
        'atualizar_motorista.html',
        form_motorista=form_motorista,
        motorista=motorista
    )


@app.route('/motoristas-desclassificados', methods=['GET', 'POST'])
@login_required
def motoristas_desclassificados():
    form_filtro = FormFiltroMotorista()

    form_filtro.obm_id.choices = [
        ('', '-- Selecione OBM --')] + [(obm.id, obm.sigla) for obm in Obm.query.all()]
    form_filtro.posto_grad_id.choices = [('', '-- Selecione Posto/Grad --')] + [
        (posto.id, posto.sigla) for posto in PostoGrad.query.all()]
    form_filtro.categoria_id.choices = [('', '-- Selecione uma categoria --')] + [(
        categoria.id, categoria.sigla) for categoria in Categoria.query.all()]

    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str)
    obm_id = request.args.get('obm_id', '', type=str)
    posto_grad_id = request.args.get('posto_grad_id', '', type=str)
    categoria_id = request.args.get('categoria_id', '', type=str)

    # Query base: apenas desclassificados atuais (registros não modificados e desclassificar == 'SIM')
    query = Motoristas.query.join(Militar).filter(
        Motoristas.desclassificar == 'SIM', Motoristas.modified.is_(None))

    # Filtro por OBM
    if obm_id:
        subquery = MilitarObmFuncao.query.filter_by(
            obm_id=obm_id).with_entities(MilitarObmFuncao.militar_id)
        query = query.filter(Motoristas.militar_id.in_(subquery))

    # Filtro por Posto/Graduação
    if posto_grad_id:
        query = query.filter(Militar.posto_grad_id == posto_grad_id)

    # Filtro por Categoria
    if categoria_id:
        query = query.filter(Motoristas.categoria_id == categoria_id)

    # Filtro por Nome
    if search:
        query = query.filter(Militar.nome_completo.ilike(f'%{search}%'))

    # Paginação
    motoristas_paginados = query.order_by(Motoristas.desclassificar_em.desc(
    ).nullslast(), Militar.nome_completo.asc()).paginate(page=page, per_page=per_page)

    # Contagem total de desclassificados (para resumo)
    total_desclassificados = Motoristas.query.filter(
        Motoristas.desclassificar == 'SIM', Motoristas.modified.is_(None)).count()

    # Gráfico 1: Desclassificados por categoria
    categorias = database.session.query(
        Categoria.sigla,
        database.func.count(Motoristas.id)
    ).join(Motoristas).filter(Motoristas.desclassificar == 'SIM', Motoristas.modified.is_(None)).group_by(Categoria.sigla).all()
    labels_categorias = [c[0] for c in categorias]
    values_categorias = [c[1] for c in categorias]
    fig_categorias = go.Figure(
        data=[go.Pie(labels=labels_categorias, values=values_categorias, hole=0.4)])
    grafico_categorias = pio.to_json(fig_categorias)

    # Gráfico 2: Desclassificados por OBM
    obms = database.session.query(
        Obm.sigla,
        database.func.count(Motoristas.id)
    ).join(MilitarObmFuncao, Obm.id == MilitarObmFuncao.obm_id).join(
        Motoristas, MilitarObmFuncao.militar_id == Motoristas.militar_id
    ).filter(Motoristas.desclassificar == 'SIM', Motoristas.modified.is_(None)).group_by(Obm.sigla).all()
    labels_obms = [o[0] for o in obms]
    values_obms = [o[1] for o in obms]
    fig_obms = go.Figure(
        data=[go.Pie(labels=labels_obms, values=values_obms, hole=0.4)])
    grafico_obms = pio.to_json(fig_obms)

    # Gráfico 3: Evolução mensal de desclassificados (últimos 12 meses)
    # Usa date_trunc para agrupar por mês; funciona em PostgreSQL (Supabase)
    mensal = database.session.query(
        database.func.date_trunc(
            'month', Motoristas.desclassificar_em).label('mes'),
        database.func.count(Motoristas.id)
    ).filter(
        Motoristas.desclassificar == 'SIM',
        Motoristas.modified.is_(None),
        Motoristas.desclassificar_em.isnot(None)
    ).group_by('mes').order_by('mes').all()

    meses = [row[0].strftime('%Y-%m') for row in mensal] if mensal else []
    valores_mensais = [row[1] for row in mensal] if mensal else []
    fig_mensal = go.Figure(data=[go.Bar(x=meses, y=valores_mensais)])
    fig_mensal.update_layout(xaxis_title='Mês', yaxis_title='Desclassificados')
    grafico_mensal = pio.to_json(fig_mensal)

    return render_template(
        'motoristas_desclassificados.html',
        motoristas=motoristas_paginados,
        search=search,
        form_filtro=form_filtro,
        total_desclassificados=total_desclassificados,
        grafico_categorias=grafico_categorias,
        grafico_obms=grafico_obms,
        grafico_mensal=grafico_mensal
    )


@app.route("/viaturas", methods=["GET"])
def escolher_obm():
    obms = Obm.query.order_by(Obm.sigla.asc()).all()
    return render_template("viaturas_escolher_obm.html", obms=obms)


@app.route("/<int:obm_id>/viaturas", methods=["GET"])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER', 'DRH', 'CHEFE DRH')
def gerenciar_viaturas(obm_id):
    obm = Obm.query.get_or_404(obm_id)

    # Viaturas já dessa OBM
    viaturas_da_obm = (Viaturas.query
                       .filter(Viaturas.obm_id == obm_id)
                       .order_by(Viaturas.prefixo.asc(), Viaturas.placa.asc())
                       .all())

    # Viaturas sem OBM (ou de outra OBM)
    viaturas_sem_obm = (Viaturas.query
                        .filter(Viaturas.obm_id.is_(None))
                        .order_by(Viaturas.prefixo.asc(), Viaturas.placa.asc())
                        .all())

    # Verifica se é o Chefe do CSM (obm_id_1 == 32)
    is_chefe_csm = (current_user.obm_id_1 == 32)

    motoristas = []
    motoristas_por_viatura = {}

    # Só carrega os motoristas se NÃO for o Chefe do CSM
    if not is_chefe_csm:
        motoristas = (database.session.query(Motoristas)
                      .join(Militar, Motoristas.militar_id == Militar.id)
                      .join(MilitarObmFuncao, MilitarObmFuncao.militar_id == Militar.id)
                      .filter(
                          MilitarObmFuncao.obm_id == obm_id,
                          MilitarObmFuncao.data_fim.is_(None),
                          Motoristas.modified.is_(None),
                          or_(
                              Motoristas.desclassificar.is_(None),
                              Motoristas.desclassificar != 'SIM'
                          )
                      )
                      .order_by(Militar.nome_completo.asc())
                      .all())

        for v in viaturas_da_obm:
            vms = (ViaturaMilitar.query
                   .filter_by(viatura_id=v.id)
                   .all())
            motoristas_por_viatura[v.id] = [vm.militar_id for vm in vms]

    return render_template(
        "viaturas_gerenciar.html",
        obm=obm,
        viaturas_da_obm=viaturas_da_obm,
        viaturas_sem_obm=viaturas_sem_obm,
        motoristas=motoristas,
        motoristas_por_viatura=motoristas_por_viatura,
        is_chefe_csm=is_chefe_csm # Passamos a variável para o template
    )


@app.route("/<int:obm_id>/viaturas/atribuir", methods=["POST"])
def atribuir_viaturas_obm(obm_id):
    """
    Recebe a lista 'assigned_ids[]' (viaturas que devem ficar na OBM).
    Vamos:
      - Setar obm_id = obm_id para as IDs enviadas
      - Remover desta OBM (setar NULL) todas as que estavam e não vieram no POST
    """
    obm = Obm.query.get_or_404(obm_id)
    ids_enviados = request.form.getlist("assigned_ids[]")
    ids_enviados = [int(x) for x in ids_enviados]

    # Viaturas que hoje estão nessa OBM
    atuais = Viaturas.query.filter_by(obm_id=obm_id).all()
    ids_atuais = {v.id for v in atuais}

    # 1) adicionar/mover para esta OBM os enviados que não estão
    if ids_enviados:
        (Viaturas.query
         .filter(Viaturas.id.in_(ids_enviados))
         .update({Viaturas.obm_id: obm_id}, synchronize_session=False))

    # 2) remover desta OBM os que estavam e não estão mais na lista
    ids_remover = list(ids_atuais - set(ids_enviados))
    if ids_remover:
        (Viaturas.query
         .filter(Viaturas.id.in_(ids_remover))
         .update({Viaturas.obm_id: None}, synchronize_session=False))

    database.session.commit()
    flash("Atribuições de viaturas atualizadas para a OBM {}.".format(
        obm.sigla), "success")
    return redirect(url_for("gerenciar_viaturas", obm_id=obm_id))


@app.route("/viaturas/<int:viatura_id>/motoristas", methods=["POST"])
def salvar_motoristas_viatura(viatura_id):
    """
    Salva até 5 motoristas (militar_id) para a viatura.
    O gatilho no banco impede >5, mas também checamos aqui para UX.
    """
    v = Viaturas.query.get_or_404(viatura_id)
    selecionados = request.form.getlist("motoristas[]")
    selecionados = [int(x) for x in selecionados if x]

    if len(selecionados) > 5:
        flash("Selecione no máximo 5 motoristas para a viatura.", "warning")
        return redirect(url_for("viaturas_admin.gerenciar_viaturas", obm_id=v.obm_id or 0))

    # Atualiza o conjunto: remove os que saíram e adiciona os novos
    atuais = ViaturaMilitar.query.filter_by(viatura_id=viatura_id).all()
    ids_atuais = {vm.militar_id for vm in atuais}
    novos = set(selecionados) - ids_atuais
    remover = ids_atuais - set(selecionados)

    if remover:
        (ViaturaMilitar.query
            .filter(ViaturaMilitar.viatura_id == viatura_id,
                    ViaturaMilitar.militar_id.in_(remover))
            .delete(synchronize_session=False))

    for mid in novos:
        database.session.add(ViaturaMilitar(
            viatura_id=viatura_id, militar_id=mid))

    try:
        database.session.commit()
        flash("Motoristas atualizados para a viatura {}.".format(
            v.prefixo or v.placa), "success")
    except IntegrityError:
        database.session.rollback()
        flash("Não foi possível salvar: limite atingido ou motorista duplicado.", "danger")

    return redirect(url_for("gerenciar_viaturas", obm_id=v.obm_id or 0))


@app.route("/motoristas/exportar-excel", methods=["GET"])
@login_required
def exportar_motoristas_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Motoristas Classificados"

        headers = [
            "ID Militar",
            "Nome Completo",
            "Nome de Guerra",
            "CPF",
            "Matrícula",
            "Posto/Grad",
            "Quadro",
            "OBM",
            "Categoria CNH",
            "Vencimento CNH",
            "SIGED",
            "Boletim Geral",
            "Qtd. Viaturas da OBM",
            "Viaturas da OBM",
            "Criado Em",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="B5121B")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # SOMENTE classificados atuais:
        # - modified IS NULL
        # - desclassificar diferente de 'SIM' OU NULL
        subquery = (
            database.session.query(
                Motoristas.militar_id,
                func.max(Motoristas.id).label("ultimo_motorista_id")
            )
            .filter(
                Motoristas.militar_id.isnot(None),
                Motoristas.modified.is_(None),
                or_(
                    Motoristas.desclassificar.is_(None),
                    func.upper(Motoristas.desclassificar) != "SIM"
                )
            )
            .group_by(Motoristas.militar_id)
            .subquery()
        )

        motoristas = (
            Motoristas.query
            .join(subquery, Motoristas.id == subquery.c.ultimo_motorista_id)
            .options(
                joinedload(Motoristas.militar).joinedload(Militar.posto_grad),
                joinedload(Motoristas.militar).joinedload(Militar.quadro),
                joinedload(Motoristas.militar).selectinload(
                    Militar.obm_funcoes
                ).joinedload(MilitarObmFuncao.obm).selectinload(Obm.viaturas_obm),
                joinedload(Motoristas.categoria),
            )
            .order_by(Motoristas.id.asc())
            .all()
        )

        for motorista in motoristas:
            militar = motorista.militar

            militar_id = militar.id if militar else ""
            nome_completo = militar.nome_completo if militar else ""
            nome_guerra = militar.nome_guerra if militar else ""
            cpf = militar.cpf if militar else ""
            matricula = militar.matricula if militar else ""
            posto_grad = militar.posto_grad.sigla if militar and militar.posto_grad else ""
            quadro = militar.quadro.quadro if militar and militar.quadro else ""
            categoria = motorista.categoria.sigla if motorista.categoria else ""

            vencimento_cnh = ""
            if motorista.vencimento_cnh:
                vencimento_cnh = motorista.vencimento_cnh.strftime("%d/%m/%Y")

            criado_em = ""
            if motorista.created:
                criado_em = motorista.created.strftime("%d/%m/%Y %H:%M:%S")

            obm_sigla = ""
            obm_atual = None

            if militar and militar.obm_funcoes:
                obms_ativas = [
                    rel for rel in militar.obm_funcoes
                    if rel.data_fim is None and rel.obm
                ]
                if obms_ativas:
                    obm_atual = obms_ativas[0].obm
                else:
                    relacoes_com_obm = [
                        rel for rel in militar.obm_funcoes if rel.obm
                    ]
                    if relacoes_com_obm:
                        relacoes_com_obm.sort(
                            key=lambda x: x.data_criacao or datetime.min,
                            reverse=True
                        )
                        obm_atual = relacoes_com_obm[0].obm

            if obm_atual:
                obm_sigla = obm_atual.sigla or ""

            lista_viaturas = []
            if obm_atual and obm_atual.viaturas_obm:
                for viatura in obm_atual.viaturas_obm:
                    texto = " - ".join(
                        filter(
                            None,
                            [
                                viatura.prefixo,
                                viatura.placa,
                                viatura.marca_modelo,
                            ]
                        )
                    )
                    if texto:
                        lista_viaturas.append(texto)

                lista_viaturas = list(dict.fromkeys(lista_viaturas))

            qtd_viaturas_obm = len(lista_viaturas)
            viaturas_obm = ", ".join(lista_viaturas)

            ws.append([
                militar_id,
                nome_completo,
                nome_guerra,
                cpf,
                matricula,
                posto_grad,
                quadro,
                obm_sigla,
                categoria,
                vencimento_cnh,
                motorista.siged or "",
                motorista.boletim_geral or "",
                qtd_viaturas_obm,
                viaturas_obm,
                criado_em,
            ])

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                try:
                    cell_value = str(
                        cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[get_column_letter(
                column)].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"motoristas_classificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        registrar_log_download(
            nome_relatorio="Lista de Motoristas Classificados",
            colunas_lista=headers,
            filtros_dict={"status": "Apenas Classificados Ativos"}
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return {"ok": False, "erro": str(e)}, 500


@app.route('/listar-cnhs')
@login_required
def listar_cnhs():
    # Lista do root do bucket motoristas (com path vazio!)
    arquivos = app.supabase.storage.from_('motoristas').list('')

    print("🟡 Arquivos retornados do Supabase:")
    if arquivos:
        for item in arquivos:
            print(" -", item['name'])
    else:
        print("⚠️ Nenhum arquivo retornado!")

    # Garante que estamos pegando só os arquivos (não pastas)
    nomes_arquivos = [item['name']
                      for item in arquivos if item['name'] and not item['name'].endswith('/')]

    return render_template('listar_cnhs.html', arquivos=nomes_arquivos)
