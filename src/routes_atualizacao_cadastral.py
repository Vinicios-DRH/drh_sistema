from datetime import datetime
from zoneinfo import ZoneInfo
import unicodedata
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from src.authz import require_perm
from src import database
from src.models import (
    Militar,
    EstadoCivil,
    MilitarConjuge,
    MilitarContatoEmergencia,
    MilitarGraduacao,
    PostoGrad,
    AuditoriaAtualizacaoCadastral,
    MilitarConferenciaCadastral,
    MilitarObmFuncao,
    Obm,
    User,  # se tua model existir no mesmo arquivo
)
from src.forms import AtualizacaoCadastralForm
from src.utils.cadastro_status import (
    cadastro_esta_completo,
    get_campos_pendentes_cadastro,
)
from io import BytesIO
from flask import send_file
from sqlalchemy import or_, and_, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas


# Se tu tiver decorator de permissão, descomenta
# from src.authz import require_perm


bp_atualizacao_cadastral = Blueprint(
    "atualizacao_cadastral",
    __name__,
    url_prefix="/atualizacao-cadastral"
)


def _eh_al_sd(militar):
    if not militar or not militar.posto_grad:
        return False

    sigla = (militar.posto_grad.sigla or "").strip().upper()
    return sigla in {"AL SD", "ALUNO SOLDADO", "AL SD BM"}


MANAUS_TZ = ZoneInfo("America/Manaus")


def agora_manaus():
    return datetime.now(MANAUS_TZ)


def _normalizar_texto(valor: str) -> str:
    if not valor:
        return ""
    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(c for c in valor if not unicodedata.combining(c))
    return " ".join(valor.strip().upper().split())


def _get_client_ip():
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr


def _get_militar_do_usuario():
    if not current_user.is_authenticated:
        return None

    if not current_user.militar_id:
        return None

    return Militar.query.get(current_user.militar_id)


def _registrar_auditoria(militar_id, observacao="Militar realizou atualização cadastral."):
    auditoria = AuditoriaAtualizacaoCadastral(
        militar_id=militar_id,
        user_id=current_user.id,
        acao="ATUALIZACAO_CADASTRAL",
        ip_address=_get_client_ip(),
        user_agent=request.headers.get("User-Agent"),
        criado_em=agora_manaus(),
        observacao=observacao
    )
    database.session.add(auditoria)


@bp_atualizacao_cadastral.route("/", methods=["GET", "POST"])
@login_required
def atualizar():
    militar = _get_militar_do_usuario()
    eh_al_sd = _eh_al_sd(militar)

    if not militar:
        flash("Seu usuário não está vinculado a um cadastro de militar.", "danger")
        return redirect(url_for("home"))

    form = AtualizacaoCadastralForm()

    form.estado_civil.choices = [("", "Selecione")] + [
        (str(item.id), item.estado)
        for item in EstadoCivil.query.order_by(EstadoCivil.estado.asc()).all()
    ]

    graduacoes = MilitarGraduacao.query.filter_by(
        militar_id=militar.id
    ).order_by(MilitarGraduacao.id.asc()).all()

    contatos_emergencia = MilitarContatoEmergencia.query.filter_by(
        militar_id=militar.id
    ).order_by(MilitarContatoEmergencia.id.asc()).all()

    conjuge = MilitarConjuge.query.filter_by(militar_id=militar.id).first()

    if form.validate_on_submit():
        try:
            # ===== Campos principais =====
            militar.grau_instrucao = form.grau_instrucao.data or None
            militar.graduacao = form.graduacao.data or None
            militar.pos_graduacao = form.pos_graduacao.data or None
            militar.mestrado = form.mestrado.data or None
            militar.doutorado = form.doutorado.data or None
            militar.raca = form.raca.data or None

            militar.nome_pai = form.nome_pai.data or None
            militar.nome_mae = form.nome_mae.data or None
            militar.estado_civil = int(
                form.estado_civil.data) if form.estado_civil.data else None
            militar.data_nascimento = form.data_nascimento.data

            militar.endereco = form.endereco.data or None
            militar.complemento = form.complemento.data or None
            militar.cidade = form.cidade.data or None
            militar.estado = form.estado.data or None
            militar.cep = form.cep.data or None
            militar.celular = form.celular.data or None
            militar.email = form.email.data or None

            # ===== Novos campos =====
            militar.local_nascimento = form.local_nascimento.data or None
            militar.altura = form.altura.data
            militar.cor_olhos = form.cor_olhos.data or None
            militar.cor_cabelos = form.cor_cabelos.data or None
            militar.bigode = bool(form.bigode.data)

            militar.medida_cabeca = form.medida_cabeca.data or None
            militar.numero_sapato = form.numero_sapato.data or None
            militar.medida_calca = form.medida_calca.data or None
            militar.medida_camisa = form.medida_camisa.data or None

            militar.tipo_sanguineo = form.tipo_sanguineo.data or None
            militar.sinais_particulares = form.sinais_particulares.data or None

            militar.tatuagem = bool(form.tatuagem.data)
            militar.local_tatuagem = (
                form.local_tatuagem.data or None if form.tatuagem.data else None
            )

            # ===== Graduações múltiplas =====
            graduacoes_curso = request.form.getlist("graduacoes_curso[]")
            graduacoes_instituicao = request.form.getlist(
                "graduacoes_instituicao[]")
            graduacoes_ano = request.form.getlist("graduacoes_ano[]")

            MilitarGraduacao.query.filter_by(militar_id=militar.id).delete()

            # ===== Campos extras somente para AL SD =====
            if eh_al_sd:
                militar.sexo = (request.form.get("sexo") or "").strip() or None
                militar.pis_pasep = (request.form.get(
                    "pis_pasep") or "").strip() or None
                militar.num_titulo_eleitor = (request.form.get(
                    "num_titulo_eleitor") or "").strip() or None
                militar.digito_titulo_eleitor = (request.form.get(
                    "digito_titulo_eleitor") or "").strip() or None
                militar.zona = (request.form.get("zona") or "").strip() or None
                militar.secao = (request.form.get(
                    "secao") or "").strip() or None

            for i, curso in enumerate(graduacoes_curso):
                curso = (curso or "").strip()
                if not curso:
                    continue

                instituicao = (graduacoes_instituicao[i] or "").strip(
                ) if i < len(graduacoes_instituicao) else ""
                ano_raw = (graduacoes_ano[i] or "").strip(
                ) if i < len(graduacoes_ano) else ""

                nova_graduacao = MilitarGraduacao(
                    militar_id=militar.id,
                    curso=curso,
                    instituicao=instituicao or None,
                    ano_conclusao=int(ano_raw) if ano_raw.isdigit() else None,
                    criado_em=agora_manaus()
                )
                database.session.add(nova_graduacao)

            # ===== Contatos de emergência =====
            contato_nome = request.form.getlist("contato_nome[]")
            contato_parentesco = request.form.getlist("contato_parentesco[]")
            contato_telefone = request.form.getlist("contato_telefone[]")
            contato_telefone_secundario = request.form.getlist(
                "contato_telefone_secundario[]")
            contato_observacao = request.form.getlist("contato_observacao[]")

            MilitarContatoEmergencia.query.filter_by(
                militar_id=militar.id).delete()

            for i, nome in enumerate(contato_nome):
                nome = (nome or "").strip()
                telefone = (contato_telefone[i] or "").strip(
                ) if i < len(contato_telefone) else ""

                if not nome or not telefone:
                    continue

                novo_contato = MilitarContatoEmergencia(
                    militar_id=militar.id,
                    nome=nome,
                    parentesco=(contato_parentesco[i] or "").strip(
                    ) if i < len(contato_parentesco) else None,
                    telefone=telefone,
                    telefone_secundario=(contato_telefone_secundario[i] or "").strip(
                    ) if i < len(contato_telefone_secundario) else None,
                    observacao=(contato_observacao[i] or "").strip(
                    ) if i < len(contato_observacao) else None,
                    criado_em=agora_manaus()
                )
                database.session.add(novo_contato)

            # ===== Cônjuge =====
            estado_civil_obj = EstadoCivil.query.get(
                militar.estado_civil) if militar.estado_civil else None
            estado_nome = _normalizar_texto(
                estado_civil_obj.estado if estado_civil_obj else "")

            tem_conjuge = (
                "CASADO" in estado_nome or
                "CASADA" in estado_nome or
                "UNIAO ESTAVEL" in estado_nome
            )

            conjuge_nome = (request.form.get("conjuge_nome") or "").strip()
            conjuge_cpf = (request.form.get("conjuge_cpf") or "").strip()
            conjuge_telefone = (request.form.get(
                "conjuge_telefone") or "").strip()
            conjuge_data_nascimento = (request.form.get(
                "conjuge_data_nascimento") or "").strip()
            conjuge_endereco = (request.form.get(
                "conjuge_endereco") or "").strip()
            conjuge_observacao = (request.form.get(
                "conjuge_observacao") or "").strip()

            conjuge = MilitarConjuge.query.filter_by(
                militar_id=militar.id).first()

            if tem_conjuge and conjuge_nome:
                if not conjuge:
                    conjuge = MilitarConjuge(
                        militar_id=militar.id,
                        criado_em=agora_manaus()
                    )
                    database.session.add(conjuge)

                conjuge.nome = conjuge_nome
                conjuge.cpf = conjuge_cpf or None
                conjuge.telefone = conjuge_telefone or None
                conjuge.data_nascimento = (
                    datetime.strptime(
                        conjuge_data_nascimento, "%Y-%m-%d").date()
                    if conjuge_data_nascimento else None
                )
                conjuge.endereco = conjuge_endereco or None
                conjuge.observacao = conjuge_observacao or None
                conjuge.atualizado_em = agora_manaus()

            elif conjuge:
                database.session.delete(conjuge)

            # ===== Auditoria / controle =====
            militar.atualizacao_cadastral_em = agora_manaus()
            militar.ip_address = _get_client_ip()

            _registrar_auditoria(
                militar_id=militar.id,
                observacao="Militar realizou atualização cadastral no próprio perfil."
            )

            database.session.flush()

            pendentes = get_campos_pendentes_cadastro(militar) or []
            militar.cadastro_atualizado = len(pendentes) == 0

            database.session.commit()
            database.session.refresh(militar)

            pendentes = get_campos_pendentes_cadastro(militar)

            if militar.cadastro_atualizado:
                flash(
                    "Atualização cadastral salva com sucesso. Seu cadastro está completo.",
                    "success"
                )
            else:
                flash(
                    f"Atualização salva, mas ainda existem {len(pendentes)} campo(s) pendente(s).",
                    "warning"
                )

            return redirect(url_for("atualizacao_cadastral.atualizar"))

        except Exception as e:
            database.session.rollback()
            flash(f"Erro ao salvar atualização cadastral: {str(e)}", "danger")

    elif request.method == "GET":
        form.grau_instrucao.data = militar.grau_instrucao
        form.graduacao.data = militar.graduacao
        form.pos_graduacao.data = militar.pos_graduacao
        form.mestrado.data = militar.mestrado
        form.doutorado.data = militar.doutorado
        form.raca.data = militar.raca

        form.nome_pai.data = militar.nome_pai
        form.nome_mae.data = militar.nome_mae
        form.estado_civil.data = str(
            militar.estado_civil) if militar.estado_civil else ""
        form.data_nascimento.data = militar.data_nascimento

        form.endereco.data = militar.endereco
        form.complemento.data = militar.complemento
        form.cidade.data = militar.cidade
        form.estado.data = militar.estado
        form.cep.data = militar.cep
        form.celular.data = militar.celular
        form.email.data = militar.email

        form.local_nascimento.data = militar.local_nascimento
        form.altura.data = militar.altura
        form.cor_olhos.data = militar.cor_olhos
        form.cor_cabelos.data = militar.cor_cabelos
        form.bigode.data = militar.bigode

        form.medida_cabeca.data = militar.medida_cabeca
        form.numero_sapato.data = militar.numero_sapato
        form.medida_calca.data = militar.medida_calca
        form.medida_camisa.data = militar.medida_camisa
        form.tipo_sanguineo.data = militar.tipo_sanguineo
        form.sinais_particulares.data = militar.sinais_particulares

        form.tatuagem.data = militar.tatuagem
        form.local_tatuagem.data = militar.local_tatuagem

    else:
        if request.method == "POST":
            print("ERROS DO FORM:", form.errors)
            flash(f"Erros no formulário: {form.errors}", "danger")

    campos_pendentes = get_campos_pendentes_cadastro(militar)
    cadastro_completo = len(campos_pendentes) == 0

    return render_template(
        "atualizacao/atualizacao_cadastral.html",
        form=form,
        militar=militar,
        cadastro_completo=cadastro_completo,
        campos_pendentes=campos_pendentes,
        graduacoes=graduacoes,
        contatos_emergencia=contatos_emergencia,
        conjuge=conjuge,
        eh_al_sd=eh_al_sd,
    )


@bp_atualizacao_cadastral.route("/painel", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_AUDITORIA_READ")
def painel():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()

    query = (
        Militar.query
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .filter(Militar.inativo.is_(False))
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Militar.nome_completo.ilike(like),
                Militar.matricula.ilike(like),
                Militar.nome_guerra.ilike(like),
            )
        )

    if status == "atualizado":
        query = query.filter(Militar.cadastro_atualizado.is_(True))
    elif status == "pendente":
        query = query.filter(
            or_(
                Militar.cadastro_atualizado.is_(False),
                Militar.cadastro_atualizado.is_(None)
            )
        )

    militares = query.order_by(Militar.nome_completo.asc()).all()

    total = Militar.query.filter(Militar.inativo.is_(False)).count()

    total_atualizado = Militar.query.filter(
        Militar.inativo.is_(False),
        Militar.cadastro_atualizado.is_(True)
    ).count()

    total_pendente = Militar.query.filter(
        Militar.inativo.is_(False),
        or_(
            Militar.cadastro_atualizado.is_(False),
            Militar.cadastro_atualizado.is_(None)
        )
    ).count()

    percentual = round((total_atualizado / total * 100), 1) if total else 0

    return render_template(
        "atualizacao/painel_atualizacao_cadastral.html",
        militares=militares,
        total=total,
        total_atualizado=total_atualizado,
        total_pendente=total_pendente,
        percentual=percentual,
        q=q,
        status=status
    )


@bp_atualizacao_cadastral.route("/painel/auditoria", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_PAINEL_READ")
def auditoria():
    q = (request.args.get("q") or "").strip()

    query = (
        AuditoriaAtualizacaoCadastral.query
        .join(Militar, Militar.id == AuditoriaAtualizacaoCadastral.militar_id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Militar.nome_completo.ilike(like),
                Militar.matricula.ilike(like),
                Militar.nome_guerra.ilike(like),
                AuditoriaAtualizacaoCadastral.ip_address.ilike(like),
            )
        )

    auditorias = query.order_by(
        AuditoriaAtualizacaoCadastral.criado_em.desc()
    ).limit(500).all()

    return render_template(
        "atualizacao/auditoria_atualizacao_cadastral.html",
        auditorias=auditorias,
        q=q
    )


# ---------------- AUDITORIA DE CONFERÊNCIA CADASTRAL ----------------

# ---------------- Rotas para conferência cadastral (marcar como conferido, listar conferências, etc) ----------------
@bp_atualizacao_cadastral.route("/conferencia", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_READ")
def conferencia_lista():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    modo = (request.args.get("modo") or "meu").strip().lower()
    # modo = "meu" -> conferência do usuário logado
    # modo = "global" -> mostra se qualquer usuário conferiu

    query = (
        database.session.query(
            Militar,
            MilitarConferenciaCadastral.id.label("conferencia_id"),
            MilitarConferenciaCadastral.conferido_em.label("conferido_em"),
            Obm.sigla.label("obm_sigla"),
            User.nome.label("conferido_por_nome"),
        )
        .outerjoin(
            MilitarObmFuncao,
            and_(
                MilitarObmFuncao.militar_id == Militar.id,
                MilitarObmFuncao.data_fim.is_(None)
            )
        )
        .outerjoin(Obm, Obm.id == MilitarObmFuncao.obm_id)
        .filter(Militar.inativo.is_(False))
    )

    if modo == "global":
        sub = (
            database.session.query(
                MilitarConferenciaCadastral.militar_id.label("militar_id"),
                func.max(MilitarConferenciaCadastral.id).label("max_id")
            )
            .group_by(MilitarConferenciaCadastral.militar_id)
            .subquery()
        )

        query = (
            query
            .outerjoin(sub, sub.c.militar_id == Militar.id)
            .outerjoin(
                MilitarConferenciaCadastral,
                MilitarConferenciaCadastral.id == sub.c.max_id
            )
            .outerjoin(User, User.id == MilitarConferenciaCadastral.user_id)
        )
    else:
        query = (
            query
            .outerjoin(
                MilitarConferenciaCadastral,
                and_(
                    MilitarConferenciaCadastral.militar_id == Militar.id,
                    MilitarConferenciaCadastral.user_id == current_user.id
                )
            )
            .outerjoin(User, User.id == MilitarConferenciaCadastral.user_id)
        )

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Militar.nome_completo.ilike(like),
                Militar.nome_guerra.ilike(like),
                Militar.matricula.ilike(like),
                Militar.cpf.ilike(like),
            )
        )

    if status == "conferido":
        query = query.filter(MilitarConferenciaCadastral.id.isnot(None))
    elif status == "pendente":
        query = query.filter(MilitarConferenciaCadastral.id.is_(None))

    registros = query.order_by(Militar.nome_completo.asc()).all()

    total = Militar.query.filter(Militar.inativo.is_(False)).count()

    if modo == "global":
        total_conferido = (
            database.session.query(func.count(func.distinct(
                MilitarConferenciaCadastral.militar_id)))
            .join(Militar, Militar.id == MilitarConferenciaCadastral.militar_id)
            .filter(Militar.inativo.is_(False))
            .scalar()
        ) or 0
    else:
        total_conferido = (
            database.session.query(MilitarConferenciaCadastral.id)
            .join(Militar, Militar.id == MilitarConferenciaCadastral.militar_id)
            .filter(
                Militar.inativo.is_(False),
                MilitarConferenciaCadastral.user_id == current_user.id
            )
            .count()
        )

    total_pendente = max(total - total_conferido, 0)
    percentual = round((total_conferido / total * 100), 1) if total else 0

    return render_template(
        "atualizacao/conferencia_lista.html",
        registros=registros,
        total=total,
        total_conferido=total_conferido,
        total_pendente=total_pendente,
        percentual=percentual,
        q=q,
        status=status,
        modo=modo,
    )


@bp_atualizacao_cadastral.route("/conferencia/proximo-pendente", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_READ")
def conferencia_proximo_pendente():
    modo = (request.args.get("modo") or "meu").strip().lower()

    query = Militar.query.filter(Militar.inativo.is_(False))

    if modo == "global":
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .distinct()
            .subquery()
        )
        proximo = (
            query
            .filter(~Militar.id.in_(sub))
            .order_by(Militar.nome_completo.asc())
            .first()
        )
    else:
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .filter(MilitarConferenciaCadastral.user_id == current_user.id)
            .subquery()
        )
        proximo = (
            query
            .filter(~Militar.id.in_(sub))
            .order_by(Militar.nome_completo.asc())
            .first()
        )

    if not proximo:
        flash("Não há militares pendentes para conferência.", "success")
        return redirect(url_for("atualizacao_cadastral.conferencia_lista", modo=modo))

    return redirect(
        url_for("atualizacao_cadastral.conferencia_detalhe",
                militar_id=proximo.id, modo=modo)
    )
# ---------------- Detalhes da conferência cadastral (marcar/desmarcar conferência, ver detalhes do militar, etc) ----------------


@bp_atualizacao_cadastral.route("/conferencia/<int:militar_id>", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_AUDITORIA_READ")
def conferencia_detalhe(militar_id):
    modo = (request.args.get("modo") or "meu").strip().lower()

    militar = (
        Militar.query
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .filter(
            Militar.id == militar_id,
            Militar.inativo.is_(False)
        )
        .first_or_404()
    )

    graduacoes = MilitarGraduacao.query.filter_by(
        militar_id=militar.id
    ).order_by(MilitarGraduacao.id.asc()).all()

    contatos_emergencia = MilitarContatoEmergencia.query.filter_by(
        militar_id=militar.id
    ).order_by(MilitarContatoEmergencia.id.asc()).all()

    conjuge = MilitarConjuge.query.filter_by(militar_id=militar.id).first()

    conferencia = MilitarConferenciaCadastral.query.filter_by(
        militar_id=militar.id,
        user_id=current_user.id
    ).first()

    conferencia_global = (
        MilitarConferenciaCadastral.query
        .join(User, User.id == MilitarConferenciaCadastral.user_id)
        .filter(MilitarConferenciaCadastral.militar_id == militar.id)
        .order_by(MilitarConferenciaCadastral.conferido_em.desc())
        .first()
    )

    campos_pendentes = get_campos_pendentes_cadastro(militar)
    cadastro_completo = len(campos_pendentes) == 0

    return render_template(
        "atualizacao/conferencia_detalhe.html",
        militar=militar,
        graduacoes=graduacoes,
        contatos_emergencia=contatos_emergencia,
        conjuge=conjuge,
        conferencia=conferencia,
        conferencia_global=conferencia_global,
        cadastro_completo=cadastro_completo,
        campos_pendentes=campos_pendentes,
        modo=modo,
    )


@bp_atualizacao_cadastral.route("/conferencia/<int:militar_id>/marcar", methods=["POST"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_CHECK")
def conferencia_marcar(militar_id):
    modo = (request.args.get("modo") or "meu").strip().lower()

    militar = Militar.query.filter(
        Militar.id == militar_id,
        Militar.inativo.is_(False)
    ).first_or_404()

    conferencia = MilitarConferenciaCadastral.query.filter_by(
        militar_id=militar.id,
        user_id=current_user.id
    ).first()

    if not conferencia:
        conferencia = MilitarConferenciaCadastral(
            militar_id=militar.id,
            user_id=current_user.id,
            conferido_em=agora_manaus(),
            observacao="Conferência cadastral realizada no painel."
        )
        database.session.add(conferencia)
        database.session.commit()
        flash("Militar marcado como conferido com sucesso.", "success")
    else:
        flash("Esse militar já foi conferido por você.", "info")

    return redirect(url_for("atualizacao_cadastral.conferencia_proximo_pendente", modo=modo))


@bp_atualizacao_cadastral.route("/conferencia/<int:militar_id>/desmarcar", methods=["POST"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_CHECK")
def conferencia_desmarcar(militar_id):
    modo = (request.args.get("modo") or "meu").strip().lower()

    militar = Militar.query.filter(
        Militar.id == militar_id,
        Militar.inativo.is_(False)
    ).first_or_404()

    conferencia = MilitarConferenciaCadastral.query.filter_by(
        militar_id=militar.id,
        user_id=current_user.id
    ).first()

    if conferencia:
        database.session.delete(conferencia)
        database.session.commit()
        flash("Check removido com sucesso.", "success")
    else:
        flash("Esse militar ainda não foi marcado por você.", "warning")

    return redirect(url_for("atualizacao_cadastral.conferencia_detalhe", militar_id=militar.id, modo=modo))


# ---------------- Exportação para Excel dos militares pendentes de conferência ----------------
@bp_atualizacao_cadastral.route("/conferencia/exportar-excel", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_EXPORT")
def conferencia_exportar_excel():
    modo = (request.args.get("modo") or "meu").strip().lower()

    query = (
        database.session.query(
            Militar.nome_completo,
            Militar.nome_guerra,
            Militar.matricula,
            Militar.cpf,
            PostoGrad.sigla.label("posto_grad"),
            Obm.sigla.label("obm_sigla"),
        )
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .outerjoin(
            MilitarObmFuncao,
            and_(
                MilitarObmFuncao.militar_id == Militar.id,
                MilitarObmFuncao.data_fim.is_(None)
            )
        )
        .outerjoin(Obm, Obm.id == MilitarObmFuncao.obm_id)
        .filter(Militar.inativo.is_(False))
    )

    if modo == "global":
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .distinct()
            .subquery()
        )
        query = query.filter(~Militar.id.in_(sub))
    else:
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .filter(MilitarConferenciaCadastral.user_id == current_user.id)
            .subquery()
        )
        query = query.filter(~Militar.id.in_(sub))

    pendentes = query.order_by(Militar.nome_completo.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendentes"

    headers = ["Nome Completo", "Nome de Guerra",
               "Matrícula", "CPF", "Posto/Grad", "OBM"]

    ws.append(headers)

    fill = PatternFill("solid", fgColor="0B5ED7")
    font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill
        cell.font = font

    for item in pendentes:
        ws.append([
            item.nome_completo or "",
            item.nome_guerra or "",
            item.matricula or "",
            item.cpf or "",
            item.posto_grad or "",
            item.obm_sigla or "",
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value or "")
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="militares_pendentes_conferencia.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- Exportação para PDF dos militares pendentes de conferência ----------------


@bp_atualizacao_cadastral.route("/conferencia/exportar-pdf", methods=["GET"])
@login_required
@require_perm("ATUALIZACAO_CADASTRAL_CONFERENCIA_EXPORT")
def conferencia_exportar_pdf():
    modo = (request.args.get("modo") or "meu").strip().lower()

    query = (
        database.session.query(
            Militar.nome_completo,
            Militar.matricula,
            PostoGrad.sigla.label("posto_grad"),
            Obm.sigla.label("obm_sigla"),
        )
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .outerjoin(
            MilitarObmFuncao,
            and_(
                MilitarObmFuncao.militar_id == Militar.id,
                MilitarObmFuncao.data_fim.is_(None)
            )
        )
        .outerjoin(Obm, Obm.id == MilitarObmFuncao.obm_id)
        .filter(Militar.inativo.is_(False))
    )

    if modo == "global":
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .distinct()
            .subquery()
        )
        query = query.filter(~Militar.id.in_(sub))
    else:
        sub = (
            database.session.query(MilitarConferenciaCadastral.militar_id)
            .filter(MilitarConferenciaCadastral.user_id == current_user.id)
            .subquery()
        )
        query = query.filter(~Militar.id.in_(sub))

    pendentes = query.order_by(Militar.nome_completo.asc()).all()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    y = height - 40
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(30, y, "Relatório de Militares Pendentes de Conferência")
    y -= 30

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(30, y, "Nome")
    pdf.drawString(340, y, "Matrícula")
    pdf.drawString(450, y, "Posto/Grad")
    pdf.drawString(560, y, "OBM")
    y -= 20

    pdf.setFont("Helvetica", 9)

    for item in pendentes:
        if y < 40:
            pdf.showPage()
            y = height - 40
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(30, y, "Nome")
            pdf.drawString(340, y, "Matrícula")
            pdf.drawString(450, y, "Posto/Grad")
            pdf.drawString(560, y, "OBM")
            y -= 20
            pdf.setFont("Helvetica", 9)

        nome = (item.nome_completo or "")[:55]
        matricula = item.matricula or "-"
        posto = item.posto_grad or "-"
        obm = item.obm_sigla or "-"

        pdf.drawString(30, y, nome)
        pdf.drawString(340, y, matricula)
        pdf.drawString(450, y, posto)
        pdf.drawString(560, y, obm)
        y -= 16

    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="militares_pendentes_conferencia.pdf",
        mimetype="application/pdf"
    )
