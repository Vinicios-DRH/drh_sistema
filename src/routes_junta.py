from io import BytesIO
from math import ceil
from pathlib import Path

from flask import Blueprint, jsonify, render_template, request, redirect, send_file, url_for, flash, send_from_directory
from flask_login import login_required, current_user
from sqlalchemy import or_, func
from sqlalchemy.orm import joinedload

from src import database
from src.models import JuntaFechamentoBg, Militar, Licencas
from src.forms import FormLicencas
from src.services.junta_medica import (
    calcular_data_fim,
    calcular_situacao_atual,
    calcular_status_registro,
    calcular_status_atual,
    label_status,
    label_tipo,
    exige_inspecao_pos_lts,
    montar_dados_licencas,
    STATUS_LABELS,
    TIPO_LICENCA_LABELS,
)
from src.services.junta_bg_generator import gerar_nota_bg_docx
from src.services.junta_periodos import montar_blocos_por_militar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
from calendar import monthrange
from zoneinfo import ZoneInfo

junta_bp = Blueprint("junta", __name__, url_prefix="/junta")

MANAUS_TZ = ZoneInfo("America/Manaus")


def hoje_manaus():
    return datetime.now(MANAUS_TZ).date()


def _obter_obm_atual(militar):
    obm_sigla = ""

    if hasattr(militar, "obm_funcoes") and militar.obm_funcoes:
        obms_ativas = [x for x in militar.obm_funcoes if getattr(
            x, "data_fim", None) is None]

        if obms_ativas:
            obm = obms_ativas[-1].obm
            obm_sigla = obm.sigla if obm else ""
        else:
            obm = militar.obm_funcoes[-1].obm
            obm_sigla = obm.sigla if obm else ""

    return obm_sigla or ""


MESES_PT = [
    "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
    "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
]


def data_por_extenso_maiuscula(dt):
    if not dt:
        return ""
    return f"{dt.day} DE {MESES_PT[dt.month - 1]} DE {dt.year}"


@junta_bp.route("/nova-licenca", methods=["GET", "POST"])
@login_required
def nova_licenca():
    form = FormLicencas()
    hoje = hoje_manaus()
    data_extenso_hoje = data_por_extenso_maiuscula(hoje)

    if form.validate_on_submit():
        try:
            militar_id = int(form.militar_id.data)
            militar = Militar.query.get(militar_id)

            if not militar:
                flash("Militar não encontrado.", "danger")
                return redirect(url_for("junta.nova_licenca"))

            historico = (
                Licencas.query
                .filter_by(militar_id=militar.id)
                .order_by(Licencas.data_inicio.desc(), Licencas.id.desc())
                .all()
            )

            situacao = calcular_situacao_atual(historico)
            status_atual = situacao["status_atual"]

            tipo = form.tipo_licenca.data

            numero_bg_curso = None
            data_extenso_curso = None

            if tipo == "CURSO":
                resultado_curso = (form.resultado_curso.data or "").strip()
                numero_bg_curso = (form.numero_bg_curso.data or "").strip()

                if resultado_curso not in {"CURSO_APTO", "CURSO_INAPTO"}:
                    flash("Informe o resultado para fins de curso.", "danger")
                    return redirect(url_for("junta.nova_licenca"))

                if not numero_bg_curso:
                    flash("Informe o número do BG para fins de curso.", "danger")
                    return redirect(url_for("junta.nova_licenca"))

                qtd_dias = 1
                data_inicio = hoje
                data_fim = hoje
                status_registro = resultado_curso
                data_extenso_curso = data_extenso_hoje

            elif tipo == "AGREGADO":
                data_inicio = form.data_inicio.data

                if not data_inicio:
                    flash("Informe a data para o registro de agregação.", "danger")
                    return redirect(url_for("junta.nova_licenca"))

                if status_atual != "APTO_RESTR":
                    flash(
                        "A agregação manual só pode ser registrada quando a situação atual do militar estiver como APTO COM RESTRIÇÕES.",
                        "danger"
                    )
                    return redirect(url_for("junta.nova_licenca"))

                qtd_dias = 1
                data_fim = data_inicio
                status_registro = calcular_status_registro(tipo)

            else:
                data_inicio = form.data_inicio.data
                qtd_dias = form.qtd_dias.data

                if not data_inicio:
                    flash("Informe a data de início.", "danger")
                    return redirect(url_for("junta.nova_licenca"))

                if not qtd_dias:
                    flash("Informe a quantidade de dias.", "danger")
                    return redirect(url_for("junta.nova_licenca"))

                data_fim = calcular_data_fim(data_inicio, qtd_dias)
                status_registro = calcular_status_registro(tipo)

            nova = Licencas(
                militar_id=militar.id,
                tipo_licenca=tipo,
                recebimento_bg=None,
                qtd_dias=qtd_dias,
                data_inicio=data_inicio,
                data_fim=data_fim,
                status=status_registro,
                sessao=form.sessao.data.strip(),
                numero_bg_curso=numero_bg_curso,
                data_extenso_curso=data_extenso_curso,
                observacao=form.observacao.data.strip() if form.observacao.data else None,
                usuario_id=current_user.id
            )

            database.session.add(nova)
            database.session.commit()

            flash("Registro da Junta Médica adicionado com sucesso!", "success")
            return redirect(url_for("junta.nova_licenca"))

        except Exception as e:
            database.session.rollback()
            flash(f"Erro ao salvar licença: {str(e)}", "danger")

    pendentes_hoje = (
        Licencas.query
        .filter(
            func.date(Licencas.created_at) == hoje,
            Licencas.fechamento_bg_id.is_(None)
        )
        .count()
    )

    return render_template(
        "junta/nova_licenca.html",
        form=form,
        tipo_labels=TIPO_LICENCA_LABELS,
        status_labels=STATUS_LABELS,
        pendentes_hoje=pendentes_hoje,
        hoje=hoje,
        data_extenso_hoje=data_extenso_hoje
    )


@junta_bp.route("/licencas", methods=["GET"])
@login_required
def listar_licencas():
    page = request.args.get("page", 1, type=int)
    per_page = 20

    filtro_q = (request.args.get("q") or "").strip()
    filtro_tipo = (request.args.get("tipo") or "").strip()
    filtro_status = (request.args.get("status") or "").strip()
    filtro_status_atual = (request.args.get("status_atual") or "").strip()
    filtro_nota_bg = (request.args.get("nota_bg") or "").strip()

    dados, resumo = montar_dados_licencas(
        filtro_q=filtro_q,
        filtro_tipo=filtro_tipo,
        filtro_status=filtro_status,
        filtro_status_atual=filtro_status_atual,
        filtro_nota_bg=filtro_nota_bg,
    )

    total = len(dados)
    total_pages = max(1, ceil(total / per_page)) if total else 1

    if page > total_pages:
        page = total_pages

    inicio = (page - 1) * per_page
    fim = inicio + per_page
    dados_pagina = dados[inicio:fim]

    tipos_filtro = [
        ("LTS", "LTS"),
        ("LTSPF", "LTSPF"),
        ("LM", "Licença Maternidade"),
        ("APTO_RECOM", "Apto com Recomendações"),
        ("APTO_RESTR", "Apto com Restrições"),
        ("APTO", "Apto sem Restrição"),
        ("AGREGADO", "Agregado"),
    ]

    status_filtro = [
        ("LTS", "LTS"),
        ("LTSPF", "LTSPF"),
        ("LM", "Licença Maternidade"),
        ("APTO_RECOM", "Apto com Recomendações"),
        ("APTO_RESTR", "Apto com Restrições"),
        ("APTO", "Apto sem Restrição"),
        ("AGREGADO", "Agregado"),
    ]

    status_atual_filtro = [
        ("LTS", "LTS"),
        ("LTSPF", "LTSPF"),
        ("LM", "Licença Maternidade"),
        ("APTO_RECOM", "Apto com Recomendações"),
        ("APTO_RESTR", "Apto com Restrições"),
        ("APTO", "Apto sem Restrição"),
        ("AGUARDANDO_INSPECAO", "Aguardando Inspeção"),
        ("AGREGADO", "Agregado"),
    ]

    return render_template(
        "junta/listar_licencas.html",
        dados=dados_pagina,
        resumo=resumo,
        current_page=page,
        total_pages=total_pages,
        total=total,
        filtro_q=filtro_q,
        filtro_tipo=filtro_tipo,
        filtro_status=filtro_status,
        filtro_status_atual=filtro_status_atual,
        filtro_nota_bg=filtro_nota_bg,
        tipos_filtro=tipos_filtro,
        status_filtro=status_filtro,
        status_atual_filtro=status_atual_filtro,
    )


@junta_bp.route("/historico/<int:militar_id>", methods=["GET"])
@login_required
def historico_militar(militar_id):
    militar = (
        Militar.query
        .options(
            joinedload(Militar.posto_grad),
            joinedload(Militar.quadro),
        )
        .get_or_404(militar_id)
    )

    registros = (
        Licencas.query
        .filter_by(militar_id=militar_id)
        .order_by(Licencas.data_inicio.desc(), Licencas.id.desc())
        .all()
    )

    situacao = calcular_situacao_atual(registros)

    return render_template(
        "junta/historico_militar.html",
        militar=militar,
        registros=registros,
        status_atual=situacao["status_atual"],
        status_atual_label=situacao["status_atual_label"],
        agregacao=situacao["agregacao"],
        label_status=label_status,
        label_tipo=label_tipo,
    )


@junta_bp.route("/api/militares/buscar", methods=["GET"])
@login_required
def buscar_militares():
    q = (request.args.get("q") or "").strip()

    if len(q) < 2:
        return jsonify([])

    militares = (
        Militar.query
        .options(joinedload(Militar.posto_grad))
        .filter(
            Militar.inativo == False,
            or_(
                Militar.nome_completo.ilike(f"%{q}%"),
                Militar.nome_guerra.ilike(f"%{q}%"),
            )
        )
        .order_by(Militar.nome_completo.asc())
        .limit(20)
        .all()
    )

    resultados = []
    for m in militares:
        resultados.append({
            "id": m.id,
            "nome": m.nome_completo or "",
            "nome_guerra": m.nome_guerra or "",
            "matricula": m.matricula or "",
            "posto_grad": m.posto_grad.sigla if m.posto_grad else "",
            "label": f"{m.nome_completo}"
        })

    return jsonify(resultados)


@junta_bp.route("/api/militar/<int:militar_id>", methods=["GET"])
@login_required
def get_militar_info(militar_id):
    militar = (
        Militar.query
        .options(
            joinedload(Militar.posto_grad),
            joinedload(Militar.quadro),
        )
        .get_or_404(militar_id)
    )

    return jsonify({
        "id": militar.id,
        "nome": militar.nome_completo or "",
        "nome_guerra": militar.nome_guerra or "",
        "matricula": militar.matricula or "",
        "posto_grad": militar.posto_grad.sigla if militar.posto_grad else "N/D",
        "quadro": militar.quadro.quadro if militar.quadro else "N/D",
        "obm": _obter_obm_atual(militar)
    })


@junta_bp.route("/licencas/exportar-excel", methods=["GET"])
@login_required
def exportar_licencas_excel():
    filtro_q = (request.args.get("q") or "").strip()
    filtro_tipo = (request.args.get("tipo") or "").strip()
    filtro_status = (request.args.get("status") or "").strip()
    filtro_status_atual = (request.args.get("status_atual") or "").strip()

    dados, resumo = montar_dados_licencas(
        filtro_q=filtro_q,
        filtro_tipo=filtro_tipo,
        filtro_status=filtro_status,
        filtro_status_atual=filtro_status_atual,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Licenças Junta"

    headers = [
        "Militar",
        "Tipo",
        "Status do Registro",
        "Status Atual",
        "BG",
        "Dias",
        "Data Início",
        "Data Fim",
        "Sessão",
        "Agregação",
        "Observação",
        "Criado em",
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="0B2F4F")
    font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in dados:
        reg = item["registro"]
        agregacao = item["agregacao"]

        if agregacao["atingiu_limite"]:
            agg_texto = "AGREGADO/Apto à agregação"
        elif agregacao["alerta"]:
            agg_texto = agregacao["mensagem"] or "Próximo da agregação"
        else:
            agg_texto = "-"

        ws.append([
            f"{reg.militar.posto_grad.sigla if reg.militar and reg.militar.posto_grad else ''} {reg.militar.nome_completo if reg.militar else ''}".strip(),
            item["tipo_label"],
            item["status_label"],
            item["status_atual_label"],
            reg.recebimento_bg,
            reg.qtd_dias,
            reg.data_inicio.strftime("%d/%m/%Y") if reg.data_inicio else "",
            reg.data_fim.strftime("%d/%m/%Y") if reg.data_fim else "",
            reg.sessao,
            agg_texto,
            reg.observacao or "",
            reg.created_at.strftime(
                "%d/%m/%Y %H:%M") if reg.created_at else "",
        ])

    larguras = {
        "A": 42, "B": 24, "C": 28, "D": 28, "E": 16, "F": 10,
        "G": 14, "H": 14, "I": 18, "J": 38, "K": 45, "L": 18
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="licencas_junta.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@junta_bp.route("/licencas/relatorio", methods=["GET"])
@login_required
def relatorio_licencas():
    filtro_q = (request.args.get("q") or "").strip()
    filtro_tipo = (request.args.get("tipo") or "").strip()
    filtro_status = (request.args.get("status") or "").strip()
    filtro_status_atual = (request.args.get("status_atual") or "").strip()

    dados, resumo = montar_dados_licencas(
        filtro_q=filtro_q,
        filtro_tipo=filtro_tipo,
        filtro_status=filtro_status,
        filtro_status_atual=filtro_status_atual,
    )

    return render_template(
        "junta/relatorio_licencas.html",
        dados=dados,
        resumo=resumo,
        filtro_q=filtro_q,
        filtro_tipo=filtro_tipo,
        filtro_status=filtro_status,
        filtro_status_atual=filtro_status_atual,
    )


@junta_bp.route("/licencas/finalizar-bg", methods=["POST"])
@login_required
def finalizar_bg_dia():
    try:
        data_ref_str = (request.form.get("data_referencia") or "").strip()
        nota_bg = (request.form.get("nota_bg") or "").strip()
        sessao_bg = (request.form.get("sessao_bg") or "").strip()
        observacao_bg = (request.form.get("observacao_bg") or "").strip()

        if not data_ref_str:
            flash("Informe a data de referência do fechamento.", "danger")
            return redirect(url_for("junta.nova_licenca"))

        if not nota_bg:
            flash("Informe a nota para BG.", "danger")
            return redirect(url_for("junta.nova_licenca"))

        if not sessao_bg:
            flash("Informe a sessão do fechamento.", "danger")
            return redirect(url_for("junta.nova_licenca"))

        data_referencia = datetime.strptime(data_ref_str, "%Y-%m-%d").date()

        pendentes = (
            Licencas.query
            .filter(
                func.date(Licencas.created_at) == data_referencia,
                Licencas.fechamento_bg_id.is_(None)
            )
            .order_by(Licencas.created_at.asc(), Licencas.id.asc())
            .all()
        )

        if not pendentes:
            flash(
                "Não há lançamentos pendentes para finalizar nessa data e sessão.", "warning")
            return redirect(url_for("junta.nova_licenca"))

        fechamento = JuntaFechamentoBg(
            data_referencia=data_referencia,
            nota_bg=nota_bg,
            sessao=sessao_bg,
            observacao=observacao_bg or None,
            usuario_id=current_user.id
        )

        database.session.add(fechamento)
        database.session.flush()

        for lic in pendentes:
            lic.fechamento_bg_id = fechamento.id

        # ainda sem commit final
        arquivo = gerar_nota_bg_docx(fechamento.id, commit_db=False)

        fechamento.arquivo_docx = arquivo
        database.session.commit()

        flash(
            f"Fechamento realizado com sucesso. {len(pendentes)} lançamento(s) vinculados à nota BG {nota_bg}.",
            "success"
        )

    except Exception as e:
        database.session.rollback()
        flash(f"Erro ao finalizar BG do dia: {str(e)}", "danger")

    return redirect(
        url_for(
            "junta.baixar_docx_fechamento",
            fechamento_id=fechamento.id
        )
    )


@junta_bp.route("/fechamento-bg/<int:fechamento_id>/baixar-docx", methods=["GET"])
@login_required
def baixar_docx_fechamento(fechamento_id):
    fechamento = JuntaFechamentoBg.query.get_or_404(fechamento_id)

    if not fechamento.arquivo_docx:
        flash("Este fechamento ainda não possui documento gerado.", "warning")
        return redirect(url_for("junta.listar_licencas"))

    pasta = Path("src/static/junta_bg")
    return send_from_directory(
        directory=str(pasta.resolve()),
        path=fechamento.arquivo_docx,
        as_attachment=True
    )


@junta_bp.route("/renovacoes", methods=["GET"])
@login_required
def painel_renovacoes():
    mes = request.args.get("mes", type=int)
    ano = request.args.get("ano", type=int)
    tipo = (request.args.get("tipo") or "").strip()

    query = (
        Licencas.query
        .options(
            joinedload(Licencas.militar).joinedload(Militar.posto_grad),
            joinedload(Licencas.militar).joinedload(Militar.quadro),
        )
        .order_by(Licencas.militar_id.asc(), Licencas.data_inicio.asc(), Licencas.id.asc())
    )

    if tipo:
        query = query.filter(Licencas.tipo_licenca == tipo)

    if mes and ano:
        inicio_mes = date(ano, mes, 1)
        fim_mes = date(ano, mes, monthrange(ano, mes)[1])

        query = query.filter(
            Licencas.data_inicio <= fim_mes,
            Licencas.data_fim >= inicio_mes
        )

    registros = query.all()
    blocos_por_militar = montar_blocos_por_militar(registros)

    linhas = []

    for militar_id, blocos in blocos_por_militar.items():
        for bloco in blocos:
            primeiro_reg = bloco["registros"][0]
            militar = primeiro_reg.militar

            linhas.append({
                "militar": militar,
                "tipo_licenca": bloco["tipo_licenca"],
                "tipo_label": bloco["tipo_label"],
                "inicio_bloco": bloco["inicio_bloco"],
                "fim_bloco": bloco["fim_bloco"],
                "dias_continuos": bloco["dias_continuos"],
                "renovacoes": bloco["renovacoes"],
                "quantidade_registros": bloco["quantidade_registros"],
                "ultima_renovacao": bloco["ultima_renovacao"],
                "meses_abrangidos": bloco["meses_abrangidos"],
                "datas_renovacao": bloco["datas_renovacao"],
                "suspeito": bloco["suspeito"],
                "motivo_suspeita": bloco["motivo_suspeita"],
            })

    linhas.sort(key=lambda x: (
        x["suspeito"], x["renovacoes"], x["dias_continuos"]), reverse=True)

    resumo = {
        "total_blocos": len(linhas),
        "com_renovacao": sum(1 for x in linhas if x["renovacoes"] > 0),
        "suspeitos": sum(1 for x in linhas if x["suspeito"]),
    }

    return render_template(
        "junta/painel_renovacoes.html",
        linhas=linhas,
        resumo=resumo,
        filtro_mes=mes,
        filtro_ano=ano,
        filtro_tipo=tipo,
    )
