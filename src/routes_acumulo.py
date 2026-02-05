from src.formatar_cpf import get_militar_por_user
from src import app
from io import BytesIO
from flask import Blueprint, abort, flash, redirect, render_template, send_file, session, url_for, request, jsonify
from flask_login import login_required, current_user
from openpyxl import Workbook
from sqlalchemy import case, exists, literal, or_, and_, func
from src.formatar_cpf import get_militar_por_user
from src.decorators.utils_acumulo import b2_presigned_get, b2_upload_fileobj, b2_check, b2_put_test, build_prefix
from src.models import AuditoriaDeclaracao, DocumentoMilitar, Funcao, User, database as db, Militar, PostoGrad, Obm, DeclaracaoAcumulo, MilitarObmFuncao, VinculoExterno, DraftDeclaracaoAcumulo, FichaAlunos
from src.decorators.control import USERS_ANALISE_VINCULO, checar_ocupacao, _is_super_user, _is_drh_like, _user_obm_ids, _militar_permitido, _can_analise_vinculo, _can_editar_declaracao, _prazo_envio_ate, _prazo_fechado, ANO_ATUAL, _is_privilegiado, _parse_date, _parse_time, require_analise_vinculo, _usuario_ja_tem_declaracao, _is_auditor_vinculo
from sqlalchemy.orm import joinedload, selectinload
from datetime import datetime, date, time
from src.decorators.formatar_datas import formatar_data_extenso, formatar_data_sem_zero
from src.decorators.helpers_docx import render_docx_from_template
from sqlalchemy.sql import functions
from sqlalchemy.orm import aliased
from src.identificacao import get_aluno_por_user, ensure_militar_from_aluno

bp_acumulo = Blueprint("acumulo", __name__, url_prefix="/acumulo")


@app.route("/home-atualizacao", methods=["GET"])
@login_required
def home_atualizacao():
    M, PG, O, MOF = Militar, PostoGrad, Obm, MilitarObmFuncao
    D = DeclaracaoAcumulo
    DM = DocumentoMilitar

    U = current_user

    prazo_limite = _prazo_envio_ate()
    prazo_fechado = date.today() > prazo_limite

    # -------- 1) Resolver Militar do usuário --------
    militar = None

    mid = getattr(U, "militar_id", None)
    if mid:
        militar = db.session.get(M, mid)

    if not militar or getattr(U, "funcao_user_id", None) == 12:
        try:
            m_by_helper = get_militar_por_user(U)  # usa CPF do User
            if m_by_helper:
                militar = m_by_helper
        except Exception:
            pass

    # ⚠️ NÃO redireciona mais se não achar militar: aluno não tem Militar mesmo.
    # if not militar and getattr(U, "funcao_user_id", None) == 12:
    #     flash("Não foi possível localizar seus dados de militar.", "danger")
    #     return redirect(url_for("home"))

    militar_id = militar.id if militar else None

    aluno = None
    if not militar:
        try:
            aluno = get_aluno_por_user(U)
        except Exception:
            pass

    pessoa_tipo_atual = "militar" if militar else ("aluno" if aluno else None)
    pessoa_id_atual = militar.id if militar else (aluno.id if aluno else None)

    # -------- 2) Dados do cabeçalho --------
    user_pg = None
    try:
        if militar and getattr(militar, "posto_grad_id", None):
            user_pg = db.session.query(PG.sigla).filter(
                PG.id == militar.posto_grad_id).scalar()
    except Exception:
        pass

    user_obm_siglas = []
    try:
        if militar_id:
            rows = (
                db.session.query(O.sigla)
                .join(MOF, MOF.obm_id == O.id)
                .filter(MOF.militar_id == militar_id, MOF.data_fim.is_(None))
                .all()
            )
            user_obm_siglas = [s for (s,) in rows] or []
    except Exception:
        pass

    # 🔁 Fallback de OBM pela relação do User (cobre caso aluno sem Militar)
    if not user_obm_siglas:
        try:
            if getattr(U, "obm1", None) and getattr(U.obm1, "sigla", None):
                user_obm_siglas.append(U.obm1.sigla)
            if getattr(U, "obm2", None) and getattr(U.obm2, "sigla", None):
                user_obm_siglas.append(U.obm2.sigla)
        except Exception:
            pass

    militar_funcoes = []
    try:
        if militar and militar.obm_funcoes:
            militar_funcoes = [
                f.funcao.ocupacao
                for f in militar.obm_funcoes
                if f.data_fim is None and f.funcao
            ]
    except Exception:
        pass

    # Funções (FuncaoUser) no User
    user_funcoes = []
    fun = getattr(U, "user_funcao", None)
    if isinstance(fun, (list, tuple)):
        user_funcoes = [
            f.ocupacao for f in fun if getattr(f, "ocupacao", None)]
    elif fun and getattr(fun, "ocupacao", None):
        user_funcoes = [fun.ocupacao]

    ano_atual = datetime.now().year

    # -------- 3) KPIs (só se tiver militar_id) --------
    kpi_decl_total = kpi_decl_pendentes = kpi_decl_validadas = kpi_decl_inconformes = 0
    minhas_declaracoes = []
    inconforme_info_map = {}

    if militar_id:
        kpi_q = (
            db.session.query(D.status, func.count(D.id))
            .filter(D.militar_id == militar_id, D.ano_referencia == ano_atual)
            .group_by(D.status)
            .all()
        )
        kpi_map = {s: n for s, n in kpi_q}
        kpi_decl_total = sum(kpi_map.values())
        kpi_decl_pendentes = kpi_map.get("pendente", 0)
        kpi_decl_validadas = kpi_map.get("validado", 0)
        kpi_decl_inconformes = kpi_map.get("inconforme", 0)

        minhas_declaracoes = (
            db.session.query(D)
            .filter(D.militar_id == militar_id, D.ano_referencia == ano_atual)
            .order_by(D.updated_at.desc().nullslast(), D.id.desc())
            .limit(6)
            .all()
        )

        if minhas_declaracoes:
            decl_ids = [d.id for d in minhas_declaracoes]
            auds = (
                db.session.query(AuditoriaDeclaracao)
                .filter(
                    AuditoriaDeclaracao.declaracao_id.in_(decl_ids),
                    AuditoriaDeclaracao.para_status == "inconforme",
                    AuditoriaDeclaracao.motivo.isnot(None),
                )
                .order_by(
                    AuditoriaDeclaracao.declaracao_id.asc(),
                    AuditoriaDeclaracao.data_alteracao.desc(),
                )
                .all()
            )
            for a in auds:
                if a.declaracao_id not in inconforme_info_map:
                    inconforme_info_map[a.declaracao_id] = {
                        "motivo": a.motivo or "",
                        "quando": a.data_alteracao,
                    }

    # -------- 4) Atividades recentes (por user_id) --------
    atividades = []
    try:
        aud = (
            db.session.query(AuditoriaDeclaracao)
            .filter(AuditoriaDeclaracao.alterado_por_user_id == U.id)
            .order_by(AuditoriaDeclaracao.data_alteracao.desc())
            .limit(8)
            .all()
        )
        for a in aud:
            atividades.append(type("A", (), {
                "quando": a.data_alteracao,
                "texto": f"Alterou declaração #{a.declaracao_id} ({a.de_status} → {a.para_status})"
            }))
    except Exception:
        pass

    last_login_dt = getattr(U, "last_login_at", None)
    last_ip = getattr(U, "last_login_ip", None)
    drh_like = _is_drh_like()

    nome_exibicao = (
        getattr(militar, "nome_completo", None)
        or getattr(current_user, "nome", None)
        or getattr(current_user, "login", None)
        or "Usuário"
    )

    # -------- 5) Documentos pendentes para o usuário --------
    docs_pendentes = []
    docs_pendentes_qtd = 0
    try:
        if getattr(current_user, "cpf", None):
            docs_pendentes = (
                db.session.query(DM)
                .filter(DM.destinatario_cpf == current_user.cpf)
                .filter(DM.baixado_em.is_(None))
                .order_by(DM.criado_em.desc())
                .limit(5)
                .all()
            )
            docs_pendentes_qtd = (
                db.session.query(func.count(DM.id))
                .filter(DM.destinatario_cpf == current_user.cpf)
                .filter(DM.baixado_em.is_(None))
                .scalar() or 0
            )
    except Exception:
        pass

    # ⚠️ Define a variável usada no template/JS
    militar_id_atual = militar_id
    tem_inconforme_no_ano = (kpi_decl_inconformes or 0) > 0

    return render_template(
        "home_atualizacao.html",
        ano_atual=ano_atual,
        user_pg=user_pg,
        user_obm_siglas=user_obm_siglas,
        user_funcoes=user_funcoes,
        kpi_decl_total=kpi_decl_total,
        kpi_decl_pendentes=kpi_decl_pendentes,
        kpi_decl_validadas=kpi_decl_validadas,
        kpi_decl_inconformes=kpi_decl_inconformes,
        minhas_declaracoes=minhas_declaracoes,
        atividades=atividades,
        last_login_dt=last_login_dt,
        last_ip=last_ip,
        twofa_enabled=getattr(U, "twofa_enabled", False),
        drh_like=drh_like,
        militar=militar,
        nome_exibicao=nome_exibicao,
        militar_funcoes=militar_funcoes,
        inconforme_info_map=inconforme_info_map,
        militar_id_atual=militar_id_atual,           # mantém (pode ser None)
        pessoa_tipo_atual=pessoa_tipo_atual,
        pessoa_id_atual=pessoa_id_atual,
        docs_pendentes=docs_pendentes,
        docs_pendentes_qtd=docs_pendentes_qtd,
        prazo_fechado=prazo_fechado,
        prazo_limite=prazo_limite,
        tem_inconforme_no_ano=tem_inconforme_no_ano,
    )


@bp_acumulo.route("/lista", methods=["GET"])
@login_required
@checar_ocupacao('DIRETOR', 'CHEFE', 'SUPER USER', 'DIRETOR DRH')
def lista():
    # filtros
    ano = request.args.get("ano", type=int)
    q = (request.args.get("q") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    ano_ref = ano or datetime.now().year

    MOF = MilitarObmFuncao

    # Base: Militar -> PostoGrad -> (função ativa) -> OBM
    qry = (
        db.session.query(Militar)
        .select_from(Militar)
        .outerjoin(PostoGrad, Militar.posto_grad_id == PostoGrad.id)
        .outerjoin(MOF, and_(MOF.militar_id == Militar.id, MOF.data_fim.is_(None)))
        .outerjoin(Obm, Obm.id == MOF.obm_id)
        .distinct(Militar.id)  # DISTINCT ON (militar.id) no Postgres
    )

    # 🔒 Escopo por OBM via User (só super vê tudo)
    if not _is_super_user():
        obm_ids = _user_obm_ids()
        if obm_ids:
            qry = qry.filter(Obm.id.in_(obm_ids))
        else:
            # sem OBM atribuída no usuário → nada a mostrar
            qry = qry.filter(db.text("1=0"))

    # Busca
    if q:
        ilike = f"%{q}%"
        qry = qry.filter(or_(
            Militar.nome_completo.ilike(ilike),
            Militar.matricula.ilike(ilike),
            PostoGrad.sigla.ilike(ilike),
            Obm.sigla.ilike(ilike),
        ))

    # Postgres exige que ORDER BY comece pelo campo do DISTINCT ON
    qry = qry.order_by(Militar.id.asc(), Militar.nome_completo.asc())

    # Paginação
    pagination = qry.paginate(page=page, per_page=per_page, error_out=False)
    militares = pagination.items

    # OBM "ativa" para exibir no template
    ids = [m.id for m in militares]
    obm_map = {}
    if ids:
        rows = (
            db.session.query(MOF.militar_id, Obm.sigla)
            .join(Obm, Obm.id == MOF.obm_id)
            .filter(MOF.militar_id.in_(ids), MOF.data_fim.is_(None))
            .all()
        )
        for mid, sigla in rows:
            obm_map.setdefault(mid, sigla)  # primeira OBM ativa encontrada
    for m in militares:
        m.obm_sigla = obm_map.get(m.id, "-")

    # "Última declaração" por militar
    ultima_por_militar = {}
    if ids:
        if ano:  # do ano filtrado
            decls = (
                db.session.query(DeclaracaoAcumulo)
                .filter(
                    DeclaracaoAcumulo.militar_id.in_(ids),
                    DeclaracaoAcumulo.ano_referencia == ano
                )
                .all()
            )
            ultima_por_militar = {d.militar_id: d for d in decls}
        else:    # mais recente
            sub = (
                db.session.query(
                    DeclaracaoAcumulo.militar_id,
                    func.max(DeclaracaoAcumulo.ano_referencia).label(
                        "max_ano"),
                )
                .filter(DeclaracaoAcumulo.militar_id.in_(ids))
                .group_by(DeclaracaoAcumulo.militar_id)
                .subquery()
            )
            decls = (
                db.session.query(DeclaracaoAcumulo)
                .join(
                    sub,
                    and_(
                        DeclaracaoAcumulo.militar_id == sub.c.militar_id,
                        DeclaracaoAcumulo.ano_referencia == sub.c.max_ano,
                    ),
                )
                .all()
            )
            ultima_por_militar = {d.militar_id: d for d in decls}
    for m in militares:
        m.ultima_declaracao = ultima_por_militar.get(m.id)

    # KPIs (mesmo escopo da lista)
    base_total = (
        db.session.query(Militar.id)
        .select_from(Militar)
        .outerjoin(PostoGrad, Militar.posto_grad_id == PostoGrad.id)
        .outerjoin(MOF, and_(MOF.militar_id == Militar.id, MOF.data_fim.is_(None)))
        .outerjoin(Obm, Obm.id == MOF.obm_id)
        .distinct(Militar.id)
    )
    if not _is_super_user():
        obm_ids = _user_obm_ids()
        if obm_ids:
            base_total = base_total.filter(Obm.id.in_(obm_ids))
        else:
            base_total = base_total.filter(db.text("1=0"))

    total_militares_visiveis = base_total.count()

    entregues = 0
    if total_militares_visiveis:
        base_decl = (
            db.session.query(DeclaracaoAcumulo.militar_id)
            .join(Militar, Militar.id == DeclaracaoAcumulo.militar_id)
            .join(MOF, and_(MOF.militar_id == Militar.id, MOF.data_fim.is_(None)), isouter=True)
            .join(Obm, Obm.id == MOF.obm_id, isouter=True)
            .distinct()
        )
        if not _is_super_user():
            obm_ids = _user_obm_ids()
            if obm_ids:
                base_decl = base_decl.filter(Obm.id.in_(obm_ids))
            else:
                base_decl = base_decl.filter(db.text("1=0"))

        entregues = (
            base_decl
            .filter(DeclaracaoAcumulo.ano_referencia == ano_ref)
            .count()
        )

    pendentes = max(total_militares_visiveis - entregues, 0)
    adesao = f"{(entregues / total_militares_visiveis * 100):.0f}%" if total_militares_visiveis else "0%"

    # Paginação HTML
    def render_pagination(p):
        if p.pages <= 1:
            return ""
        items = []
        for num in range(1, p.pages + 1):
            cls = "active" if num == p.page else ""
            href = f"?page={num}&per_page={per_page}&ano={ano or ''}&q={q}"
            items.append(
                f'<li class="page-item {cls}"><a class="page-link" href="{href}">{num}</a></li>')
        return '<nav><ul class="pagination pagination-sm justify-content-end mt-3">' + "".join(items) + "</ul></nav>"

    pagination_html = render_pagination(pagination)

    return render_template(
        "acumulo_lista.html",
        ano=ano,
        militares=militares,
        entregues=entregues,
        pendentes=pendentes,
        adesao=adesao,
        pagination=pagination_html,
        tem_permissao=True,
    )


# def _obms_ativas_do_militar(militar_id: int):
#     MOF = MilitarObmFuncao
#     rows = (db.session.query(MOF.obm_id)
#             .filter(and_(MOF.militar_id == militar_id, MOF.data_fim.is_(None)))
#             # ajuste a ordenação se você tiver prioridade
#             .order_by(MOF.id.asc())
#             .all())
#     ids = [r.obm_id for r in rows]
#     obm_id_1 = ids[0] if len(ids) > 0 else None
#     obm_id_2 = ids[1] if len(ids) > 1 else None
#     return obm_id_1, obm_id_2


# @bp_acumulo.route("/novo/<int:militar_id>", methods=["GET", "POST"])
# @login_required
# def novo(militar_id):
#     # Se for usuário comum, força militar_id a ser o do próprio usuário
#     if current_user.funcao_user_id == 12:
#         militar = get_militar_por_user(current_user)
#         if not militar:
#             flash("Não foi possível localizar seus dados de militar.", "danger")
#             return redirect(url_for("home"))
#         militar_id = militar.id

#     ano = request.values.get("ano", type=int) or datetime.now().year

#     if _prazo_fechado() and not _is_drh_like():
#         ja_tem = db.session.query(DeclaracaoAcumulo.id).filter(
#             DeclaracaoAcumulo.militar_id == militar_id,
#             DeclaracaoAcumulo.ano_referencia == ano
#         ).first()
#         if ja_tem:
#             flash("Prazo encerrado. Reaberto apenas para: (1) quem não enviou nenhuma declaração; ou (2) corrigir declarações INCONFORME.", "warning")
#             return redirect(url_for("acumulo.minhas_declaracoes", ano=ano))

#     pendente = db.session.query(DeclaracaoAcumulo.id).filter_by(
#         militar_id=militar_id, ano_referencia=ano, status='pendente'
#     ).first()
#     if pendente:
#         flash("Você já possui uma declaração enviada e pendente de análise para este ano.", "warning")
#         return redirect(url_for("acumulo.lista", ano=ano))

#     MOF = MilitarObmFuncao
#     row = (
#         db.session.query(Militar, PostoGrad.sigla.label(
#             "pg_sigla"), Obm.sigla.label("obm_sigla"))
#         .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
#         .outerjoin(MOF, and_(MOF.militar_id == Militar.id, MOF.data_fim.is_(None)))
#         .outerjoin(Obm, Obm.id == MOF.obm_id)
#         .filter(Militar.id == militar_id)
#         .first()
#     )
#     if not row:
#         abort(404)
#     militar, pg_sigla, obm_sigla = row

#     if request.method == "GET":
#         # carrega rascunho (se existir)
#         draft_row = (
#             db.session.query(DraftDeclaracaoAcumulo)
#             .filter(DraftDeclaracaoAcumulo.militar_id == militar.id,
#                     DraftDeclaracaoAcumulo.ano_referencia == ano)
#             .first()
#         )
#         draft_payload = draft_row.payload if draft_row else None

#         return render_template(
#             "acumulo_novo.html",
#             militar_id=militar.id,
#             militar=militar,
#             posto_grad_sigla=pg_sigla or "-",
#             obm_sigla=obm_sigla or "-",
#             ano=ano,
#             draft=draft_payload,
#         )

#     # ----------------- POST -----------------
#     tipo = (request.form.get("tipo") or "").strip().lower()
#     if tipo not in {"positiva", "negativa"}:
#         flash("Informe o tipo de declaração (positiva/negativa).", "alert-danger")
#         return redirect(request.url)

#     meio_entrega = (request.form.get("meio_entrega")
#                     or "digital").strip().lower()
#     if meio_entrega not in {"digital", "presencial"}:
#         meio_entrega = "digital"

#     observacoes = (request.form.get("observacoes") or "").strip()

#     # arquivo obrigatório
#     arquivo_fs = request.files.get("arquivo_declaracao")
#     if not arquivo_fs or not (arquivo_fs.filename or "").strip():
#         flash("Anexe o arquivo da declaração (PDF/JPG/PNG).", "alert-danger")
#         return redirect(request.url)

#     # vínculo único (apenas se positiva)
#     vinculo_row = None
#     if tipo == "positiva":
#         emp_nome = (request.form.get("empregador_nome") or "").strip()

#         # AGORA: esfera do órgão público (municipal/estadual/federal)
#         emp_esfera = (request.form.get("empregador_tipo")
#                       or "").strip().lower()

#         licenca = (request.form.get("licenca") or "").strip().lower()
#         emp_doc = _digits(request.form.get("empregador_doc") or "")

#         # natureza é fixa = efetivo (não ler mais do form)
#         natureza = "efetivo"

#         jornada = (request.form.get("jornada_trabalho") or "").strip().lower()
#         cargo = (request.form.get("cargo_funcao") or "").strip()
#         try:
#             carga = int(
#                 (request.form.get("carga_horaria_semanal") or "0").strip() or "0")
#         except Exception:
#             carga = 0
#         h_ini = _parse_time(request.form.get("horario_inicio"))
#         h_fim = _parse_time(request.form.get("horario_fim"))
#         d_ini = _parse_date(request.form.get("data_inicio"))

#         # validações atualizadas
#         if emp_esfera not in {"municipal", "estadual", "federal"}:
#             flash(
#                 "Esfera do órgão inválida. Use Municipal, Estadual ou Federal.", "alert-danger")
#             return redirect(request.url)
#         if jornada not in {"escala", "expediente"}:
#             flash("Jornada de trabalho do vínculo inválida.", "alert-danger")
#             return redirect(request.url)
#         if not (emp_nome and emp_doc and cargo and carga > 0 and h_ini and h_fim and d_ini):
#             flash("Preencha todos os campos do vínculo externo.", "alert-danger")
#             return redirect(request.url)

#         vinculo_row = dict(
#             empregador_nome=emp_nome,
#             empregador_tipo=emp_esfera,      # << agora guarda a ESFERA
#             licenca=licenca,
#             empregador_doc=emp_doc,
#             natureza_vinculo=natureza,       # << sempre 'efetivo'
#             jornada_trabalho=jornada,
#             cargo_funcao=cargo,
#             carga_horaria_semanal=carga,
#             horario_inicio=h_ini,
#             horario_fim=h_fim,
#             data_inicio=d_ini,
#         )

#     # ✅ Antes de salvar a declaração, atualize o User (apenas se for o próprio militar)
#     if current_user.funcao_user_id == 12:
#         # não dá pop aqui se você quiser reutilizar depois
#         email_sess = session.get('email_atualizacao')
#         obm_id_1, obm_id_2 = _obms_ativas_do_militar(militar.id)
#         localidade_id = getattr(militar, 'localidade_id', None)

#         user_changed = False
#         # e-mail: prioriza o que veio da sessão
#         if hasattr(current_user, 'email') and email_sess and current_user.email != email_sess:
#             current_user.email = email_sess
#             user_changed = True

#         # OBMs e localidade (só se as colunas existirem em User)
#         if hasattr(current_user, 'obm_id_1') and current_user.obm_id_1 != obm_id_1:
#             current_user.obm_id_1 = obm_id_1
#             user_changed = True
#         if hasattr(current_user, 'obm_id_2') and current_user.obm_id_2 != obm_id_2:
#             current_user.obm_id_2 = obm_id_2
#             user_changed = True
#         if hasattr(current_user, 'localidade_id') and current_user.localidade_id != localidade_id:
#             current_user.localidade_id = localidade_id
#             user_changed = True

#         if user_changed:
#             try:
#                 db.session.add(current_user)
#                 db.session.commit()
#             except Exception as e:
#                 db.session.rollback()
#                 # Não bloqueia o fluxo da declaração; apenas avisa
#                 flash(
#                     f"Não foi possível atualizar seus dados de usuário: {e}", "warning")

#     # upload Backblaze
#     try:
#         object_key = b2_upload_fileobj(
#             arquivo_fs, key_prefix=f"acumulo/{ano}/{militar_id}")
#     except Exception as e:
#         db.session.rollback()
#         flash(f"Falha no upload do arquivo: {e}", "alert-danger")
#         return redirect(request.url)

#     # UPSERT
#     try:
#         decl = (
#             db.session.query(DeclaracaoAcumulo)
#             .filter(
#                 DeclaracaoAcumulo.militar_id == militar_id,
#                 DeclaracaoAcumulo.ano_referencia == ano
#             ).first()
#         )

#         if decl:
#             de_status = decl.status
#             decl.tipo = tipo
#             decl.meio_entrega = meio_entrega
#             decl.observacoes = observacoes or None
#             decl.arquivo_declaracao = object_key
#             decl.status = "pendente"
#             decl.updated_at = datetime.utcnow()
#         else:
#             de_status = None
#             decl = DeclaracaoAcumulo(
#                 militar_id=militar_id,
#                 ano_referencia=ano,
#                 tipo=tipo,
#                 meio_entrega=meio_entrega,
#                 arquivo_declaracao=object_key,
#                 observacoes=observacoes or None,
#                 status="pendente",
#                 created_at=datetime.utcnow(),
#             )
#             db.session.add(decl)
#             db.session.flush()

#         # zera vínculos antigos (uma vez só)
#         for v in list(getattr(decl, "vinculos", [])):
#             db.session.delete(v)

#         # insere vínculo único se positiva
#         if tipo == "positiva" and vinculo_row:
#             db.session.add(VinculoExterno(
#                 declaracao_id=decl.id, **vinculo_row))

#         # auditoria opcional se mudou
#         if de_status and de_status != decl.status:
#             db.session.add(AuditoriaDeclaracao(
#                 declaracao_id=decl.id,
#                 de_status=de_status,
#                 para_status=decl.status,
#                 motivo="Reenvio/atualização de declaração.",
#                 alterado_por_user_id=getattr(current_user, "id", None),
#                 data_alteracao=datetime.utcnow()
#             ))

#         db.session.commit()
#         flash("Declaração salva com sucesso!", "alert-success")
#         return redirect(url_for("home", ano=ano))

#     except Exception as e:
#         db.session.rollback()
#         flash(f"Erro ao salvar a declaração: {e}", "alert-danger")
#         print(f"Erro ao salvar a declaração: {e}")
#         return redirect(request.url)


def _digits(s: str) -> str:
    return ''.join(ch for ch in (s or '') if ch.isdigit())


def _sanitize_vinculo(v: dict) -> dict:
    """
    Mantém somente os campos aceitos e aplica regras novas:
    - empregador_tipo: municipal/estadual/federal
    - natureza_vinculo: sempre 'efetivo'
    """
    v = v or {}
    esfera = (v.get('empregador_tipo') or '').strip().lower()
    if esfera not in {'municipal', 'estadual', 'federal', ''}:
        esfera = ''  # não invalidamos rascunho — só normalizamos

    licenca = (v.get('licenca') or '').strip().lower()

    return {
        'empregador_nome': (v.get('empregador_nome') or '').strip(),
        # municipal | estadual | federal
        'empregador_tipo': esfera,
        'licenca': licenca,
        'empregador_doc': _digits(v.get('empregador_doc') or ''),
        'natureza_vinculo': 'efetivo',                        # fixo
        'jornada_trabalho': (v.get('jornada_trabalho') or '').strip().lower(),
        'cargo_funcao': (v.get('cargo_funcao') or '').strip(),
        'carga_horaria_semanal': (v.get('carga_horaria_semanal') or '').strip(),
        'horario_inicio': (v.get('horario_inicio') or '').strip(),
        'horario_fim': (v.get('horario_fim') or '').strip(),
        'data_inicio': (v.get('data_inicio') or '').strip(),
    }


@bp_acumulo.route("/salvar-rascunho/<int:militar_id>", methods=["POST"])
@login_required
def salvar_rascunho(militar_id):
    try:
        # segurança: usuário comum só salva do próprio militar
        if current_user.funcao_user_id == 12:
            mil_user = get_militar_por_user(current_user)
            if not mil_user or mil_user.id != militar_id:
                return jsonify(ok=False, error="Sem permissão para este militar."), 403

        data = request.get_json(silent=True) or {}
        ano = int(data.get("ano") or datetime.now().year)

        # tipo pode estar vazio (rascunho!), mas se vier, normalize
        tipo = (data.get("tipo") or '').strip().lower()
        if tipo not in {'', 'positiva', 'negativa'}:
            tipo = ''  # não invalidar rascunho

        observacoes = (data.get("observacoes") or '').strip()

        # vínculos: lista; mantemos só o 1º por enquanto (sua UI atual é 1)
        vinculos_in = data.get("vinculos") or []
        vinculos = []
        for v in vinculos_in:
            vinculos.append(_sanitize_vinculo(v))

        payload = {
            'ano': ano,
            'tipo': tipo,
            'observacoes': observacoes,
            'vinculos': vinculos,
        }

        # upsert (1 rascunho por militar/ano)
        draft = (db.session.query(DraftDeclaracaoAcumulo)
                 .filter(DraftDeclaracaoAcumulo.militar_id == militar_id,
                         DraftDeclaracaoAcumulo.ano_referencia == ano)
                 .first())
        if draft:
            draft.payload = payload
            draft.updated_at = datetime.utcnow()
            db.session.add(draft)
        else:
            draft = DraftDeclaracaoAcumulo(
                militar_id=militar_id,
                ano_referencia=ano,
                payload=payload,
            )
            db.session.add(draft)

        db.session.commit()
        return jsonify(ok=True), 200

    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, error=str(e)), 500


@bp_acumulo.route("/prepara-geracao", methods=["POST"])
@login_required
def prepara_geracao():
    try:
        ano = request.form.get("ano", type=int) or datetime.now().year
        militar_id = request.form.get("militar_id", type=int)
        tipo = (request.form.get("tipo") or "").strip().lower()
        if tipo not in {"positiva", "negativa"}:
            return {"ok": False, "error": "Tipo inválido."}, 400

        # força o próprio militar se for usuário comum
        if current_user.funcao_user_id == 12:
            mil_user = get_militar_por_user(current_user)
            if not mil_user:
                return {"ok": False, "error": "Militar não encontrado para o usuário."}, 400
            militar_id = mil_user.id

        militar = Militar.query.filter_by(id=militar_id).first()
        if not militar:
            return {"ok": False, "error": "Militar não encontrado."}, 404

        # 1) atualizar USER (email vindo da sessão / OBMs / localidade)
        email_sess = session.get('email_atualizacao')

        MOF = MilitarObmFuncao
        rows = (db.session.query(MOF.obm_id)
                .filter(and_(MOF.militar_id == militar.id, MOF.data_fim.is_(None)))
                .order_by(MOF.id.asc()).all())
        ids = [r.obm_id for r in rows]
        obm_id_1 = ids[0] if len(ids) > 0 else None
        obm_id_2 = ids[1] if len(ids) > 1 else None
        localidade_id = getattr(militar, 'localidade_id', None)

        changed = False
        if hasattr(current_user, 'email') and email_sess and current_user.email != email_sess:
            current_user.email = email_sess
            changed = True
        if hasattr(current_user, 'obm_id_1') and current_user.obm_id_1 != obm_id_1:
            current_user.obm_id_1 = obm_id_1
            changed = True
        if hasattr(current_user, 'obm_id_2') and current_user.obm_id_2 != obm_id_2:
            current_user.obm_id_2 = obm_id_2
            changed = True
        if hasattr(current_user, 'localidade_id') and current_user.localidade_id != localidade_id:
            current_user.localidade_id = localidade_id
            changed = True
        if changed:
            db.session.add(current_user)
            db.session.commit()

        # 2) guardar dados na sessão (chave única militar+ano)
        pre = {
            "ano": ano,
            "tipo": tipo,
            "observacoes": (request.form.get("observacoes") or "").strip(),
        }
        if tipo == "positiva":
            pre.update({
                "empregador_nome": (request.form.get("empregador_nome") or "").strip(),
                "empregador_tipo": (request.form.get("empregador_tipo") or "").strip().lower(),
                "licenca": (request.form.get("licenca") or "").strip().lower(),
                "empregador_doc": ''.join(filter(str.isdigit, request.form.get("empregador_doc") or "")),
                "natureza_vinculo": (request.form.get("natureza_vinculo") or "").strip().lower(),
                "jornada_trabalho": (request.form.get("jornada_trabalho") or "").strip().lower(),
                "cargo_funcao": (request.form.get("cargo_funcao") or "").strip(),
                "carga_horaria_semanal": (request.form.get("carga_horaria_semanal") or "").strip(),
                "horario_inicio": (request.form.get("horario_inicio") or "").strip(),
                "horario_fim": (request.form.get("horario_fim") or "").strip(),
                "data_inicio": (request.form.get("data_inicio") or "").strip(),
            })
        session[f'pre_decl_{militar.id}_{ano}'] = pre

        return {"ok": True}, 200
    except Exception as e:
        db.session.rollback()
        return {"ok": False, "error": str(e)}, 500


@bp_acumulo.route("/upload-assinado/<int:militar_id>", methods=["GET", "POST"])
@login_required
def upload_assinado(militar_id):
    # segurança: militar comum só o próprio
    if current_user.funcao_user_id == 12:
        mil_user = get_militar_por_user(current_user)
        if not mil_user or mil_user.id != militar_id:
            flash("Você não pode enviar arquivos desta declaração.", "alert-danger")
            return redirect(url_for("home_atualizacao"))

    militar = Militar.query.filter_by(id=militar_id).first_or_404()
    ano = request.args.get("ano", type=int) or datetime.now().year

    pre = session.get(f'pre_decl_{militar.id}_{ano}')
    if not pre:
        flash(
            "Dados da declaração não encontrados. Gere o modelo novamente.", "alert-warning")
        return redirect(url_for("acumulo.novo", militar_id=militar.id))

    tipo = pre["tipo"]
    observacoes = pre.get("observacoes")

    if request.method == "GET":
        return render_template(
            "acumulo_upload_assinado.html",
            militar=militar, ano=ano, tipo=tipo
        )

    # ===== 2 ARQUIVOS =====
    # 1) Modelo DOCX gerado por vocês + assinado pelo militar (sempre obrigatório)
    arquivo_modelo_fs = request.files.get("arquivo_modelo_assinado")
    if not arquivo_modelo_fs or not (arquivo_modelo_fs.filename or "").strip():
        flash("Anexe o modelo assinado pelo militar (PDF).", "alert-danger")
        return redirect(request.url)

    # 2) Declaração do órgão público (exigida apenas para tipo 'positiva')
    arquivo_orgao_fs = request.files.get("arquivo_declaracao_orgao")
    if tipo == "positiva":
        if not arquivo_orgao_fs or not (arquivo_orgao_fs.filename or "").strip():
            flash("Anexe a declaração emitida pelo órgão público (PDF).",
                  "alert-danger")
            return redirect(request.url)
    else:
        # tipo 'negativa' ignora segundo arquivo se vier vazio
        if arquivo_orgao_fs and not (arquivo_orgao_fs.filename or "").strip():
            arquivo_orgao_fs = None

    vinculo_row = None
    if tipo == "positiva":
        from datetime import datetime as _dt, time as _time

        def _p_time(s):
            s = (s or "").strip()
            if not s:
                return None
            try:
                h, m = s.split(":")
                return _time(int(h), int(m))
            except Exception:
                return None

        def _p_date(s):
            s = (s or "").strip()
            if not s:
                return None
            try:
                return _dt.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return None

        def _p_int(s):
            s = (s or "").strip()
            if not s:
                return None
            try:
                return int(s)
            except Exception:
                return None

        def _norm_lic(v):
            v = (v or "").strip().lower()
            if v in ("nao", "não", "n", "no"):
                return "não"
            if v in ("sim", "s", "yes", "y"):
                return "sim"
            return ""

        lic = _norm_lic(pre.get("licenca"))
        is_licenca = (lic == "sim")

        # === Carga horária: NOT NULL no banco -> use 0 quando estiver de licença ===
        carga = _p_int(pre.get("carga_horaria_semanal"))
        if is_licenca:
            carga = 0
        elif carga is None:
            # Se NÃO estiver de licença, exija valor válido
            flash("Informe a carga horária semanal.", "alert-danger")
            return redirect(request.url)

        # === Horários: NOT NULL no banco -> use 00:00 quando estiver de licença ===
        if is_licenca:
            h_ini = _time(0, 0)
            h_fim = _time(0, 0)
            jornada = None  # pode ser None; se seu banco exigir NOT NULL aqui também, troque para 'escala' ou 'expediente'
        else:
            h_ini = _p_time(pre.get("horario_inicio"))
            h_fim = _p_time(pre.get("horario_fim"))
            jornada = (pre.get("jornada_trabalho")
                       or "").strip().lower() or None
            # Valida se não estiver de licença
            if h_ini is None or h_fim is None:
                flash("Informe horário de início e fim.", "alert-danger")
                return redirect(request.url)
            if not jornada:
                flash("Informe a jornada de trabalho.", "alert-danger")
                return redirect(request.url)

        vinculo_row = dict(
            empregador_nome=(pre.get("empregador_nome") or "").strip() or None,
            empregador_tipo=(pre.get("empregador_tipo") or "").strip() or None,
            licenca=lic,  # "sim" ou "não"
            empregador_doc=(pre.get("empregador_doc") or "").strip() or None,
            natureza_vinculo=(pre.get("natureza_vinculo")
                              or "").strip() or None,

            cargo_funcao=(pre.get("cargo_funcao") or "").strip() or None,
            jornada_trabalho=jornada,

            carga_horaria_semanal=carga,
            horario_inicio=h_ini,
            horario_fim=h_fim,
            data_inicio=_p_date(pre.get("data_inicio")),
        )

    # ===== Upload Backblaze (mesma pasta) =====
    try:
        prefix = build_prefix(ano, militar.id)  # "acumulo/{ano}/{militar_id}"
        key_modelo = b2_upload_fileobj(arquivo_modelo_fs, key_prefix=prefix)
        key_orgao = b2_upload_fileobj(
            arquivo_orgao_fs, key_prefix=prefix) if arquivo_orgao_fs else None
    except Exception as e:
        db.session.rollback()
        flash(f"Falha no upload do arquivo: {e}", "alert-danger")
        return redirect(request.url)

    # ===== UPSERT =====
    try:
        decl = (db.session.query(DeclaracaoAcumulo)
                .filter(DeclaracaoAcumulo.militar_id == militar.id,
                        DeclaracaoAcumulo.ano_referencia == ano)
                .first())

        if decl:
            de_status = decl.status
            decl.tipo = tipo
            decl.meio_entrega = "digital"
            decl.observacoes = observacoes or None

            # grava os 2 arquivos (modelo sempre; órgão só se existir)
            decl.arquivo_declaracao = key_modelo
            if key_orgao:
                decl.arquivo_declaracao_orgao = key_orgao

            decl.status = "pendente"
            decl.updated_at = datetime.utcnow()
        else:
            de_status = None
            decl = DeclaracaoAcumulo(
                militar_id=militar.id,
                ano_referencia=ano,
                tipo=tipo,
                meio_entrega="digital",
                arquivo_declaracao=key_modelo,
                arquivo_declaracao_orgao=key_orgao,
                observacoes=observacoes or None,
                status="pendente",
                created_at=datetime.utcnow(),
            )
            db.session.add(decl)
            db.session.flush()

        # vínculos: limpa e regrava (1) quando positiva
        for v in list(getattr(decl, "vinculos", [])):
            db.session.delete(v)
        if tipo == "positiva" and vinculo_row:
            db.session.add(VinculoExterno(
                declaracao_id=decl.id, **vinculo_row))

        # auditoria de mudança de status (se houve)
        if de_status and de_status != decl.status:
            db.session.add(AuditoriaDeclaracao(
                declaracao_id=decl.id,
                de_status=de_status,
                para_status=decl.status,
                motivo="Envio de documentos assinados (modelo e órgão).",
                alterado_por_user_id=getattr(current_user, "id", None),
                data_alteracao=datetime.utcnow()
            ))

        # negativas: marcar "enviado_drh" (uma única vez)
        if tipo == "negativa":
            ja = (db.session.query(AuditoriaDeclaracao.id)
                  .filter(AuditoriaDeclaracao.declaracao_id == decl.id,
                          AuditoriaDeclaracao.motivo == "enviado_drh")
                  .first())
            if not ja:
                db.session.add(AuditoriaDeclaracao(
                    declaracao_id=decl.id,
                    de_status=decl.status,
                    para_status=decl.status,  # permanece 'pendente'
                    motivo="enviado_drh",
                    alterado_por_user_id=getattr(current_user, "id", None),
                    data_alteracao=datetime.utcnow()
                ))

        db.session.commit()

        # limpa cache da sessão + apaga rascunho do ano
        session.pop(f'pre_decl_{militar.id}_{ano}', None)
        try:
            draft = (db.session.query(DraftDeclaracaoAcumulo)
                     .filter(DraftDeclaracaoAcumulo.militar_id == militar.id,
                             DraftDeclaracaoAcumulo.ano_referencia == ano)
                     .first())
            if draft:
                db.session.delete(draft)
                db.session.commit()
        except Exception:
            db.session.rollback()

        flash("Declaração enviada com sucesso!", "alert-success")
        return redirect(url_for("home_atualizacao", ano=ano))

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao salvar a declaração: {e}", "alert-danger")
        return redirect(request.url)


@bp_acumulo.route("/editar/<int:decl_id>", methods=["GET", "POST"])
@login_required
def editar(decl_id):
    # carrega a declaração com militar e vínculos
    decl = (db.session.query(DeclaracaoAcumulo)
            .options(
                joinedload(DeclaracaoAcumulo.militar).joinedload(
                    Militar.posto_grad),
                selectinload(DeclaracaoAcumulo.vinculos)
    ).get(decl_id))
    if not decl:
        abort(404)

    # segurança: usuário comum só edita a sua
    if current_user.funcao_user_id == 12:
        mil_user = get_militar_por_user(current_user)
        if not mil_user or mil_user.id != decl.militar_id:
            flash("Sem permissão para editar esta declaração.", "alert-danger")
            return redirect(url_for("home_atualizacao"))

    st = (decl.status or "").lower()
    if _prazo_fechado() and st != "inconforme" and not _is_drh_like():
        flash("Prazo encerrado. Edição liberada apenas para declarações INCONFORME.", "warning")
        return redirect(url_for("acumulo.minhas_declaracoes", ano=decl.ano_referencia))

    # SIGLA da OBM ativa (só pra exibição)
    MOF = MilitarObmFuncao
    obm_sigla = (db.session.query(Obm.sigla)
                 .join(MOF, Obm.id == MOF.obm_id)
                 .filter(MOF.militar_id == decl.militar_id, MOF.data_fim.is_(None))
                 .limit(1).scalar()) or "-"

    if request.method == "GET":
        # link temporário do arquivo atual (se houver)
        url_arquivo = None
        if decl.arquivo_declaracao:
            url_arquivo = b2_presigned_get(
                decl.arquivo_declaracao, expires_seconds=600)

        return render_template(
            "acumulo_editar.html",
            decl=decl,
            militar=decl.militar,
            posto_grad_sigla=getattr(decl.militar.posto_grad, "sigla", "-"),
            obm_sigla=obm_sigla,
            ano=decl.ano_referencia,
            url_arquivo=url_arquivo,
        )

    # ----------------- POST -----------------
    ano = decl.ano_referencia
    militar = decl.militar

    tipo = (request.form.get("tipo") or "").strip().lower()
    if tipo not in {"positiva", "negativa"}:
        flash("Informe o tipo de declaração (positiva/negativa).", "alert-danger")
        return redirect(request.url)

    observacoes = (request.form.get("observacoes") or "").strip()

    # === Atualiza User (igual ao prepara_geracao) — não bloqueia fluxo ===
    try:
        email_sess = session.get('email_atualizacao')

        rows = (db.session.query(MOF.obm_id)
                .filter(and_(MOF.militar_id == militar.id, MOF.data_fim.is_(None)))
                .order_by(MOF.id.asc()).all())
        ids = [r.obm_id for r in rows]
        obm_id_1 = ids[0] if len(ids) > 0 else None
        obm_id_2 = ids[1] if len(ids) > 1 else None
        localidade_id = getattr(militar, 'localidade_id', None)

        changed = False
        if hasattr(current_user, 'email') and email_sess and current_user.email != email_sess:
            current_user.email = email_sess
            changed = True
        if hasattr(current_user, 'obm_id_1') and current_user.obm_id_1 != obm_id_1:
            current_user.obm_id_1 = obm_id_1
            changed = True
        if hasattr(current_user, 'obm_id_2') and current_user.obm_id_2 != obm_id_2:
            current_user.obm_id_2 = obm_id_2
            changed = True
        if hasattr(current_user, 'localidade_id') and current_user.localidade_id != localidade_id:
            current_user.localidade_id = localidade_id
            changed = True
        if changed:
            db.session.add(current_user)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(
            f"Não foi possível atualizar seus dados de usuário: {e}", "warning")

    # === Monta pacote pré-declaração p/ sessão (fluxo de geração DOCX) ===
    pre = {
        "ano": ano,
        "tipo": tipo,
        "observacoes": observacoes,
    }

    if tipo == "positiva":
        nomes = request.form.getlist("empregador_nome[]")
        # municipal/estadual/federal
        tipos = request.form.getlist("empregador_tipo[]")

        licencas = request.form.getlist("licenca[]")
        docs = request.form.getlist("empregador_doc[]")
        jornada = request.form.getlist("jornada_trabalho[]")
        cargos = request.form.getlist("cargo_funcao[]")
        cargas = request.form.getlist("carga_horaria_semanal[]")
        h_ini = request.form.getlist("horario_inicio[]")
        h_fim = request.form.getlist("horario_fim[]")
        d_ini = request.form.getlist("data_inicio[]")

        def first_or_empty(lst, i):
            return (lst[i] if i < len(lst) else "").strip()

        # pega o primeiro vínculo com nome preenchido
        idx = None
        for i, nome in enumerate(nomes or []):
            if (nome or "").strip():
                idx = i
                break
        if idx is None:
            flash(
                "Inclua ao menos um vínculo público para declaração positiva.", "alert-danger")
            return redirect(request.url)

        esfera = first_or_empty(tipos, idx).lower()
        if esfera not in {"municipal", "estadual", "federal"}:
            flash(
                "Tipo do órgão inválido (use Municipal/Estadual/Federal).", "alert-danger")
            return redirect(request.url)

        # natureza = sempre efetivo
        pre.update({
            "empregador_nome": first_or_empty(nomes, idx),
            "empregador_tipo": esfera,
            "empregador_doc": _digits(first_or_empty(docs, idx)),
            "licenca": first_or_empty(licencas, idx).lower(),
            "natureza_vinculo": "efetivo",
            "jornada_trabalho": first_or_empty(jornada, idx).lower(),
            "cargo_funcao": first_or_empty(cargos, idx),
            "carga_horaria_semanal": (first_or_empty(cargas, idx) or "0").strip(),
            "horario_inicio": first_or_empty(h_ini, idx)[:5],
            "horario_fim":    first_or_empty(h_fim, idx)[:5],
            "data_inicio":    first_or_empty(d_ini, idx)[:10],
        })

    session[f'pre_decl_{militar.id}_{ano}'] = pre

    url_arquivo_modelo = b2_presigned_get(
        decl.arquivo_declaracao, 600) if decl.arquivo_declaracao else None
    url_arquivo_orgao = b2_presigned_get(decl.arquivo_declaracao_orgao, 600) if getattr(
        decl, "arquivo_declaracao_orgao", None) else None

    return render_template(
        "acumulo_editar.html",
        decl=decl,
        militar=decl.militar,
        posto_grad_sigla=getattr(decl.militar.posto_grad, "sigla", "-"),
        obm_sigla=obm_sigla,
        ano=decl.ano_referencia,
        url_arquivo_modelo=url_arquivo_modelo,
        url_arquivo_orgao=url_arquivo_orgao,
    )


@bp_acumulo.route("/recebimento", methods=["GET"])
@login_required
@require_analise_vinculo
def recebimento():
    ano = request.args.get("ano", type=int) or datetime.now().year
    status = (request.args.get("status") or "todos").lower()
    q = (request.args.get("q") or "").strip()
    obm_id = request.args.get("obm_id", type=int)
    page = max(request.args.get("page", 1, type=int), 1)
    per_page = max(request.args.get("per_page", 20, type=int), 1)

    M, D, PG, MOF, O, F = Militar, DeclaracaoAcumulo, PostoGrad, MilitarObmFuncao, Obm, Funcao
    IS_AUDITOR = getattr(current_user, "id", None) in USERS_ANALISE_VINCULO  # ou _is_auditor_vinculo()

    IS_DRH_LIKE = _is_drh_like() or IS_AUDITOR      # <- bypass do filtro por OBM
    IS_SUPER = _is_super_user()
    IS_PRIV = _is_privilegiado() or IS_AUDITOR      # <- “vê tudo” nas regras DRH

    rn = func.row_number().over(
        partition_by=D.militar_id,
        order_by=(D.updated_at.desc().nullslast(), D.id.desc())
    ).label("rn")

    ultimo_subq = (
        db.session.query(
            D.militar_id.label("m_id"),
            D.id.label("decl_id"),
            D.status.label("decl_status"),
            D.tipo.label("decl_tipo"),
            D.meio_entrega.label("decl_meio"),
            D.arquivo_declaracao.label("decl_arquivo_modelo"),
            D.arquivo_declaracao_orgao.label("decl_arquivo_orgao"),
            rn
        )
        .filter(D.ano_referencia == ano)
        .subquery()
    )

    base_q = (
        db.session.query(M.id.label("m_id"))
        .join(MOF, and_(MOF.militar_id == M.id, MOF.data_fim.is_(None)))
        .join(F, F.id == MOF.funcao_id, isouter=True)
        .filter(M.inativo.is_(False))
    )
    if obm_id:
        base_q = base_q.filter(MOF.obm_id == obm_id)
    if not IS_DRH_LIKE:
        obms_user = [getattr(current_user, "obm_id_1", None),
                     getattr(current_user, "obm_id_2", None)]
        obms_user = [x for x in obms_user if x]
        base_q = base_q.filter(MOF.obm_id.in_(
            obms_user)) if obms_user else base_q.filter(literal(False))
    if q:
        ilike = f"%{q}%"
        base_q = (base_q.join(PG, PG.id == M.posto_grad_id, isouter=True)
                        .filter(or_(M.nome_completo.ilike(ilike),
                                    M.matricula.ilike(ilike),
                                    PG.sigla.ilike(ilike))))
    base_q = base_q.distinct()

    total_militares = base_q.order_by(None).count()
    if total_militares == 0:
        return render_template(
            "acumulo_recebimento.html",
            ano=ano, status=status, q=q, obm_id=obm_id,
            obms=db.session.query(O).order_by(O.sigla).all(),
            linhas=[], url_map={},
            total_ano=0, total_pendentes=0, total_validados=0, total_inconformes=0, total_nao_enviaram=0,
            pagination="", pode_editar_map={}, enviado_drh_map={},
            is_drh_like=IS_DRH_LIKE,
            is_super_user=IS_SUPER,
        )

    base_page_q = (
        db.session.query(M.id.label("m_id"), M.nome_completo.label("nome"))
        .join(MOF, and_(MOF.militar_id == M.id, MOF.data_fim.is_(None)))
        .join(F, F.id == MOF.funcao_id, isouter=True)
        .filter(M.inativo.is_(False))
    )
    if obm_id:
        base_page_q = base_page_q.filter(MOF.obm_id == obm_id)
    if not IS_DRH_LIKE:
        obms_user = [getattr(current_user, "obm_id_1", None),
                     getattr(current_user, "obm_id_2", None)]
        obms_user = [x for x in obms_user if x]
        base_page_q = base_page_q.filter(MOF.obm_id.in_(
            obms_user)) if obms_user else base_page_q.filter(literal(False))
    if q:
        ilike = f"%{q}%"
        base_page_q = (base_page_q.join(PG, PG.id == M.posto_grad_id, isouter=True)
                       .filter(or_(M.nome_completo.ilike(ilike),
                                   M.matricula.ilike(ilike),
                                   PG.sigla.ilike(ilike))))

    base_page_q = base_page_q.outerjoin(ultimo_subq, and_(
        ultimo_subq.c.m_id == M.id, ultimo_subq.c.rn == 1))

    enc_exists = exists().where(
        and_(
            AuditoriaDeclaracao.declaracao_id == ultimo_subq.c.decl_id,
            AuditoriaDeclaracao.motivo == "enviado_drh"
        )
    )

    # VISIBILIDADE
    if not IS_DRH_LIKE:
        # Chefia/OBM: NÃO veem negativas; positivas sempre
        base_page_q = base_page_q.filter(
            or_(ultimo_subq.c.decl_id.is_(None),
                ultimo_subq.c.decl_tipo != "negativa")
        )
    else:
        if IS_PRIV:
            # SUPER USER / CHEFE DRH: vê TUDO
            pass
        else:
            # DRH normal: vê "não enviou", TODAS as negativas e
            # positivas SOMENTE se encaminhadas
            base_page_q = base_page_q.filter(
                or_(
                    # <<< reintroduz "não enviou"
                    ultimo_subq.c.decl_id.is_(None),
                    ultimo_subq.c.decl_tipo == "negativa",
                    enc_exists                         # positiva encaminhada
                )
            )

    # Chegada à DRH = Somente encaminhadas
    enc_only = (request.args.get("enc") == "1")
    if IS_DRH_LIKE and enc_only:
        base_page_q = base_page_q.filter(
            or_(
                # negativas contam como "encaminhadas"
                ultimo_subq.c.decl_tipo == "negativa",
                # positivas somente se houver registro "enviado_drh"
                and_(ultimo_subq.c.decl_tipo == "positiva", enc_exists)
            )
        )

    if status == "pendente":
        base_page_q = base_page_q.filter(
            ultimo_subq.c.decl_status == "pendente")
    elif status == "validado":
        base_page_q = base_page_q.filter(
            ultimo_subq.c.decl_status == "validado")
    elif status == "inconforme":
        base_page_q = base_page_q.filter(
            ultimo_subq.c.decl_status == "inconforme")
    elif status == "nao_enviou":
        base_page_q = base_page_q.filter(ultimo_subq.c.decl_id.is_(None))
    # "todos" -> sem filtro adicional

    base_page_q = base_page_q.distinct()

    # Conte o TOTAL já FILTRADO para a paginação
    total_filtrado = base_page_q.order_by(None).count()

    # Paginado
    page_rows = (base_page_q.order_by(M.nome_completo.asc())
                 .limit(per_page).offset((page - 1) * per_page).all())
    page_ids = [r.m_id for r in page_rows]

    rn_enc = func.row_number().over(
        partition_by=AuditoriaDeclaracao.declaracao_id,
        order_by=AuditoriaDeclaracao.data_alteracao.desc()
    ).label("rn")

    enc_subq = (
        db.session.query(
            AuditoriaDeclaracao.declaracao_id.label("decl_id"),
            AuditoriaDeclaracao.alterado_por_user_id.label("enc_uid"),
            AuditoriaDeclaracao.data_alteracao.label("enc_quando"),
            rn_enc
        )
        .filter(AuditoriaDeclaracao.motivo == "enviado_drh")
        .subquery()
    )

    rn_val = func.row_number().over(
        partition_by=AuditoriaDeclaracao.declaracao_id,
        order_by=AuditoriaDeclaracao.data_alteracao.desc()
    ).label("rn")

    val_subq = (
        db.session.query(
            AuditoriaDeclaracao.declaracao_id.label("decl_id"),
            AuditoriaDeclaracao.alterado_por_user_id.label("val_uid"),
            AuditoriaDeclaracao.data_alteracao.label("val_quando"),
            rn_val
        )
        .filter(AuditoriaDeclaracao.para_status == "validado")
        .subquery()
    )

    rn_inc = func.row_number().over(
        partition_by=AuditoriaDeclaracao.declaracao_id,
        order_by=AuditoriaDeclaracao.data_alteracao.desc()
    ).label("rn")

    inc_subq = (
        db.session.query(
            AuditoriaDeclaracao.declaracao_id.label("decl_id"),
            AuditoriaDeclaracao.alterado_por_user_id.label("inc_uid"),
            AuditoriaDeclaracao.data_alteracao.label("inc_quando"),
            rn_inc
        )
        .filter(AuditoriaDeclaracao.para_status == "inconforme")
        .subquery()
    )

    U3 = aliased(User)
    U = User
    U2 = aliased(User)
    V = VinculoExterno

    lic_count_subq = (
        db.session.query(
            V.declaracao_id.label("decl_id"),
            func.sum(case((V.licenca == "sim", 1), else_=0)).label("lic_sim"),
            func.count(V.id).label("lic_total"),
        )
        .group_by(V.declaracao_id)
        .subquery()
    )

    rows = (
        db.session.query(
            M, PG.sigla.label("pg_sigla"), O.sigla.label("obm_sigla"),
            ultimo_subq.c.decl_id, ultimo_subq.c.decl_status, ultimo_subq.c.decl_tipo,
            ultimo_subq.c.decl_meio,
            ultimo_subq.c.decl_arquivo_modelo, ultimo_subq.c.decl_arquivo_orgao,

            # quem encaminhou
            U.nome.label("encaminhado_por"),
            enc_subq.c.enc_quando.label("encaminhado_quando"),

            # quem validou
            U2.nome.label("validado_por"),
            val_subq.c.val_quando.label("validado_quando"),

            # inconforme  <<< NOVO
            U3.nome.label("inconforme_por"),
            inc_subq.c.inc_quando.label("inconforme_quando"),

            lic_count_subq.c.lic_sim,
            lic_count_subq.c.lic_total,
        )
        .join(MOF, and_(MOF.militar_id == M.id, MOF.data_fim.is_(None)))
        .join(O, O.id == MOF.obm_id)
        .join(PG, PG.id == M.posto_grad_id, isouter=True)
        .outerjoin(ultimo_subq, and_(ultimo_subq.c.m_id == M.id, ultimo_subq.c.rn == 1))
        .outerjoin(enc_subq, and_(enc_subq.c.decl_id == ultimo_subq.c.decl_id, enc_subq.c.rn == 1))
        .outerjoin(U, U.id == enc_subq.c.enc_uid)

        # joins da validação
        .outerjoin(val_subq, and_(val_subq.c.decl_id == ultimo_subq.c.decl_id, val_subq.c.rn == 1))
        .outerjoin(U2, U2.id == val_subq.c.val_uid)
        .outerjoin(inc_subq, and_(inc_subq.c.decl_id == ultimo_subq.c.decl_id, inc_subq.c.rn == 1))
        .outerjoin(U3, U3.id == inc_subq.c.inc_uid)
        .outerjoin(lic_count_subq, lic_count_subq.c.decl_id == ultimo_subq.c.decl_id)
        .filter(M.id.in_(page_ids))
        .order_by(M.nome_completo.asc())
        .all()
    )

    decl_ids = [r.decl_id for r in rows if r.decl_id]
    enviado_drh_map = {}
    if decl_ids:
        enviado_drh_set = {aid for (aid,) in (
            db.session.query(AuditoriaDeclaracao.declaracao_id)
            .filter(AuditoriaDeclaracao.declaracao_id.in_(decl_ids),
                    AuditoriaDeclaracao.motivo == "enviado_drh")
            .distinct()
            .all()
        )}
        enviado_drh_map = {did: (did in enviado_drh_set) for did in decl_ids}

    elegiveis_cte = base_q.subquery()
    kpi_q = (
        db.session.query(
            func.count().label("total"),
            func.sum(case((ultimo_subq.c.decl_id.isnot(None), 1), else_=0)).label(
                "enviaram"),
            func.sum(case((ultimo_subq.c.decl_status == "pendente", 1), else_=0)).label(
                "pendentes"),
            func.sum(case((ultimo_subq.c.decl_status == "validado", 1), else_=0)).label(
                "validados"),
            func.sum(case((ultimo_subq.c.decl_status == "inconforme", 1), else_=0)).label(
                "inconformes"),
        )
        .select_from(elegiveis_cte)
        .outerjoin(ultimo_subq, and_(ultimo_subq.c.m_id == elegiveis_cte.c.m_id, ultimo_subq.c.rn == 1))
    )

    if not IS_DRH_LIKE:
        kpi_q = kpi_q.filter(or_(ultimo_subq.c.decl_id.is_(
            None), ultimo_subq.c.decl_tipo != "negativa"))

    kpi_rows = kpi_q.one()

    total_ano = kpi_rows.total or 0
    enviados = kpi_rows.enviaram or 0
    total_pendentes = kpi_rows.pendentes or 0
    total_validados = kpi_rows.validados or 0
    total_inconformes = kpi_rows.inconformes or 0
    total_nao_enviaram = max(total_ano - enviados, 0)

    url_map, linhas = {}, []
    for (militar, pg_sigla, obm_sigla, decl_id, st, tp, me, arq_modelo, arq_orgao,
         enc_por, enc_quando, val_por, val_quando,
         inc_por, inc_quando,
         lic_sim, lic_total) in rows:

        if decl_id:
            urls = {}
            if arq_modelo:
                urls["modelo"] = b2_presigned_get(arq_modelo, 600)
            if arq_orgao:
                urls["orgao"] = b2_presigned_get(arq_orgao, 600)
            url_map[decl_id] = urls

        # regra: "está de licença" = existe algum vínculo com licenca == 'sim'
        licenca_flag = bool(lic_sim and lic_sim >
                            0) if decl_id and tp == "positiva" else None

        linhas.append(dict(
            militar=militar,
            posto_grad_sigla=pg_sigla or "-",
            obm_sigla=obm_sigla or "-",
            decl_id=decl_id,
            decl_status=st,
            decl_tipo=tp,
            decl_meio=me,

            encaminhado_por=enc_por or "",
            encaminhado_quando=enc_quando,
            encaminhado_flag=bool(enc_quando),

            validado_por=val_por or "",
            validado_quando=val_quando,
            validado_flag=bool(val_quando),

            inconforme_por=inc_por or "",
            inconforme_quando=inc_quando,
            inconforme_flag=bool(inc_quando),

            # True = tem algum SIM; False = (0 SIM); None = não aplicável
            licenca_flag=licenca_flag,
            lic_sim=lic_sim or 0,
            lic_total=lic_total or 0,
        ))

    def render_pagination(total, page, per_page):
        pages = max((total + per_page - 1) // per_page, 1)
        if pages <= 1:
            return ""
        items = []
        for num in range(1, pages + 1):
            cls = "active" if num == page else ""
            params = f"page={num}&per_page={per_page}&ano={ano}&status={status}&q={q}&obm_id={(obm_id or '')}"
            items.append(
                f'<li class="page-item {cls}"><a class="page-link" href="?{params}">{num}</a></li>')
        return '<nav><ul class="pagination pagination-sm justify-content-end mt-3">' + "".join(items) + "</ul></nav>"

    pagination_html = render_pagination(total_filtrado, page, per_page)

    pode_editar_map = {decl_id: (IS_DRH_LIKE or _can_editar_declaracao(m.id))
                       for (m, _, _, decl_id, *_rest) in rows if decl_id}

    obms = db.session.query(O).order_by(O.sigla.asc()).all()
    print('IS_DRH_LIKE=', IS_DRH_LIKE, 'IS_PRIV=',
          IS_PRIV, 'enc_only=', request.args.get("enc"))

    return render_template(
        "acumulo_recebimento.html",
        ano=ano, status=status, q=q, obm_id=obm_id, obms=obms,
        linhas=linhas, url_map=url_map,
        total_ano=total_ano, total_pendentes=total_pendentes,
        total_validados=total_validados, total_inconformes=total_inconformes,
        total_nao_enviaram=total_nao_enviaram,
        total_filtrado=total_filtrado,
        pagination=pagination_html,
        pode_editar_map=pode_editar_map,
        enviado_drh_map=enviado_drh_map,
        is_drh_like=IS_DRH_LIKE,
        is_super_user=IS_SUPER,
    )


@bp_acumulo.route("/recebimento/<int:decl_id>/status", methods=["POST"])
@login_required
@require_analise_vinculo
def recebimento_mudar_status(decl_id):
    novo = (request.form.get("status") or "").lower()
    if novo not in {"validado", "inconforme", "pendente", "enviar_drh"}:
        flash("Status inválido.", "alert-danger")
        return redirect(url_for("acumulo.recebimento"))

    decl = db.session.get(DeclaracaoAcumulo, decl_id) or abort(404)
    IS_DRH_LIKE = _is_drh_like()
    IS_SUPER = _is_super_user()

    # chefia NÃO mexe em negativas (nem encaminhar)
    if decl.tipo == "negativa" and not IS_DRH_LIKE:
        flash("Declarações negativas seguem direto para a DRH.", "alert-warning")
        return redirect(url_for("acumulo.recebimento"))

    # chefia comum só pode agir sobre seus militares
    if not IS_DRH_LIKE and not _militar_permitido(decl.militar_id):
        abort(403)

    # depois de carregar decl e conferir 'validado'
    if novo == "validado" and IS_DRH_LIKE and not IS_SUPER:
        enc_reg = db.session.query(AuditoriaDeclaracao.id).filter(
            AuditoriaDeclaracao.declaracao_id == decl.id,
            AuditoriaDeclaracao.motivo == "enviado_drh"
        ).first()
        if decl.tipo == "positiva" and not enc_reg:
            flash(
                "Positiva ainda não foi encaminhada pela chefia. Sem validação.", "alert-warning")
            return redirect(url_for(
                "acumulo.recebimento",
                ano=request.args.get("ano"), status=request.args.get("status"),
                q=request.args.get("q"), obm_id=request.args.get("obm_id"),
                page=request.args.get("page"), per_page=request.args.get("per_page"),
            ))

    de = decl.status
    if novo == "enviar_drh":
        novo_status = "pendente"
        motivo = "enviado_drh"
    else:
        novo_status = novo
        motivo = (request.form.get("motivo") or None)

    # >>> NOVO: registra quem recebeu (apenas se ainda não foi registrado)
    if decl.recebido_por_user_id is None:
        decl.recebido_por_user_id = current_user.id
        decl.recebido_em = datetime.utcnow()

    # atualiza status
    decl.status = novo_status

    # auditoria
    db.session.add(AuditoriaDeclaracao(
        declaracao_id=decl.id,
        de_status=de, para_status=novo_status,
        motivo=motivo,
        alterado_por_user_id=current_user.id
    ))

    try:
        db.session.commit()
        flash("Status atualizado.", "alert-success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao atualizar status: {e}", "alert-danger")

    return redirect(url_for(
        "acumulo.recebimento",
        ano=request.args.get("ano"),
        status=request.args.get("status"),
        q=request.args.get("q"),
        obm_id=request.args.get("obm_id"),
        page=request.args.get("page"),
        per_page=request.args.get("per_page"),
    ))


@bp_acumulo.route("/arquivo/<int:decl_id>", methods=["GET"])
@login_required
@require_analise_vinculo
# @checar_ocupacao('DRH', 'DIRETOR DRH', 'SUPER USER')
def arquivo(decl_id):
    decl = db.session.get(DeclaracaoAcumulo, decl_id)
    if not decl or not decl.arquivo_declaracao:
        abort(404)
    return redirect(b2_presigned_get(decl.arquivo_declaracao, expires_seconds=600))


@bp_acumulo.route("/recebimento/export", methods=["GET"])
@login_required
@require_analise_vinculo
def recebimento_export():
    def fmt_dt(dt):
        return dt.strftime("%d/%m/%Y %H:%M") if dt else ""

    def fmt_d(d):
        return d.strftime("%d/%m/%Y") if d else ""

    def fmt_t(t):
        return t.strftime("%H:%M") if t else ""

    # mapeamentos legíveis
    TIPO_PT = {
        "positiva": "Positiva (com vínculo)",
        "negativa": "Negativa (sem vínculo)"
    }
    STATUS_PT = {"pendente": "Pendente",
                 "validado": "Validado", "inconforme": "Inconforme"}
    MEIO_PT = {"digital": "Digital", "presencial": "Presencial"}
    EMPREG_TIPO_PT = {"publico": "Público", "privado": "Privado",
                      "cooperativa": "Cooperativa", "profissional_liberal": "Profissional liberal"}
    NATUREZA_PT = {"efetivo": "Efetivo", "contratado": "Contratado",
                   "prestacao_servicos": "Prestação de serviços", "profissional_liberal": "Profissional liberal"}
    JORNADA_PT = {"escala": "Escala", "expediente": "Expediente"}

    def label(mapper, value):
        if value is None:
            return ""
        return mapper.get(str(value).lower(), str(value))

    ano = request.args.get("ano", type=int) or datetime.now().year
    status = (request.args.get("status") or "todos").lower()
    q = (request.args.get("q") or "").strip()
    obm_id = request.args.get("obm_id", type=int)

    # apelidos de modelos
    D, M, PG, U = DeclaracaoAcumulo, Militar, PostoGrad, User
    VE = VinculoExterno

    # coluna nome do usuário que recebeu
    user_name_col = functions.coalesce(U.nome)

    # ---------------------------
    # BASE (filtros comuns)
    # ---------------------------
    base_qry = (
        db.session.query(
            D,
            M,
            PG.sigla.label("pg_sigla"),
            user_name_col.label("recebido_por_nome"),
        )
        .join(M, M.id == D.militar_id)
        .outerjoin(PG, PG.id == M.posto_grad_id)
        .outerjoin(U, U.id == D.recebido_por_user_id)
        .filter(D.ano_referencia == ano)
    )

    if obm_id:
        base_qry = base_qry.filter(exists().where(and_(
            MilitarObmFuncao.militar_id == D.militar_id,
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
        )))

    if status in {"pendente", "validado", "inconforme"}:
        base_qry = base_qry.filter(D.status == status)

    if q:
        ilike = f"%{q}%"
        base_qry = base_qry.filter(or_(
            M.nome_completo.ilike(ilike),
            M.matricula.ilike(ilike),
            PG.sigla.ilike(ilike),
        ))

    # ---------------------------
    # 1) RESUMO (com selectinload p/ evitar N+1 no len(vinculos))
    # ---------------------------
    resumo_rows = (
        base_qry
        # carrega todos vínculos das declarações em 1 SELECT
        .options(selectinload(D.vinculos))
        .order_by(D.data_entrega.desc(), D.id.desc())
        .all()
    )

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = f"Resumo {ano}"
    ws_resumo.append([
        "Ano", "Status", "Tipo", "Meio", "Data de entrega",
        "Recebido por (nome)", "Recebido em",
        "Militar (nome completo)", "Matrícula", "Posto/Grad.", "Qtd vínculos", "Observações",
        "Arquivo modelo (URL/caminho)", "Arquivo órgão (URL/caminho)"
    ])

    for d, m, pg_sigla, recebido_por_nome in resumo_rows:
        ws_resumo.append([
            d.ano_referencia,
            label(STATUS_PT, d.status),
            label(TIPO_PT, d.tipo),
            label(MEIO_PT, d.meio_entrega),
            fmt_dt(d.data_entrega),
            recebido_por_nome or "",
            fmt_dt(d.recebido_em),
            m.nome_completo,
            m.matricula,
            pg_sigla or "",
            len(d.vinculos),  # NÃO vai gerar N+1 graças ao selectinload
            d.observacoes or "",
            d.arquivo_declaracao or "",
            getattr(d, "arquivo_declaracao_orgao", "") or "",
        ])

    # ---------------------------
    # 2) VÍNCULOS das POSITIVAS (sem session.get no loop)
    # ---------------------------
    vinc_qry = (
        db.session.query(
            D.id.label("decl_id"),
            D.status, D.meio_entrega, D.data_entrega,
            user_name_col.label("recebido_por_nome"), D.recebido_em,
            M.nome_completo, M.matricula, PG.sigla.label("pg_sigla"),
            VE.empregador_nome, VE.empregador_tipo, VE.empregador_doc,
            VE.natureza_vinculo, VE.cargo_funcao, VE.jornada_trabalho,
            VE.carga_horaria_semanal, VE.horario_inicio, VE.horario_fim,
            VE.data_inicio,
            D.arquivo_declaracao,                 # << já vem na mesma linha
            D.arquivo_declaracao_orgao,           # << idem
        )
        .join(M, M.id == D.militar_id)
        .outerjoin(PG, PG.id == M.posto_grad_id)
        .outerjoin(U, U.id == D.recebido_por_user_id)
        .join(VE, VE.declaracao_id == D.id)
        .filter(D.ano_referencia == ano, D.tipo == 'positiva')
    )

    if obm_id:
        vinc_qry = vinc_qry.filter(exists().where(and_(
            MilitarObmFuncao.militar_id == D.militar_id,
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
        )))

    if status in {"pendente", "validado", "inconforme"}:
        vinc_qry = vinc_qry.filter(D.status == status)

    if q:
        ilike = f"%{q}%"
        vinc_qry = vinc_qry.filter(or_(
            M.nome_completo.ilike(ilike),
            M.matricula.ilike(ilike),
            PG.sigla.ilike(ilike),
            VE.empregador_nome.ilike(ilike),
            VE.empregador_doc.ilike(ilike),
            VE.cargo_funcao.ilike(ilike),
        ))

    vinc_rows = vinc_qry.order_by(D.data_entrega.desc(), D.id.desc()).all()

    ws_vinc = wb.create_sheet("Vínculos (positivas)")
    ws_vinc.append([
        "Declaração ID", "Status", "Meio", "Data de entrega",
        "Recebido por (nome)", "Recebido em",
        "Militar (nome completo)", "Matrícula", "Posto/Grad.",
        "Empregador", "Tipo do empregador", "CPF/CNPJ",
        "Natureza do vínculo", "Cargo/Função", "Jornada",
        "Carga semanal (h)", "Entrada", "Saída", "Início do vínculo",
        "Arquivo modelo (URL/caminho)", "Arquivo órgão (URL/caminho)"
    ])
    for r in vinc_rows:
        ws_vinc.append([
            r.decl_id,
            label(STATUS_PT, r.status),
            label(MEIO_PT, r.meio_entrega),
            fmt_dt(r.data_entrega),
            r.recebido_por_nome or "",
            fmt_dt(r.recebido_em),
            r.nome_completo,
            r.matricula,
            (r.pg_sigla or ""),
            r.empregador_nome,
            label(EMPREG_TIPO_PT, r.empregador_tipo),
            (r.empregador_doc or ""),
            label(NATUREZA_PT, r.natureza_vinculo),
            r.cargo_funcao,
            label(JORNADA_PT, r.jornada_trabalho),
            r.carga_horaria_semanal,
            fmt_t(r.horario_inicio),
            fmt_t(r.horario_fim),
            fmt_d(r.data_inicio),
            r.arquivo_declaracao or "",
            (r.arquivo_declaracao_orgao or ""),
        ])

    # ---------------------------
    # 3) DECLARAÇÕES NEGATIVAS
    # ---------------------------
    neg_qry = (
        db.session.query(
            D.id.label("decl_id"),
            D.status, D.meio_entrega, D.data_entrega,
            user_name_col.label("recebido_por_nome"), D.recebido_em,
            D.observacoes, D.arquivo_declaracao,
            M.nome_completo, M.matricula, PG.sigla.label("pg_sigla"),
        )
        .join(M, M.id == D.militar_id)
        .outerjoin(PG, PG.id == M.posto_grad_id)
        .outerjoin(U, U.id == D.recebido_por_user_id)
        .filter(D.ano_referencia == ano, D.tipo == 'negativa')
    )

    if obm_id:
        neg_qry = neg_qry.filter(exists().where(and_(
            MilitarObmFuncao.militar_id == D.militar_id,
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
        )))

    if status in {"pendente", "validado", "inconforme"}:
        neg_qry = neg_qry.filter(D.status == status)

    if q:
        ilike = f"%{q}%"
        neg_qry = neg_qry.filter(or_(
            M.nome_completo.ilike(ilike),
            M.matricula.ilike(ilike),
            PG.sigla.ilike(ilike),
            D.observacoes.ilike(ilike),
        ))

    neg_rows = neg_qry.order_by(D.data_entrega.desc(), D.id.desc()).all()

    ws_neg = wb.create_sheet("Declarações negativas")
    ws_neg.append([
        "Declaração ID", "Status", "Meio", "Data de entrega",
        "Recebido por (nome)", "Recebido em",
        "Militar (nome completo)", "Matrícula", "Posto/Grad.",
        "Observações",
        "Arquivo modelo (URL/caminho)"
    ])
    for r in neg_rows:
        ws_neg.append([
            r.decl_id,
            label(STATUS_PT, r.status),
            label(MEIO_PT, r.meio_entrega),
            fmt_dt(r.data_entrega),
            r.recebido_por_nome or "",
            fmt_dt(r.recebido_em),
            r.nome_completo,
            r.matricula,
            (r.pg_sigla or ""),
            r.observacoes or "",
            r.arquivo_declaracao or "",
        ])

    # ---------------------------
    # 4) NÃO ENVIARAM  (AGORA FORA DO LOOP, roda 1x só)
    # ---------------------------
    falt_qry = (
        db.session.query(
            M.id.label("militar_id"),
            M.nome_completo,
            M.matricula,
            PG.sigla.label("pg_sigla"),
        )
        .outerjoin(PG, PG.id == M.posto_grad_id)
        .filter(~exists().where(and_(
            DeclaracaoAcumulo.militar_id == M.id,
            DeclaracaoAcumulo.ano_referencia == ano
        )))
    )

    if obm_id:
        falt_qry = falt_qry.filter(exists().where(and_(
            MilitarObmFuncao.militar_id == M.id,
            MilitarObmFuncao.obm_id == obm_id,
            MilitarObmFuncao.data_fim.is_(None),
        )))

    if q:
        ilike = f"%{q}%"
        falt_qry = falt_qry.filter(or_(
            M.nome_completo.ilike(ilike),
            M.matricula.ilike(ilike),
            PG.sigla.ilike(ilike),
        ))

    faltantes = falt_qry.order_by(M.nome_completo.asc()).all()

    # OBMs ativas em lote
    falt_ids = [r.militar_id for r in faltantes]
    obm_map = {}
    if falt_ids:
        MOF, O = MilitarObmFuncao, Obm  # ajuste se nomes forem outros
        obm_rows = (
            db.session.query(MOF.militar_id, O.sigla)
            .join(O, O.id == MOF.obm_id)
            .filter(
                MOF.militar_id.in_(falt_ids),
                MOF.data_fim.is_(None),
                *([MOF.obm_id == obm_id] if obm_id else [])
            )
            .order_by(O.sigla.asc())
            .all()
        )
        for mid, sigla in obm_rows:
            obm_map.setdefault(mid, []).append(sigla)

    ws_falt = wb.create_sheet(f"Não enviados {ano}")
    ws_falt.append(["Militar (nome completo)",
                   "Matrícula", "Posto/Grad.", "OBM(s)"])
    for r in faltantes:
        ws_falt.append([
            r.nome_completo,
            r.matricula,
            r.pg_sigla or "",
            ", ".join(obm_map.get(r.militar_id, [])),
        ])

    # ---------------------------
    # Retorno
    # ---------------------------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    bstatus = status if status != 'todos' else 'todos'
    suffix_obm = f"_obm{obm_id}" if obm_id else ""
    fname = f"declaracoes_{ano}_{bstatus}{suffix_obm}_com_vinculos_negativas_e_naoenviados.xlsx"

    resp = send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    try:
        resp.headers["Content-Length"] = str(bio.getbuffer().nbytes)
    except Exception:
        pass
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    return resp


@bp_acumulo.route("/detalhe/<int:decl_id>", methods=["GET"])
@login_required
@require_analise_vinculo
# @checar_ocupacao('DRH', 'DIRETOR DRH', 'DRH CHEFE', 'CHEFE DRH', 'SUPER USER', 'DIRETOR', 'CHEFE')
def detalhe(decl_id):
    D = DeclaracaoAcumulo
    decl = (db.session.query(D)
            .options(
                joinedload(D.militar).joinedload(Militar.posto_grad),
                selectinload(D.vinculos),
    ).get(decl_id))
    if not decl:
        abort(404)

    # OBMs ativas (pode haver mais de uma)
    siglas = (db.session.query(Obm.sigla)
              .join(MilitarObmFuncao, Obm.id == MilitarObmFuncao.obm_id)
              .filter(MilitarObmFuncao.militar_id == decl.militar_id,
                      MilitarObmFuncao.data_fim.is_(None))
              .all())
    obm_siglas = ", ".join(s for (s,) in siglas) or "-"

    url_arquivo_modelo = b2_presigned_get(
        decl.arquivo_declaracao, 600) if decl.arquivo_declaracao else None
    url_arquivo_orgao = b2_presigned_get(decl.arquivo_declaracao_orgao, 600) if getattr(
        decl, "arquivo_declaracao_orgao", None) else None

    pode_editar = _can_editar_declaracao(decl.militar_id)

    return render_template(
        "acumulo_detalhe.html",
        decl=decl,
        militar=decl.militar,
        posto_grad_sigla=getattr(decl.militar.posto_grad, "sigla", "-"),
        obm_siglas=obm_siglas,
        ano=decl.ano_referencia,
        url_arquivo_modelo=url_arquivo_modelo,
        url_arquivo_orgao=url_arquivo_orgao,
        pode_editar=pode_editar,
    )


@bp_acumulo.route("/modelo_docx/<int:militar_id>", methods=["GET"])
@login_required
def modelo_docx(militar_id):
    if not (_is_super_user() or _militar_permitido(militar_id)):
        flash("Sem permissão para este militar.", "alert-danger")
        return redirect(url_for("acumulo.lista"))

    # Você pode ler ano/tipo de querystring OU da sessão (se já guardou no prepara_geracao)
    ano = request.args.get("ano", type=int) or datetime.now().year
    tipo = (request.args.get("tipo") or "").lower()

    MOF = MilitarObmFuncao
    row = (
        db.session.query(Militar, PostoGrad.sigla.label(
            "pg_sigla"), Obm.sigla.label("obm_sigla"))
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .outerjoin(MOF, and_(MOF.militar_id == Militar.id, MOF.data_fim.is_(None)))
        .outerjoin(Obm, Obm.id == MOF.obm_id)
        .filter(Militar.id == militar_id)
        .first()
    )
    if not row:
        abort(404)
    militar, pg_sigla, obm_sigla = row

    emp_nome = (request.args.get("empregador_nome") or "").strip()
    emp_doc = _digits(request.args.get("empregador_doc") or "")
    natureza = (request.args.get("natureza_vinculo")
                or "").strip().replace("_", " ")
    jornada = (request.args.get("jornada_trabalho")
               or "").lower().replace("_", " ")
    cargo = (request.args.get("cargo_funcao") or "").strip()
    carga = (request.args.get("carga_horaria_semanal") or "").strip()
    hi = (request.args.get("horario_inicio") or "").strip()
    hf = (request.args.get("horario_fim") or "").strip()
    dinicio_raw = request.args.get("data_inicio") or ""
    dinicio = formatar_data_sem_zero(dinicio_raw) if dinicio_raw else ""

    mapping = {
        "posto_grad": pg_sigla or "-",
        "nome":       militar.nome_completo,
        "obm":        obm_sigla or "-",
        "ano":        str(ano),
        "x_vinculo_sim": "X" if tipo == "positiva" else "",
        "x_vinculo_nao": "X" if tipo == "negativa" else "",
        "empregador_nome":       emp_nome if tipo == "positiva" else "",
        "empregador_doc":        emp_doc if tipo == "positiva" else "",
        "x_escala": "X" if (tipo == "positiva" and jornada == "escala") else "",
        "x_expediente": "X" if (tipo == "positiva" and jornada == "expediente") else "",
        "natureza_vinculo":      natureza if tipo == "positiva" else "",
        "cargo_funcao":          cargo if tipo == "positiva" else "",
        "carga_horaria_semanal": carga if tipo == "positiva" else "",
        "horario":               (f"{hi} – {hf}" if (hi and hf) else "") if tipo == "positiva" else "",
        "data_inicio":           dinicio if tipo == "positiva" else "",
        "data_atual":            datetime.today().strftime("%d/%m/%Y"),
    }

    buf = render_docx_from_template(
        "src/template/declaracao_vinculo.docx", mapping)
    filename = f"Declaracao_{(militar.nome_guerra or militar.nome_completo).replace(' ', '_')}_{ano}.docx"

    # >>> cabeçalhos extras p/ iOS
    resp = send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    # Garante Content-Length
    try:
        resp.headers["Content-Length"] = str(buf.getbuffer().nbytes)
    except Exception:
        pass
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    return resp


@bp_acumulo.route("/modelo_docx_page/<int:militar_id>")
@login_required
def modelo_docx_page(militar_id):
    from urllib.parse import urlencode

    if not (_is_super_user() or _militar_permitido(militar_id)):
        flash("Sem permissão para este militar.", "alert-danger")
        return redirect(url_for("acumulo.lista"))

    qs = request.query_string.decode("utf-8")
    download_url = url_for("acumulo.modelo_docx",
                           militar_id=militar_id, _external=False)
    if qs:
        download_url = f"{download_url}?{qs}"

    return_to = request.args.get("return_to") or ""

    html = f"""
        <!doctype html>
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <title>Preparando download…</title>
        <style>
        body {{ font-family: system-ui,-apple-system,Segoe UI,Roboto; padding:16px; color:#333 }}
        .muted {{ color:#666; font-size:.9rem }}
        .hide {{ display:none }}
        </style>
        <p id="status">Preparando o download…</p>
        <p class="muted">Se não iniciar em alguns segundos, esta página tentará abrir o arquivo diretamente.</p>

        <script>
        (async function() {{
        const downloadUrl = {download_url!r};
        const returnTo = {return_to!r};
        const statusEl = document.getElementById('status');

        // Tenta baixar via fetch -> Blob -> <a download>, que é o modo mais estável no iOS.
        try {{
            statusEl.textContent = "Gerando arquivo…";
            const res = await fetch(downloadUrl, {{
            // mesma origem; envia cookies de sessão
            credentials: 'same-origin',
            cache: 'no-store'
            }});
            if (!res.ok) throw new Error("HTTP " + res.status);

            // tenta extrair o filename do Content-Disposition
            let filename = "Declaracao.docx";
            const cd = res.headers.get('Content-Disposition') || "";
            const m = cd.match(/filename\\*?=(?:UTF-8''|")?([^\";]+)/i);
            if (m && m[1]) {{
            try {{
                filename = decodeURIComponent(m[1].replace(/\\+/g, '%20')).replace(/\"/g, '');
            }} catch(_e) {{
                filename = m[1].replace(/\"/g, '');
            }}
            }}

            statusEl.textContent = "Iniciando download…";

            const blob = await res.blob();
            const url = URL.createObjectURL(blob);

            // cria link invisível e clica
            const a = document.createElement('a');
            a.href = url;
            a.download = filename || "Declaracao.docx";
            a.style.display = 'none';
            document.body.appendChild(a);
            a.click();

            // dá tempo do iOS criar o arquivo antes de fechar/navegar
            setTimeout(() => {{
            try {{ URL.revokeObjectURL(url); }} catch(_e) {{}}
            if (window.opener && !window.opener.closed) {{
                window.close();
            }} else if (returnTo) {{
                location.replace(returnTo);
            }}
            }}, 1200);

        }} catch (err) {{
            // Fallback: abre direto a URL do arquivo (Safari pode mostrar o preview nativo do DOCX)
            try {{
            statusEl.textContent = "Abrindo arquivo…";
            window.location.replace(downloadUrl);

            // Depois tenta fechar/voltar (se for popup, fecha; se não, volta p/ returnTo)
            setTimeout(() => {{
                if (window.opener && !window.opener.closed) {{
                window.close();
                }} else if (returnTo) {{
                location.replace(returnTo);
                }}
            }}, 2000);
            }} catch(_e) {{
            statusEl.textContent = "Não foi possível iniciar o download automaticamente.";
            }}
        }}
        }})();
        </script>
        """
    return html


@bp_acumulo.route("/minhas", methods=["GET"])
@login_required
def minhas_declaracoes():
    # usuário comum -> pega o próprio militar
    if current_user.funcao_user_id == 12:
        militar = get_militar_por_user(current_user)
        if not militar:
            flash("Não foi possível localizar seus dados de militar.", "danger")
            return redirect(url_for("home_atualizacao"))
        militar_id = militar.id
    else:
        # perfis DRH etc podem ver a própria também
        militar = get_militar_por_user(current_user) or None
        militar_id = getattr(militar, "id", None)

    ano = request.args.get("ano", type=int) or datetime.now().year

    q_base = db.session.query(DeclaracaoAcumulo).filter(
        DeclaracaoAcumulo.militar_id == militar_id,
        DeclaracaoAcumulo.ano_referencia == ano
    )
    minhas = q_base.order_by(DeclaracaoAcumulo.created_at.desc()).all()

    kpi_total = len(minhas)
    kpi_pend = sum(1 for d in minhas if (d.status or "").lower() == "pendente")
    kpi_val = sum(1 for d in minhas if (d.status or "").lower() == "validado")
    kpi_inc = sum(1 for d in minhas if (
        d.status or "").lower() == "inconforme")

    # 🔴 NOVO: existe pendente no ano?
    decl_pendente = (
        db.session.query(DeclaracaoAcumulo.id)
        .filter(
            DeclaracaoAcumulo.militar_id == militar_id,
            DeclaracaoAcumulo.ano_referencia == ano,
            DeclaracaoAcumulo.status == 'pendente',
        )
        .order_by(DeclaracaoAcumulo.id.desc())
        .first()
    )
    tem_pendente = bool(decl_pendente)
    decl_pendente_id = decl_pendente[0] if decl_pendente else None

    # Rascunho (único por militar/ano)
    # Ajuste o modelo/campo conforme sua implementação de rascunhos
    rasc = (db.session.query(DraftDeclaracaoAcumulo)
            .filter(DraftDeclaracaoAcumulo.militar_id == militar_id,
                    DraftDeclaracaoAcumulo.ano_referencia == ano)
            .first())

    # Para exibir OBM/PG no topo
    MOF = MilitarObmFuncao
    row = (
        db.session.query(Militar, PostoGrad.sigla.label("pg_sigla"))
        .outerjoin(PostoGrad, PostoGrad.id == Militar.posto_grad_id)
        .filter(Militar.id == militar_id)
        .first()
    )
    pg_sigla = row[1] if row else "-"

    # >>> NOVO: prazo
    prazo_limite = _prazo_envio_ate()  # sua função já existente
    prazo_fechado = (date.today() > prazo_limite)

    # >>> NOVO: já enviou algo no ano (qualquer status)
    ja_enviou_no_ano = db.session.query(DeclaracaoAcumulo.id).filter(
        DeclaracaoAcumulo.militar_id == militar_id,
        DeclaracaoAcumulo.ano_referencia == ano
    ).first() is not None

    tem_inconforme_no_ano = kpi_inc > 0

    return render_template(
        "acumulo_minhas.html",
        ano=ano,
        militar=militar,
        posto_grad_sigla=pg_sigla or "-",
        rascunho=rasc,
        declaracoes=minhas,
        kpi_decl_total=kpi_total,
        kpi_decl_pendentes=kpi_pend,
        kpi_decl_validadas=kpi_val,
        kpi_decl_inconformes=kpi_inc,
        tem_pendente=tem_pendente,
        decl_pendente_id=decl_pendente_id,
        # >>> NOVO: pro template
        prazo_fechado=prazo_fechado,
        prazo_limite=prazo_limite,
        ja_enviou_no_ano=ja_enviou_no_ano,
        tem_inconforme_no_ano=tem_inconforme_no_ano,
    )


@bp_acumulo.route("/rascunho/excluir/<int:militar_id>/<int:ano>", methods=["POST"])
@login_required
def excluir_rascunho(militar_id, ano):
    # segurança: usuário comum só pode excluir o próprio
    if current_user.funcao_user_id == 12:
        mil_user = get_militar_por_user(current_user)
        if not mil_user or mil_user.id != militar_id:
            return {"ok": False, "error": "Sem permissão para excluir este rascunho."}, 403

    try:
        r = (db.session.query(DraftDeclaracaoAcumulo)
             .filter(DraftDeclaracaoAcumulo.militar_id == militar_id,
                     DraftDeclaracaoAcumulo.ano == ano)
             .first())
        if not r:
            return {"ok": False, "error": "Rascunho não encontrado."}, 404

        db.session.delete(r)
        db.session.commit()
        return {"ok": True}, 200
    except Exception as e:
        db.session.rollback()
        return {"ok": False, "error": str(e)}, 500


@bp_acumulo.route("/novo/<pessoa_tipo>/<int:pessoa_id>", methods=["GET", "POST"])
@login_required
def novo_generico(pessoa_tipo, pessoa_id):
    if date.today() > _prazo_envio_ate() and not _usuario_ja_tem_declaracao(current_user.id, ANO_ATUAL):
        flash('O prazo para envio da declaração deste ano está encerrado.', 'warning')
        return redirect(url_for('home'))

    pessoa_tipo = (pessoa_tipo or "").lower().strip()
    if pessoa_tipo not in {"militar", "aluno"}:
        abort(404)

    if pessoa_tipo == "militar":
        m = Militar.query.get(pessoa_id)
        if not m:
            abort(404)
        # Se for usuário comum, reforça que é o próprio
        if current_user.funcao_user_id == 12:
            m_user = get_militar_por_user(current_user)
            if not m_user or m_user.id != m.id:
                flash(
                    "Você não tem permissão para abrir declaração para outro militar.", "danger")
                return redirect(url_for("home"))
        return redirect(url_for("acumulo.novo", militar_id=m.id))

    # aluno
    a = FichaAlunos.query.get(pessoa_id)
    if not a:
        abort(404)

    # garante um Militar correspondente (ou cria "shadow")
    m = ensure_militar_from_aluno(a, user=current_user, obm_padrao=26)
    if not m:
        flash("Não foi possível preparar seu cadastro para a declaração.", "danger")
        return redirect(url_for("home"))

    # usuário comum só abre pra si mesmo
    if current_user.funcao_user_id == 12:
        # opcional: amarrar por CPF do user
        pass

    return redirect(url_for("acumulo.novo", militar_id=m.id))

